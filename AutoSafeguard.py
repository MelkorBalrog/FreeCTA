#!/usr/bin/env python3
# SPDX-License-Identifier: GPL-3.0-or-later
#
# Copyright (C) 2025 Capek System Safety & Robotic Solutions
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""
===============================================================================
Risk & Assurance Gate Calculator for Autonomous Systems
===============================================================================

Overview of the Provided Risk Assessment Approach
-------------------------------
This tool is a semi-quantitative method designed to assess the safety assurance 
of an autonomous system’s subsystems. It produces an Prototype Assurance Level (PAL) (on a scale 
from 1 to 5) using qualitative labels that describe the required level of safety 
measures. For example, the scale is defined as:

   1 → Extra Low  
   2 → Low  
   3 → Moderate  
   4 → High  
   5 → High+  

The goal is to identify potential safety gaps and determine the extra assurance 
(i.e. additional testing, validation, design modifications) needed before a 
prototype is approved for public road trials.

Inputs – Confidence, Robustness, and Direct Assurance Metrics
-------------------------------
Each subsystem is evaluated using three main inputs (each rated from 1 to 5):

  1. **Confidence Level (CL):** Reflects the quality and extent of testing/validation.
  2. **Robustness Score (RS):** Reflects the strength of design safeguards and redundancy.
     (Different criteria are applied for system functions versus human tasks.)
  3. **Direct Assurance:** Pre-assessed assurance values derived from safety analyses.

For basic (leaf) nodes the provided ratings are used directly.

Computation Logic and Manual Calculation
-------------------------------

### 1. Deriving an Prototype Assurance Level (PAL) from Base Inputs
When only Confidence and Robustness values are provided, the tool “inverts” these 
inputs to yield a base Prototype Assurance Level (PAL). In this method, low confidence and low robustness 
result in a high assurance requirement (i.e. “High+”), while high confidence and high robustness 
yield a low assurance requirement (i.e. “Extra Low”).

**Assurance Matrix for Base Inputs (Qualitative Labels)**

|                           | **Confidence: Extra Low** | **Confidence: Low**   | **Confidence: Moderate** | **Confidence: High** | **Confidence: High+** |
|---------------------------|---------------------------|-----------------------|--------------------------|----------------------|-----------------------|
| **Robustness: Extra Low** | High+                     | High+                 | High                     | High                 | High                  |
| **Robustness: Low**       | High+                     | High+                 | High                     | Moderate             | Moderate              |
| **Robustness: Moderate**  | High                      | High                  | Moderate                 | Moderate             | Extra Low             |
| **Robustness: High**      | High                      | Moderate              | Moderate                 | Extra Low            | Extra Low             |
| **Robustness: High+**     | High                      | Moderate              | Extra Low                | Extra Low            | Extra Low             |

*Interpretation:*  
– Very poor testing and design (i.e. both “Extra Low”) lead to a “High+” assurance requirement.  
– Excellent testing and design (i.e. both “High+”) result in an “Extra Low” requirement.  
– Mixed values yield intermediate Prototype Assurance Levels (PAL).

---

### 2. Aggregating Prototype Assurance Levels (PAL) from Child Nodes

When a parent node aggregates Prototype Assurance Levels (PAL) from its children, the aggregation method 
depends on the logical gate connecting them:

#### For an **AND Gate**:
All components must be robust, so the overall assurance is determined by combining the child 
levels using a reliability-inspired approach. Use the following qualitative guideline:

**Aggregation Table for AND Gate (Qualitative Labels)**

|                         | **Child 2: Extra Low** | **Child 2: Low**   | **Child 2: Moderate** | **Child 2: High**   | **Child 2: High+**  |
|-------------------------|------------------------|--------------------|-----------------------|---------------------|---------------------|
| **Child 1: Extra Low**  | Extra Low              | Extra Low          | Low                   | High                | High+               |
| **Child 1: Low**        | Extra Low              | Low                | Moderate              | High                | High+               |
| **Child 1: Moderate**   | Low                    | Moderate           | High                  | High+               | High+               |
| **Child 1: High**       | High                   | High               | High+                 | High+               | High+               |
| **Child 1: High+**      | High+                  | High+              | High+                 | High+               | High+               |

*Interpretation:*  
– Combining two “High+” components remains “High+.”  
– If one component is significantly lower, the overall requirement shifts toward a higher assurance need.

#### For an **OR Gate**:
When alternative options are available, a simple average (by converting the qualitative levels 
to an ordered scale) is used. A strong alternative (e.g. “High+”) can partially offset a weaker one 
(e.g. “Low”).

---

### 3. Decomposing a Parent Prototype Assurance Level (PAL) into Child Targets

A parent node’s overall assurance requirement can be decomposed into target Prototype Assurance Levels (PAL) 
for its children. The following guidelines serve as a reference for common decompositions:

**Decomposition Guidelines**

- **Parent Assurance: High+**  
  – Option A: Both children target “High.”  
  – Option B: One child may target “High+” while the other targets “Extra Low” so that their combined effect meets the “High+” requirement.
  
- **Parent Assurance: High**  
  – Children should typically target between “Moderate” and “High.”

- **Parent Assurance: Moderate**  
  – Children should have targets in the range of “Low” to “Moderate.”

- **Parent Assurance: Low**  
  – Children should target “Extra Low” or “Low.”

- **Parent Assurance: Extra Low**  
  – Both children should be “Extra Low.”

These rules ensure that when children’s Prototype Assurance Levels (PAL) are aggregated (using the AND or OR rules), 
they “reconstruct” the parent’s overall requirement.

---

### 4. Adjusting Assurance Based on Severity

Severity reflects the potential impact of a subsystem’s failure. It is used to adjust the computed 
Prototype Assurance Level (PAL) as follows:

- **General Rule (for most nodes):**  
  **Final Prototype Assurance Level (PAL) = (Aggregated Child Assurance + Highest Parent Severity) ÷ 2**  
  A higher severity (indicating more catastrophic consequences) increases the overall assurance requirement.

- **For Vehicle Level Functions:**  
  The node’s own severity is used instead of the parent’s. An example adjustment formula is:  
  **Adjusted Assurance = (2 × Computed Assurance) – (Node’s Own Severity)**  
  This modification increases the Prototype Assurance Level (PAL) when the potential impact is high.

---

### Discretization Tables
The following tables map raw numeric inputs to discrete levels that are then translated into qualitative labels:

1) **Confidence Level**

   +-------+------------------------+-----------------------------------------------+
   | Level | Description            | Expert Criteria                               |
   +-------+------------------------+-----------------------------------------------+
   |   1   | Very poor confidence   | No testing or validation evidence.           |
   |   2   | Poor confidence        | Minimal testing; incomplete evidence.        |
   |   3   | Moderate confidence    | Some validation; moderate evidence.          |
   |   4   | High confidence        | Well-tested with redundant checks.           |
   |   5   | Excellent confidence   | Comprehensive testing & strong evidence.     |
   +-------+------------------------+-----------------------------------------------+

2) **Robustness (Function)**

   +-------+--------------------------+---------------------------------------------+
   | Score | Description             | Rationale (Safety Loading)                  |
   +-------+--------------------------+---------------------------------------------+
   |   1   | Very Poor Safety Load   | Minimal redundancy; fails to mitigate risks.|
   |   2   | Poor Safety Load        | Only basic safety measures.                 |
   |   3   | Moderate Safety Load    | Standard protection; moderate redundancy.   |
   |   4   | High Safety Load        | Strong redundancy & mitigations.            |
   |   5   | Excellent Safety Load   | Full redundancy & comprehensive measures.   |
   +-------+--------------------------+---------------------------------------------+

3) **Robustness (Human Task)**

   +-------+--------------------------+----------------------------------------------+
   | Level | Description            | Expert Criteria for a Human Task             |
   +-------+--------------------------+----------------------------------------------+
   |   1   | Very poor performance  | Minimal training; slow reaction times.       |
   |   2   | Poor performance       | Limited training; suboptimal responses.      |
   |   3   | Moderate performance   | Adequately trained; acceptable reactions.    |
   |   4   | High performance       | Very experienced; quick & sound decisions.   |
   |   5   | Excellent performance  | Expert-level with flawless performance.      |
   +-------+--------------------------+----------------------------------------------+

---

### Summary of Qualitative Assurance Labels

- **Extra Low:** Minimal assurance required (system is very safe).
- **Low:** Some assurance is required.
- **Moderate:** A moderate level of additional assurance is needed.
- **High:** Significant additional assurance is required.
- **High+:** Maximum assurance is required (system is highly unsafe without improvements).

---

### Additional Notes on the Calculation Process

- **Combining Direct Inputs:**  
  If any direct assurance inputs are provided, they are combined using logical gate rules:
  
  - **OR Gate:** Inputs are averaged.
  - **AND Gate:** Inputs are combined using a “complement product” approach.

- **Adjustment with Severity:**  
  The final Prototype Assurance Level (PAL) is adjusted by incorporating the severity (using the highest parent severity unless the node is a Vehicle Level Function, in which case its own severity is used).

- **Decomposition and Aggregation:**  
  The parent node’s assurance requirement can be decomposed into target Prototype Assurance Levels (PAL) for its children (see Decomposition Guidelines above), and child Prototype Assurance Levels (PAL) are aggregated (using the AND/OR rules) to reconstruct the parent’s overall requirement.

-------------------------------
References
----------
- Rausand, M., & Høyland, A. (2004). *System Reliability Theory: Models, Statistical Methods, and Applications.* Wiley-Interscience.

===============================================================================
"""

import re
import math
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from review_toolbox import (
    ReviewToolbox,
    ReviewData,
    ReviewParticipant,
    ReviewComment,
    ParticipantDialog,
    EmailConfigDialog,
    ReviewScopeDialog,
    UserSelectDialog,
    ReviewDocumentDialog,
    VersionCompareDialog,
)
from dataclasses import asdict
from mechanisms import DiagnosticMechanism, MechanismLibrary, ANNEX_D_MECHANISMS
import json
import csv
try:
    from openpyxl import load_workbook
except Exception:  # openpyxl may not be installed
    load_workbook = None
from drawing_helper import FTADrawingHelper, fta_drawing_helper
from risk_assessment import DERIVED_MATURITY_TABLE, ASSURANCE_AGGREGATION_AND, AND_DECOMPOSITION_TABLE, OR_DECOMPOSITION_TABLE, boolify, ADRiskAssessmentHelper
from models import (
    MissionProfile,
    ReliabilityComponent,
    ReliabilityAnalysis,
    HazopEntry,
    HaraEntry,
    HazopDoc,
    HaraDoc,
    QUALIFICATIONS,
    COMPONENT_ATTR_TEMPLATES,
    RELIABILITY_MODELS,
    ASIL_LEVEL_OPTIONS,
    ASIL_ORDER,
    ASIL_TARGETS,
    ASIL_TABLE,
    calc_asil,
    global_requirements,
)
from toolboxes import ReliabilityWindow, FI2TCWindow, HazopWindow, HaraWindow, TC2FIWindow
import copy
import tkinter.font as tkFont
from PIL import Image, ImageDraw, ImageFont, ImageTk
import os
import types
os.environ["GS_EXECUTABLE"] = r"C:\Program Files\gs\gs10.04.0\bin\gswin64c.exe"
import networkx as nx
import matplotlib.pyplot as plt
# Import ReportLab for PDF export.
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak, SimpleDocTemplate, Image as RLImage
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO, StringIO
from email.utils import make_msgid
import html
import datetime
import PIL.Image as PILImage
from reportlab.platypus import LongTable
from email.message import EmailMessage
import smtplib
import socket

styles = getSampleStyleSheet()  # Create the stylesheet.
preformatted_style = ParagraphStyle(name="Preformatted", fontName="Courier", fontSize=10)
styles.add(preformatted_style)

# Characters used to display pass/fail status in metrics labels.
CHECK_MARK = "\u2713"
CROSS_MARK = "\u2717"

# Target PMHF limits per ASIL level (events per hour)
PMHF_TARGETS = {
    "D": 1e-8,
    "C": 1e-7,
    "B": 1e-7,
    "A": 1e-6,
    "QM": 1.0,
}

##########################################
# VALID_SUBTYPES dictionary
##########################################
VALID_SUBTYPES = {
    "Confidence": ["Function", "Human Task"],
    "Robustness": ["Function", "Human Task"],
    "Maturity": ["Functionality"],
    "Rigor": ["Capability", "Safety Mechanism"],
    "Prototype Assurance Level (PAL)": ["Vehicle Level Function"]
}

##########################################
# Global Unique ID Counter for Nodes
##########################################
unique_node_id_counter = 1
import uuid

dynamic_recommendations = {
    1: {
        "Testing Requirements": (
            "Perform extensive scenario-based simulations covering normal driving, sensor failures, emergency braking, "
            "and boundary conditions. Conduct rigorous lab tests and closed-course trials to verify core ADS functions under ideal conditions. "
            "No public road tests are permitted until every core function is validated in a controlled prototype environment."
        ),
        "IFTD Responsibilities": (
            "A dedicated safety driver is in the vehicle at all times along with an engineer. The IFTD must be able to take immediate manual control "
            "when abnormal conditions are detected. Training focuses on achieving short reaction times and enhanced situational awareness through frequent emergency takeover drills."
        ),
        "Preventive Maintenance Actions": (
            "Conduct pre-trip and post-trip inspections on every run. Regularly calibrate, clean, and realign all sensors (cameras, radar, LiDAR). "
            "Maintain a detailed log and perform daily component checks to promptly address any anomalies before further testing."
        ),
        "Relevant AVSC Guidelines": (
            "Adhere to AVSC Best Practice for In-Vehicle clone Test Driver Selection, Training, and Oversight (AVSC00001-2019) and SAE J3018 guidelines, "
            "with extra emphasis on ensuring the IFTD can safely intervene in a prototype environment."
        ),
        "Extra Recommendations": {
            "steering": (
                "Include operational tests that simulate sudden, unintended steering inputs and verify that dynamic steering limiters are active. "
                "Ensure that the IFTD can promptly override any abnormal steering commands."
            ),
            "lateral": (
                "Design tests to simulate faulty lateral control (e.g., drifting or incorrect lane-keeping) and verify that any deviation is corrected within safe lateral boundaries. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required by the system’s rigor and child maturity."
            ),
            "longitudinal": (
                "Simulate sudden acceleration or deceleration events and verify that smooth speed transitions are maintained. "
                "Ensure that any unexpected longitudinal changes are managed safely by the IFTD."
            ),
            "braking": (
                "Simulate unintended or excessive braking events on a closed course and verify that the IFTD can quickly restore controlled braking and vehicle stability. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "deceleration": (
                "Develop test scenarios where deceleration is either too abrupt or delayed, and verify that the IFTD's intervention results in smooth, predictable slowdowns within defined limits."
            ),
            "acceleration": (
                "Test for unintended acceleration surges by simulating load changes and external disturbances. "
                "Verify that the IFTD can promptly override acceleration commands to maintain smooth, safe speed profiles."
            ),
            "park brake": (
                "Design controlled tests that trigger parking brake faults and verify that the system engages/disengages reliably. "
                "Ensure that the IFTD can safely manage the vehicle during such events."
            ),
            "parking brake": (
                "Conduct targeted tests simulating parking brake malfunctions, ensuring reliable engagement/disengagement and that the IFTD can intervene safely."
            ),
            "mode": (
                "Simulate mode indicator anomalies to verify that the IFTD receives clear, actionable alerts and can enforce proper system state transitions."
            ),
            "notification": (
                "Verify that the alert system responds accurately to simulated sensor or system errors in a controlled test environment. "
                "Ensure that alerts are displayed clearly via visual and auditory cues and are properly logged."
            ),
            "takeover": (
                "Simulate scenarios where the ADS disengages unexpectedly, requiring the IFTD to take over. "
                "Validate that the takeover mechanism enables the IFTD to quickly and safely assume manual control."
            ),
            "rollaway": (
                "Conduct basic simulation tests to verify that the system can detect a potential rollaway condition. "
                "Test the activation of emergency brakes and initial control protocols on a slight incline in a controlled laboratory environment."
            ),
            "control": (
                "Assess the IFTD’s basic manual override capability under simulated conditions. "
                "Ensure that, despite minimal training, the driver can momentarily assume control in a laboratory environment."
            )
        }
    },
    2: {
        "Testing Requirements": (
            "Initiate limited public-road tests under tightly controlled conditions (e.g., low-speed, daylight, good weather) within a constrained ODD. "
            "Employ advanced simulations—including fault injection, emergency braking, and scenario-based tests—alongside closed-course validations to verify safe operation."
        ),
        "IFTD Responsibilities": (
            "The safety driver (with a co-driver if necessary) continuously monitors the ADS and is ready to intervene immediately. "
            "Training drills focus on rapid manual intervention and maintaining situational awareness under varying test conditions."
        ),
        "Preventive Maintenance Actions": (
            "Implement both time-based and event-triggered inspections. Prior to each test, verify that sensor calibrations and system integrity meet safety standards. "
            "Document all findings comprehensively and address any anomalies immediately to support safe operation."
        ),
        "Relevant AVSC Guidelines": (
            "Follow AVSC Best Practice for Data Collection for ADS-DVs (AVSC00004-2020), comply with SAE J3018, and meet local regulatory standards, "
            "with a focus on enhancing IFTD control and training."
        ),
        "Extra Recommendations": {
            "steering": (
                "Design tests that simulate unexpected steering deviations and verify that the IFTD can safely override these inputs. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as per the required rigor and maturity of child elements."
            ),
            "lateral": (
                "Simulate faulty lateral control scenarios to ensure that any drift or deviation is corrected by the IFTD within safe lateral boundaries. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required by the system’s rigor and maturity of child elements."
            ),
            "longitudinal": (
                "Incorporate scenarios that simulate sudden acceleration or deceleration events. "
                "Verify that the emergency override system responds to maintain safe speed profiles."
            ),
            "braking": (
                "Include tests that simulate unintended or excessive braking events and verify that the IFTD can quickly re-establish controlled braking. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "deceleration": (
                "Develop test scenarios to confirm that deceleration remains controlled, even if it is slightly delayed or abrupt. "
                "Verify that basic emergency intervention protocols are activated in a controlled environment."
            ),
            "acceleration": (
                "Test for unintended acceleration surges by simulating moderate load changes and external disturbances. "
                "Confirm that the emergency override system responds to maintain safe speed profiles."
            ),
            "park brake": (
                "Conduct controlled tests to simulate parking brake faults and verify that basic safety protocols, such as system alerts and initial brake engagement, function properly."
            ),
            "parking brake": (
                "Conduct controlled tests to simulate parking brake malfunctions and ensure that the system engages protective measures reliably."
            ),
            "mode": (
                "Simulate mode indicator anomalies to verify that the IFTD receives clear alerts and can trigger preliminary system checks."
            ),
            "notification": (
                "Test the alert system under controlled conditions to verify prompt and clear notification of sensor or system errors."
            ),
            "takeover": (
                "Simulate scenarios where the ADS unexpectedly disengages to ensure that the IFTD can assume manual control quickly."
            ),
            "rollaway": (
                "Perform controlled closed-course tests simulating a rollaway event on a mild slope. "
                "Validate that the emergency braking system engages, the transmission shifts to neutral, and that driver alerts are issued promptly."
            ),
            "control": (
                "Verify that the IFTD can take control during simple, low-speed scenarios. "
                "Ensure that the manual override interface provides clear signals for intervention under these controlled conditions."
            )
        }
    },
    3: {
        "Testing Requirements": (
            "Expand testing into a broader ODD using high-fidelity simulations and extended on-road trials. "
            "Include scenarios such as higher speeds, nighttime driving, and light rain, along with targeted fault-injection tests that challenge the ADS and verify that the IFTD can promptly intervene."
        ),
        "IFTD Responsibilities": (
            "The safety driver remains onboard as a continuous clone while the ADS handles most of the route. "
            "Enhanced training emphasizes rapid manual takeover and precise interpretation of ADS signals, reinforced by regular simulator and on-track drills."
        ),
        "Preventive Maintenance Actions": (
            "Establish a formal maintenance schedule combining regular and event-based inspections supported by on-board diagnostics and predictive analytics. "
            "Preemptively address any component degradation to ensure that the IFTD’s ability to intervene is never compromised."
        ),
        "Relevant AVSC Guidelines": (
            "Utilize AVSC Best Practice for Metrics and Methods for Assessing Safety Performance and continuous monitoring principles. "
            "Ensure periodic IFTD re-training and adhere to ISO 26262/21448 for functional safety."
        ),
        "Extra Recommendations": {
            "steering": (
                "Simulate abnormal steering responses and verify that the IFTD can override these inputs safely. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as per the required rigor and the maturity of child elements."
            ),
            "lateral": (
                "Develop test scenarios that replicate lateral control failures and verify that the IFTD restores proper lateral stability within defined limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required by the system’s rigor and child maturity."
            ),
            "longitudinal": (
                "Design tests that simulate abrupt changes in speed and verify that manual override maintains smooth acceleration and deceleration within preset control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required."
            ),
            "braking": (
                "Include tests for inconsistent braking responses and evaluate how quickly and effectively the IFTD can re-establish controlled braking within safe limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "deceleration": (
                "Test deceleration behavior under fault conditions, ensuring that even with anomalies the deceleration remains predictable and controllable. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "acceleration": (
                "Include scenarios that trigger unexpected acceleration surges and verify that the IFTD can promptly intervene to restore safe speed levels. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "park brake": (
                "Design tests that simulate parking brake faults and assess the IFTD's ability to safely manage the vehicle until normal operation is restored. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "parking brake": (
                "Design tests that simulate parking brake faults and assess the IFTD's ability to safely manage the vehicle until normal operation is restored. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "mode": (
                "Simulate mode indicator errors and confirm that the IFTD is alerted to enforce correct system state transitions. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "notification": (
                "During extended on-road trials, verify that the notification system integrates with live sensor data to produce real alerts. "
                "Ensure that alerts are clearly presented—via both visual and auditory channels—and that the IFTD can respond promptly under dynamic conditions. "
                "Also, conduct performance studies to assess the IFTD's alert perception, reaction time, and controllability."
            ),
            "takeover": (
                "Develop complex scenarios that require the IFTD to take over from the ADS during fault conditions. "
                "Monitor response time and the system’s ability to safely transition to manual control, and conduct detailed post-event analyses."
            ),
            "rollaway": (
                "Simulate a truck rollaway scenario on a declining grade under controlled conditions. "
                "Verify that the vehicle's emergency braking, transmission neutralization, and electronic stability controls engage promptly to prevent uncontrolled movement. "
                "Ensure that tests include driver override procedures and proper system logging for subsequent analysis."
            ),
            "control": (
                "Confirm that the IFTD consistently demonstrates the ability to assume control during operational tests. "
                "The manual override interface should be intuitive, providing timely feedback and clear signals to the driver."
            )
        }
    },
    4: {
        "Testing Requirements": (
            "Conduct pilot tests in a quasi-commercial setting on intended routes under realistic conditions. "
            "Test the ADS across its full ODD—including boundary scenarios—using advanced simulations and on-road trials designed to safely challenge system limits."
        ),
        "IFTD Responsibilities": (
            "An IFTD is onboard at all times as the ultimate safety net. Although interventions become less frequent, the driver must remain vigilant and undergo regular drills "
            "and attention tests to ensure sustained manual control readiness under operational conditions."
        ),
        "Preventive Maintenance Actions": (
            "Integrate comprehensive preventive maintenance into the test cycle. Perform extensive pre-run system checks (HD map verification, sensor cleaning, redundant system tests) "
            "to confirm that all components operate reliably, thereby supporting uninterrupted IFTD oversight."
        ),
        "Relevant AVSC Guidelines": (
            "Implement AVSC Best Practice for First Responder Interactions and adopt a standardized Safety Inspection Framework. "
            "Ensure continuous monitoring and compliance with regulatory requirements (e.g., FMCSA, state DOT), with extra emphasis on IFTD training and rapid intervention procedures."
        ),
        "Extra Recommendations": {
            "steering": (
                "Include operational tests verifying that unexpected steering deviations are safely managed by the IFTD, with control limits enforced. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required."
            ),
            "lateral": (
                "Simulate faulty lateral control scenarios and verify that any drift or deviation is corrected by the IFTD within safe lateral boundaries. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required by the system’s rigor and maturity of child elements."
            ),
            "longitudinal": (
                "Design tests that simulate abrupt or erratic longitudinal events. Verify that manual override smoothly restores safe acceleration and deceleration within predefined control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "braking": (
                "Conduct tests simulating unintended or excessive braking events and verify that the IFTD can rapidly re-establish controlled braking with predictable deceleration. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "deceleration": (
                "Include scenarios that ensure deceleration remains smooth and within safe limits even under abnormal conditions, with timely IFTD intervention if needed. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "acceleration": (
                "Test for unexpected acceleration surges and verify that the IFTD can safely override to restore smooth acceleration within acceptable limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "park brake": (
                "Perform targeted tests on parking brake engagement under fault conditions to verify reliable operation and that the IFTD can safely manage the vehicle within defined control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "parking brake": (
                "Perform targeted tests on parking brake engagement under fault conditions to verify reliable operation and that the IFTD can safely manage the vehicle within defined control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "mode": (
                "Simulate mode indicator anomalies and verify that the IFTD receives clear, actionable alerts to enforce correct system state transitions, "
                "while ensuring control limits are maintained. These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "notification": (
                "During pilot operations, validate that the notification system generates real-time alerts in response to actual sensor malfunctions or system deviations. "
                "Ensure that alerts are unambiguous and use multiple modalities (visual and auditory) to prompt immediate manual intervention if required. "
                "Additionally, perform targeted studies to measure the IFTD's alert perception, reaction time, and controllability, and apply these insights to refine both the alert mechanisms and driver training protocols."
            ),
            "takeover": (
                "Conduct pilot tests that incorporate controlled takeover scenarios. Assess the responsiveness, accuracy, and smoothness of manual intervention when the ADS disengages. "
                "Measure key performance metrics such as takeover speed and transition stability, and use these data to further refine both the takeover mechanism and IFTD training protocols."
            ),
            "rollaway": (
                "Under more demanding conditions, simulate a truck rollaway on a steeper decline with higher speeds. "
                "Verify that advanced emergency protocols—including enhanced trailer locking mechanisms and improved driver alert systems—are activated. "
                "Ensure that the vehicle's redundant braking and stability control systems work in tandem, and that the system facilitates timely driver override if required."
            ),
            "control": (
                "Ensure that the IFTD reliably assumes control in complex scenarios. "
                "The system should deliver clear override signals, and the driver must demonstrate enhanced situational awareness and rapid response under challenging conditions."
            )
        }
    },
    5: {
        "Testing Requirements": (
            "Subject the ADS to rigorous edge-case validations and continuous simulation exercises that safely challenge the system across its entire ODD. "
            "Design test scenarios to deliberately trigger abnormal conditions so that control limits are enforced and the IFTD remains fully prepared to intervene."
        ),
        "IFTD Responsibilities": (
            "Even at the highest automation level, an IFTD is always onboard as a failsafe. Their role is primarily supervisory, yet they undergo continuous, intensive training "
            "and periodic drills—including attention-enhancing measures such as periodic system alerts—to ensure immediate manual control if any sensor or system fault occurs."
        ),
        "Preventive Maintenance Actions": (
            "Maintain standard commercial fleet maintenance protocols with automated self-checks and condition-based preventive measures. "
            "Conduct frequent system health verifications—including sensor recalibration, hardware diagnostics, and software integrity tests—to ensure that control limits are consistently maintained and the IFTD oversight remains uncompromised."
        ),
        "Relevant AVSC Guidelines": (
            "Implement all applicable AVSC best practices—including continuous monitoring, first responder protocols, and transparency standards. "
            "Adhere to industry certifications (e.g., ANSI/UL 4600, ISO 26262/21448) while emphasizing rigorous IFTD training, enhanced system controllability, and rapid manual intervention within defined control limits."
        ),
        "Extra Recommendations": {
            "steering": (
                "Include operational tests verifying that dynamic steering limiters are active and that the IFTD can safely intervene when steering inputs exceed defined control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required."
            ),
            "lateral": (
                "Design tests to simulate faulty lateral control and verify that any drift or deviation is corrected by the IFTD within safe lateral boundaries. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required by the system’s rigor and the maturity of all child elements."
            ),
            "longitudinal": (
                "Develop scenarios that test the smooth manual override of acceleration and deceleration controls, ensuring that any unexpected longitudinal changes are managed within preset control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing as required."
            ),
            "braking": (
                "Include test cases for unintended or excessive braking, confirming that the IFTD can immediately assume control to restore safe braking within defined limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "deceleration": (
                "Verify through operational tests that deceleration remains smooth and controlled even when system signals are abnormal, ensuring timely IFTD intervention within safe deceleration limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "acceleration": (
                "Include scenarios to detect and safely manage any unintended acceleration surges, ensuring the IFTD can quickly override to maintain smooth speed transitions within acceptable limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "park brake": (
                "Perform targeted tests on parking brake engagement under fault conditions to ensure reliable performance and that the IFTD can safely manage the vehicle within defined control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "parking brake": (
                "Perform targeted tests on parking brake engagement under fault conditions to ensure reliable performance and that the IFTD can safely manage the vehicle within defined control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "mode": (
                "Simulate mode indicator anomalies and verify that the IFTD receives clear, actionable alerts to enforce correct system state transitions while maintaining operational control limits. "
                "These tests shall include simulations, closed-course testing, and silent mode testing."
            ),
            "notification": (
                "Under near-commercial conditions, monitor the notification alert system over extended periods to ensure that alerts are consistently delivered in real-time. "
                "Validate that alerts are unambiguous and use multiple modalities (visual and auditory) to prompt immediate manual intervention if required. "
                "Furthermore, conduct comprehensive studies to quantify the IFTD’s alert perception time, reaction time, and controllability, and apply these insights to refine both the alert mechanisms and driver training programs."
            ),
            "takeover": (
                "Even in near-commercial conditions, periodically simulate takeover events to ensure that the ADS remains fail-safe and that the IFTD can effectively intervene if required. "
                "Measure key performance metrics—such as takeover speed, accuracy, and smoothness of transition—and use these data to continuously refine the takeover mechanisms and improve IFTD training."
            ),
            "rollaway": (
                "Conduct exhaustive tests under worst-case rollaway scenarios, such as extended, steep grades combined with sensor or system faults. "
                "Ensure that all redundant systems, including emergency braking, transmission neutralization, and electronic stability controls, engage seamlessly. "
                "Validate the effectiveness of automated driver override protocols and comprehensive system logging to support post-incident analysis."
            ),
            "control": (
                "Validate that the IFTD can seamlessly assume complete control even under worst-case conditions. "
                "Extensive driver training, robust override interfaces, and redundant manual control mechanisms must be confirmed during rigorous testing."
            )
        }
    }
}

AD_RiskAssessment_Helper = ADRiskAssessmentHelper()

##########################################
# Edit Dialog 
##########################################
class EditNodeDialog(simpledialog.Dialog):
    def __init__(self, parent, node, app):
        self.node = node
        self.app = app
        super().__init__(parent, title="Edit Node")

    def body(self, master):
        dialog_font = tkFont.Font(family="Arial", size=10)
        ttk.Label(master, text="Node ID:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.id_entry = tk.Entry(master, font=dialog_font, state="disabled")
        self.id_entry.insert(0, f"Node {self.node.unique_id}")
        self.id_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(master, text="User Name:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.user_name_entry = tk.Entry(master, font=dialog_font)
        self.user_name_entry.insert(0, self.node.user_name)
        self.user_name_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(master, text="Description:").grid(row=2, column=0, padx=5, pady=5, sticky="ne")
        self.desc_text = tk.Text(master, width=40, height=3, font=dialog_font, wrap="word")
        self.desc_text.insert("1.0", self.node.description)
        self.desc_text.grid(row=2, column=1, padx=5, pady=5)
        self.desc_text.bind("<Return>", self.on_enter_pressed)

        ttk.Label(master, text="\nRationale:").grid(row=3, column=0, padx=5, pady=5, sticky="ne")
        self.rationale_text = tk.Text(master, width=40, height=3, font=dialog_font, wrap="word")
        self.rationale_text.insert("1.0", self.node.rationale)
        self.rationale_text.grid(row=3, column=1, padx=5, pady=5)
        self.rationale_text.bind("<Return>", self.on_enter_pressed)

        row_next = 4
        if self.node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            ttk.Label(master, text="Value (1-5):").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
            self.value_combo = ttk.Combobox(master, values=["1", "2", "3", "4", "5"],
                                            state="readonly", width=5, font=dialog_font)
            current_val = self.node.quant_value if self.node.quant_value is not None else 1
            self.value_combo.set(str(int(current_val)))
            self.value_combo.grid(row=row_next, column=1, padx=5, pady=5)
            row_next += 1

            # NEW: Safety Requirements Section for base nodes.
            # Ensure the node has the attribute.
            if not hasattr(self.node, "safety_requirements"):
                self.node.safety_requirements = []
            ttk.Label(master, text="Safety Requirements:").grid(row=row_next, column=0, padx=5, pady=5, sticky="ne")
            self.safety_req_frame = ttk.Frame(master)
            self.safety_req_frame.grid(row=row_next, column=1, padx=5, pady=5, sticky="w")
            row_next += 1

            # Create a listbox to display safety requirements.
            self.safety_req_listbox = tk.Listbox(self.safety_req_frame, height=4, width=50)
            self.safety_req_listbox.grid(row=0, column=0, columnspan=3, sticky="w")
            # Populate listbox with existing requirements.
            for req in self.node.safety_requirements:
                self.safety_req_listbox.insert(tk.END, f"[{req['id']}] [{req['req_type']}] [{req.get('asil','')}] {req['text']}")

            # Buttons for Add, Edit, and Delete.
            self.add_req_button = ttk.Button(self.safety_req_frame, text="Add New", command=self.add_safety_requirement)
            self.add_req_button.grid(row=1, column=0, padx=2, pady=2)
            self.edit_req_button = ttk.Button(self.safety_req_frame, text="Edit", command=self.edit_safety_requirement)
            self.edit_req_button.grid(row=1, column=1, padx=2, pady=2)
            self.delete_req_button = ttk.Button(self.safety_req_frame, text="Delete", command=self.delete_safety_requirement)
            self.delete_req_button.grid(row=1, column=2, padx=2, pady=2)
            self.add_existing_req_button = ttk.Button(self.safety_req_frame, text="Add Existing", command=self.add_existing_requirement)
            self.add_existing_req_button.grid(row=1, column=3, padx=2, pady=2)

        elif self.node.node_type.upper() == "BASIC EVENT":
            ttk.Label(master, text="Failure Probability:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
            self.prob_entry = tk.Entry(master, font=dialog_font)
            self.prob_entry.grid(row=row_next, column=1, padx=5, pady=5)
            row_next += 1

            ttk.Label(master, text="Represents FM:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
            modes = self.app.get_all_failure_modes()
            self.fm_map = {self.app.format_failure_mode_label(m): m.unique_id for m in modes}
            self.fm_var = tk.StringVar()
            current = ''
            if getattr(self.node, 'failure_mode_ref', None):
                n = self.app.find_node_by_id_all(self.node.failure_mode_ref)
                if n:
                    current = self.app.format_failure_mode_label(n)
            self.fm_var.set(current)
            self.fm_combo = ttk.Combobox(master, textvariable=self.fm_var, values=list(self.fm_map.keys()), state='readonly', width=40)
            self.fm_combo.grid(row=row_next, column=1, padx=5, pady=5, sticky='w')
            self.fm_combo.bind("<<ComboboxSelected>>", self.update_probability)
            row_next += 1


            ttk.Label(master, text="Probability Formula:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
            self.formula_var = tk.StringVar(value=getattr(self.node, 'prob_formula', 'linear'))
            self.formula_combo = ttk.Combobox(master, textvariable=self.formula_var,
                         values=['linear', 'exponential', 'constant'],
                         state='readonly', width=12)
            self.formula_combo.grid(row=row_next, column=1, padx=5, pady=5, sticky='w')
            self.formula_var.trace_add("write", lambda *a: self.update_probability())
            row_next += 1

            self.update_probability()

            if not hasattr(self.node, "safety_requirements"):
                self.node.safety_requirements = []
            ttk.Label(master, text="Safety Requirements:").grid(row=row_next, column=0, padx=5, pady=5, sticky="ne")
            self.safety_req_frame = ttk.Frame(master)
            self.safety_req_frame.grid(row=row_next, column=1, padx=5, pady=5, sticky="w")
            row_next += 1

            self.safety_req_listbox = tk.Listbox(self.safety_req_frame, height=4, width=50)
            self.safety_req_listbox.grid(row=0, column=0, columnspan=3, sticky="w")
            for req in self.node.safety_requirements:
                self.safety_req_listbox.insert(tk.END, f"[{req['id']}] [{req['req_type']}] [{req.get('asil','')}] {req['text']}")
            self.add_req_button = ttk.Button(self.safety_req_frame, text="Add New", command=self.add_safety_requirement)
            self.add_req_button.grid(row=1, column=0, padx=2, pady=2)
            self.edit_req_button = ttk.Button(self.safety_req_frame, text="Edit", command=self.edit_safety_requirement)
            self.edit_req_button.grid(row=1, column=1, padx=2, pady=2)
            self.delete_req_button = ttk.Button(self.safety_req_frame, text="Delete", command=self.delete_safety_requirement)
            self.delete_req_button.grid(row=1, column=2, padx=2, pady=2)
            self.add_existing_req_button = ttk.Button(self.safety_req_frame, text="Add Existing", command=self.add_existing_requirement)
            self.add_existing_req_button.grid(row=1, column=3, padx=2, pady=2)

        elif self.node.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
            ttk.Label(master, text="Gate Type:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
            self.gate_var = tk.StringVar(value=self.node.gate_type if self.node.gate_type else "AND")
            self.gate_combo = ttk.Combobox(master, textvariable=self.gate_var, values=["AND", "OR"],
                                           state="readonly", width=10)
            self.gate_combo.grid(row=row_next, column=1, padx=5, pady=5)
            row_next += 1
            if self.node.node_type.upper() == "TOP EVENT":
                ttk.Label(master, text="Severity (1-3):").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
                self.sev_combo = ttk.Combobox(master, values=["1", "2", "3"],
                                              state="disabled", width=5, font=dialog_font)
                current_sev = self.node.severity if self.node.severity is not None else 3
                self.sev_combo.set(str(int(current_sev)))
                self.sev_combo.grid(row=row_next, column=1, padx=5, pady=5)
                row_next += 1

                ttk.Label(master, text="Controllability (1-3):").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
                self.cont_combo = ttk.Combobox(master, values=["1", "2", "3"],
                                              state="disabled", width=5, font=dialog_font)
                current_cont = self.node.controllability if self.node.controllability is not None else 3
                self.cont_combo.set(str(int(current_cont)))
                self.cont_combo.grid(row=row_next, column=1, padx=5, pady=5)
                row_next += 1

                ttk.Label(master, text="Safety Goal Description:").grid(row=row_next, column=0, padx=5, pady=5, sticky="ne")
                self.safety_goal_text = tk.Text(master, width=40, height=3, font=dialog_font, wrap="word")
                self.safety_goal_text.insert("1.0", self.node.safety_goal_description)
                self.safety_goal_text.grid(row=row_next, column=1, padx=5, pady=5)
                self.safety_goal_text.bind("<Return>", self.on_enter_pressed)
                row_next += 1

                ttk.Label(master, text="Safety Goal ASIL:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
                self.sg_asil_var = tk.StringVar(value=self.node.safety_goal_asil if self.node.safety_goal_asil else "QM")
                self.sg_asil_combo = ttk.Combobox(
                    master,
                    textvariable=self.sg_asil_var,
                    values=ASIL_LEVEL_OPTIONS,
                    state="disabled",
                    width=8,
                )
                self.sg_asil_combo.grid(row=row_next, column=1, padx=5, pady=5, sticky="w")
                row_next += 1

                ttk.Label(master, text="Safe State:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
                self.safe_state_entry = tk.Entry(master, width=40, font=dialog_font)
                self.safe_state_entry.insert(0, self.node.safe_state)
                self.safe_state_entry.grid(row=row_next, column=1, padx=5, pady=5, sticky="w")
                row_next += 1

                ttk.Label(master, text="FTTI:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
                self.ftti_entry = tk.Entry(master, width=20, font=dialog_font)
                self.ftti_entry.insert(0, getattr(self.node, "ftti", ""))
                self.ftti_entry.grid(row=row_next, column=1, padx=5, pady=5, sticky="w")
                row_next += 1

                ttk.Label(master, text="Acceptance Criteria:").grid(row=row_next, column=0, padx=5, pady=5, sticky="ne")
                self.ac_text = tk.Text(master, width=40, height=3, font=dialog_font, wrap="word")
                self.ac_text.insert("1.0", getattr(self.node, "acceptance_criteria", ""))
                self.ac_text.grid(row=row_next, column=1, padx=5, pady=5)
                self.ac_text.bind("<Return>", self.on_enter_pressed)
                row_next += 1


        if self.node.node_type.upper() not in ["TOP EVENT", "BASIC EVENT"]:
            self.is_page_var = tk.BooleanVar(value=self.node.is_page)
            ttk.Checkbutton(master, text="Is Page Gate?", variable=self.is_page_var)\
                .grid(row=row_next, column=0, columnspan=2, padx=5, pady=5, sticky="w")
            row_next += 1

        if "CONFIDENCE" in self.node.node_type.upper():
            base_name = "Confidence"
        elif "ROBUSTNESS" in self.node.node_type.upper():
            base_name = "Robustness"
        elif "TOP EVENT" in self.node.node_type.upper():
            base_name = "Prototype Assurance Level (PAL)"
        elif "GATE" in self.node.node_type.upper() or "RIGOR" in self.node.node_type.upper():
            base_name = "Rigor"
        else:
            base_name = "Other"

        if self.node.display_label.startswith("Maturity"):
            base_name = "Maturity"

        valid_subtypes = VALID_SUBTYPES.get(base_name, [])
        if not valid_subtypes:
            valid_subtypes = ["None"]
        ttk.Label(master, text="Subtype:").grid(row=row_next, column=0, padx=5, pady=5, sticky="e")
        initial_subtype = self.node.input_subtype if self.node.input_subtype else valid_subtypes[0]
        self.subtype_var = tk.StringVar(value=initial_subtype)
        state = "disabled" if base_name == "Maturity" else "readonly"
        self.subtype_combo = ttk.Combobox(master, textvariable=self.subtype_var, values=valid_subtypes,
                                          state=state, width=20)
        self.subtype_combo.grid(row=row_next, column=1, padx=5, pady=5, sticky="w")
        row_next += 1

        return self.user_name_entry

    class RequirementDialog(simpledialog.Dialog):
        def __init__(self, parent, title, initial_req=None):
            self.initial_req = initial_req or {}
            super().__init__(parent, title=title)
        
        def body(self, master):
            # Instead of master.resizable(), use self.top
            self.resizable(False, False)
            dialog_font = tk.font.Font(family="Arial", size=10)
            ttk.Label(master, text="Requirement Type:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
            self.type_var = tk.StringVar()
            self.type_combo = ttk.Combobox(master, textvariable=self.type_var, 
                                           values=["vehicle", "operational"],
                                           state="readonly", width=15)
            self.type_combo.grid(row=0, column=1, padx=5, pady=5)
            
            ttk.Label(master, text="Custom Requirement ID:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
            self.custom_id_entry = tk.Entry(master, width=20, font=dialog_font)
            # Preload using "custom_id" if available; otherwise, fallback to "id"
            self.custom_id_entry.insert(0, self.initial_req.get("custom_id") or self.initial_req.get("id", ""))
            self.custom_id_entry.grid(row=1, column=1, padx=5, pady=5)

            ttk.Label(master, text="Requirement Text:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
            self.req_entry = tk.Entry(master, width=40, font=dialog_font)
            self.req_entry.grid(row=2, column=1, padx=5, pady=5)

            ttk.Label(master, text="ASIL:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
            self.req_asil_var = tk.StringVar()
            self.req_asil_combo = ttk.Combobox(
                master,
                textvariable=self.req_asil_var,
                values=ASIL_LEVEL_OPTIONS,
                state="readonly",
                width=8,
            )
            self.req_asil_combo.grid(row=3, column=1, padx=5, pady=5, sticky="w")

            self.type_var.set(self.initial_req.get("req_type", "vehicle"))
            self.req_entry.insert(0, self.initial_req.get("text", ""))
            self.req_asil_var.set(self.initial_req.get("asil", "QM"))
            return self.req_entry

        def apply(self):
            req_type = self.type_var.get().strip()
            req_text = self.req_entry.get().strip()
            custom_id = self.custom_id_entry.get().strip()
            asil = self.req_asil_var.get().strip()
            self.result = {"req_type": req_type, "text": req_text, "custom_id": custom_id, "asil": asil}

        def validate(self):
            custom_id = self.custom_id_entry.get().strip()
            # If a custom ID is provided, ensure it's unique unless we're editing this requirement
            if custom_id:
                existing = global_requirements.get(custom_id)
                if existing and custom_id not in (
                    self.initial_req.get("custom_id"),
                    self.initial_req.get("id"),
                ):
                    messagebox.showerror(
                        "Duplicate ID",
                        f"Requirement ID '{custom_id}' already exists. Please choose a unique ID.",
                    )
                    return False
            return True

    class SelectExistingRequirementsDialog(simpledialog.Dialog):
        """
        A dialog that displays all global requirements in a list with checkboxes.
        The user can select one or more existing requirements to add (as clones) to the current node.
        """
        def __init__(self, parent, title="Select Existing Requirements"):
            # We'll use a dict to track checkbox variables keyed by requirement ID.
            self.selected_vars = {}
            super().__init__(parent, title=title)

        def body(self, master):
            ttk.Label(master, text="Select one or more existing requirements:").pack(padx=5, pady=5)

            # Create a container canvas and a vertical scrollbar
            container = ttk.Frame(master)
            container.pack(fill=tk.BOTH, expand=True)

            canvas = tk.Canvas(container, borderwidth=0)
            scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
            self.check_frame = ttk.Frame(canvas)

            # Configure the scrollable region when the frame's size changes
            self.check_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )

            canvas.create_window((0, 0), window=self.check_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            # Pack canvas and scrollbar side by side
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # For each requirement in the global registry, create a Checkbutton.
            for req_id, req in global_requirements.items():
                var = tk.BooleanVar(value=False)
                self.selected_vars[req_id] = var
                text = f"[{req['id']}] [{req['req_type']}] [{req.get('asil','')}] {req['text']}"
                ttk.Checkbutton(self.check_frame, text=text, variable=var).pack(anchor="w", padx=2, pady=2)
            return self.check_frame

        def apply(self):
            # Return a list of requirement IDs that were selected.
            self.result = [req_id for req_id, var in self.selected_vars.items() if var.get()]

    def add_existing_requirement(self):
        """
        Opens a dialog to let the user select one or more existing requirements from the global registry.
        The selected requirements are then allocated to the current node (as clones sharing the same custom ID).
        """
        global global_requirements  # Ensure we refer to the module-level variable
        if not global_requirements:
            messagebox.showinfo("No Existing Requirements", "There are no existing requirements to add.")
            return
        dialog = self.SelectExistingRequirementsDialog(self, title="Select Existing Requirements")
        if dialog.result:
            # For each selected requirement, allocate it to the node if not already present.
            if not hasattr(self.node, "safety_requirements"):
                self.node.safety_requirements = []
            for req_id in dialog.result:
                req = global_requirements.get(req_id)
                if req and not any(r["id"] == req_id for r in self.node.safety_requirements):
                    # For clone semantics, we simply add the same dictionary reference.
                    self.node.safety_requirements.append(req)
                    self.safety_req_listbox.insert(tk.END, f"[{req['id']}] [{req['req_type']}] [{req.get('asil','')}] {req['text']}")
        else:
            messagebox.showinfo("No Selection", "No existing requirements were selected.")
   
    def add_new_requirement(self,custom_id, req_type, text, asil="QM"):
        # When a requirement is created, register it in the global registry.
        req = {"id": custom_id, "req_type": req_type, "text": text, "custom_id": custom_id, "asil": asil, "status": "draft", "parent_id": ""}
        global_requirements[custom_id] = req
        print(f"Added new requirement: {req}")
        return req
        
    def list_all_requirements(self):
        # This function returns a list of formatted strings for all requirements
        return [f"[{req['id']}] [{req['req_type']}] [{req.get('asil','')}] {req['text']}" for req in global_requirements.values()]

    # --- Traceability helpers ---
    def get_requirement_allocation_names(self, req_id):
        """Return a list of node or FMEA entry names where the requirement appears."""
        names = []
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                names.append(n.user_name or f"Node {n.unique_id}")
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    if isinstance(e, dict):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                    else:
                        name = getattr(e, "description", "") or getattr(e, "user_name", f"BE {getattr(e, 'unique_id', '')}")
                    names.append(f"{fmea['name']}:{name}")
        return names

    def _collect_goal_names(self, node, acc):
        if node.node_type.upper() == "TOP EVENT":
            acc.add(node.safety_goal_description or (node.user_name or f"SG {node.unique_id}"))
        for p in getattr(node, "parents", []):
            self._collect_goal_names(p, acc)

    def get_requirement_goal_names(self, req_id):
        """Return a list of safety goal names linked to the requirement."""
        goals = set()
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                self._collect_goal_names(n, goals)
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    if isinstance(e, dict):
                        parent_list = e.get("parents") or []
                    else:
                        parent_list = getattr(e, "parents", []) or []
                    parent = parent_list[0] if parent_list else None
                    if isinstance(parent, dict) and "unique_id" in parent:
                        node = self.find_node_by_id_all(parent["unique_id"])
                    else:
                        node = parent if hasattr(parent, "unique_id") else None
                    if node:
                        self._collect_goal_names(node, goals)
        return sorted(goals)

    def format_requirement_with_trace(self, req):
        """Return requirement text including allocation and safety goal lists."""
        rid = req.get("id", "")
        alloc = ", ".join(self.get_requirement_allocation_names(rid))
        goals = ", ".join(self.get_requirement_goal_names(rid))
        return (
            f"[{rid}] [{req.get('req_type','')}] [{req.get('asil','')}] {req.get('text','')}" +
            f" (Alloc: {alloc}; SGs: {goals})"
        )

    
    def add_safety_requirement(self):
        """
        Opens the custom dialog to create a new requirement.
        Also, provides a button (or similar mechanism) to add existing requirements.
        """
        global global_requirements  # Ensure we refer to the module-level global_requirements
        # Use self.master (the Toplevel parent of this dialog) instead of self.
        dialog = self.RequirementDialog(self.master, title="Add Safety Requirement")
        if dialog.result is None or dialog.result["text"] == "":
            return
        custom_id = dialog.result.get("custom_id", "").strip()
        if not custom_id:
            custom_id = str(uuid.uuid4())
        # Check global registry: if exists, update; otherwise, register new.
        if custom_id in global_requirements:
            req = global_requirements[custom_id]
            req["req_type"] = dialog.result["req_type"]
            req["text"] = dialog.result["text"]
            req["asil"] = dialog.result.get("asil", "QM")
        else:
            req = {
                "id": custom_id,
                "req_type": dialog.result["req_type"],
                "text": dialog.result["text"],
                "custom_id": custom_id,
                "asil": dialog.result.get("asil", "QM"),
                "status": "draft",
                "parent_id": ""
            }
            global_requirements[custom_id] = req

        # Allocate this requirement to the current node if not already present.
        if not hasattr(self.node, "safety_requirements"):
            self.node.safety_requirements = []
        if not any(r["id"] == custom_id for r in self.node.safety_requirements):
            self.node.safety_requirements.append(req)
            self.safety_req_listbox.insert(tk.END, f"[{req['id']}] [{req['req_type']}] [{req.get('asil','')}] {req['text']}")

    def edit_safety_requirement(self):
        """
        Opens the edit dialog for a selected safety requirement.
        After editing, updates the global registry so that all nodes sharing that requirement are synchronized.
        """
        selected = self.safety_req_listbox.curselection()
        if not selected:
            messagebox.showwarning("Edit Requirement", "Select a requirement to edit.")
            return
        index = selected[0]
        current_req = self.node.safety_requirements[index]
        initial_req = current_req.copy()
        # Pass self.master as the parent here as well.
        dialog = self.RequirementDialog(self.master, title="Edit Safety Requirement", initial_req=initial_req)
        if dialog.result is None or dialog.result["text"] == "":
            return
        new_custom_id = dialog.result["custom_id"].strip() or current_req.get("custom_id") or current_req.get("id") or str(uuid.uuid4())
        current_req["req_type"] = dialog.result["req_type"]
        current_req["text"] = dialog.result["text"]
        current_req["asil"] = dialog.result.get("asil", "QM")
        current_req["custom_id"] = new_custom_id
        current_req["id"] = new_custom_id
        global_requirements[new_custom_id] = current_req
        self.node.safety_requirements[index] = current_req
        self.safety_req_listbox.delete(index)
        self.safety_req_listbox.insert(index, f"[{current_req['id']}] [{current_req['req_type']}] [{current_req.get('asil','')}] {current_req['text']}")

    def delete_safety_requirement(self):
        selected = self.safety_req_listbox.curselection()
        if not selected:
            messagebox.showwarning("Delete Requirement", "Select a requirement to delete.")
            return
        index = selected[0]
        del self.node.safety_requirements[index]
        self.safety_req_listbox.delete(index)

    def buttonbox(self):
        box = tk.Frame(self)
        ok_button = tk.Button(box, text="OK", width=10, command=self.ok, default=tk.ACTIVE)
        ok_button.pack(side=tk.LEFT, padx=5, pady=5)
        cancel_button = tk.Button(box, text="Cancel", width=10, command=self.cancel)
        cancel_button.pack(side=tk.LEFT, padx=5, pady=5)
        self.bind("<Escape>", lambda event: self.cancel())
        box.pack()

    def on_enter_pressed(self, event):
        event.widget.insert("insert", "\n")
        return "break"

    def update_probability(self, *_):
        if hasattr(self, "prob_entry") and hasattr(self, "fm_var"):
            formula = self.formula_var.get() if hasattr(self, "formula_var") else None
            if str(formula).strip().lower() == "constant":
                if not self.prob_entry.get().strip():
                    try:
                        val = float(getattr(self.node, "failure_prob", 0.0))
                    except (TypeError, ValueError):
                        val = 0.0
                    self.prob_entry.insert(0, str(val))
                return
            label = self.fm_var.get().strip()
            ref = self.fm_map.get(label)
            prob = self.app.compute_failure_prob(self.node, failure_mode_ref=ref, formula=formula)
            self.prob_entry.delete(0, tk.END)
            self.prob_entry.insert(0, f"{prob:.10g}")

    def validate(self):
        if hasattr(self, 'fm_var'):
            label = self.fm_var.get().strip()
            ref = self.fm_map.get(label)
            if ref:
                for n in self.app.get_all_nodes_in_model():
                    if n is self.node:
                        continue
                    if not n.is_primary_instance:
                        continue
                    if getattr(n, 'failure_mode_ref', None) == ref:
                        messagebox.showerror('Failure Mode', 'Selected failure mode already assigned to another node')
                        return False
        return True

    def apply(self):
        target_node = self.node if self.node.is_primary_instance else self.node.original

        target_node.user_name = self.user_name_entry.get().strip()
        target_node.description = self.desc_text.get("1.0", "end-1c")
        target_node.rationale = self.rationale_text.get("1.0", "end-1c")
        
        if self.node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            try:
                val = float(self.value_combo.get().strip())
                if not (1 <= val <= 5):
                    raise ValueError
                target_node.quant_value = val
            except ValueError:
                messagebox.showerror("Invalid Input", "Select a value between 1 and 5.")
        elif self.node.node_type.upper() == "BASIC EVENT":
            label = self.fm_var.get().strip()
            ref = self.fm_map.get(label)
            target_node.failure_mode_ref = ref
            target_node.prob_formula = self.formula_var.get()
            if target_node.prob_formula == "constant":
                try:
                    target_node.failure_prob = float(self.prob_entry.get().strip())
                except ValueError:
                    target_node.failure_prob = 0.0
            else:
                target_node.failure_prob = self.app.compute_failure_prob(
                    target_node, failure_mode_ref=ref, formula=target_node.prob_formula)
        elif self.node.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
            target_node.gate_type = self.gate_var.get().strip().upper()
            if self.node.node_type.upper() == "TOP EVENT":
                try:
                    sev = float(self.sev_combo.get().strip())
                    if not (1 <= sev <= 3):
                        raise ValueError
                    target_node.severity = sev
                except ValueError:
                    messagebox.showerror("Invalid Input", "Select a severity between 1 and 3.")
                try:
                    cont = float(self.cont_combo.get().strip())
                    if not (1 <= cont <= 3):
                        raise ValueError
                    target_node.controllability = cont
                except ValueError:
                    messagebox.showerror("Invalid Input", "Select a controllability between 1 and 3.")
                target_node.is_page = False
                target_node.safety_goal_description = self.safety_goal_text.get("1.0", "end-1c")
                target_node.safety_goal_asil = self.sg_asil_var.get().strip()
                target_node.safe_state = self.safe_state_entry.get().strip()
                target_node.ftti = self.ftti_entry.get().strip()
                target_node.acceptance_criteria = self.ac_text.get("1.0", "end-1c")
            else:
                target_node.is_page = self.is_page_var.get()

        if hasattr(self, "subtype_var"):
            target_node.input_subtype = self.subtype_var.get()

        self.app.sync_nodes_by_id(target_node)
        AD_RiskAssessment_Helper.calculate_assurance_recursive(
            self.app.root_node,
            self.app.top_events,
        )
        self.app.update_views()

##########################################
# Main Application (Parent Diagram)
##########################################
class FaultTreeApp:
    def __init__(self, root):
        self.root = root
        self.top_events = []
        self.selected_node = None
        self.clone_offset_counter = {}
        self.root.title("Autonomous Driving Risk Assessment")
        self.zoom = 1.0
        self.diagram_font = tkFont.Font(family="Arial", size=int(8 * self.zoom))
        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except tk.TclError:
            pass
        self.style.configure("Treeview", font=("Arial", 10))
        self.clipboard_node = None
        self.cut_mode = False
        self.page_history = []
        self.project_properties = {
            "pdf_report_name": "Autonomous Driving Risk Assessment PDF Report",
            "pdf_detailed_formulas": True,
            "show_grid": True,
            "black_white": False,
        }
        self.mission_profiles = []
        self.fmeda_components = []
        self.reliability_analyses = []
        self.reliability_components = []
        self.reliability_total_fit = 0.0
        self.spfm = 0.0
        self.lpfm = 0.0
        self.reliability_dc = 0.0
        self.hazop_docs = []  # list of HazopDoc
        self.hara_docs = []   # list of HaraDoc
        self.active_hazop = None
        self.active_hara = None
        self.hazop_entries = []  # backwards compatibility for active doc
        self.hara_entries = []
        self.top_events = []
        self.reviews = []
        self.review_data = None
        self.review_window = None
        self.current_user = ""
        self.comment_target = None
        self.versions = []
        self.diff_nodes = []
        self.fi2tc_entries = []
        self.tc2fi_entries = []
        self.scenario_libraries = []
        self.odd_libraries = []
        self.odd_elements = []
        self.update_odd_elements()
        # Provide the drawing helper to dialogs that may be opened later
        self.fta_drawing_helper = fta_drawing_helper

        self.mechanism_libraries = []
        self.selected_mechanism_libraries = []
        self.fmedas = []  # list of FMEDA documents
        self.load_default_mechanisms()

        self.mechanism_libraries = []
        self.selected_mechanism_libraries = []
        self.fmedas = []  # list of FMEDA documents
        self.load_default_mechanisms()

        self.mechanism_libraries = []
        self.load_default_mechanisms()

        menubar = tk.Menu(root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="New", command=self.new_model, accelerator="Ctrl+N")
        file_menu.add_command(label="Save Model", command=self.save_model, accelerator="Ctrl+S")
        file_menu.add_command(label="Load Model", command=self.load_model, accelerator="Ctrl+O")
        file_menu.add_command(label="FMEA Manager", command=self.show_fmea_list)
        file_menu.add_command(label="New Vehicle Level Function", command=self.add_top_level_event)
        file_menu.add_command(label="Project Properties", command=self.edit_project_properties)
        file_menu.add_command(label="Save PDF Report", command=self.generate_pdf_report)
        file_menu.add_command(label="Save PDF Without Assurance", command=self.generate_pdf_without_assurance)
        file_menu.add_command(label="Export SG Requirements", command=self.export_safety_goal_requirements)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.confirm_close)
        menubar.add_cascade(label="File", menu=file_menu)
        edit_menu = tk.Menu(menubar, tearoff=0)
        edit_menu.add_command(label="Add Confidence", command=lambda: self.add_node_of_type("Confidence Level"), accelerator="Ctrl+Shift+C")
        edit_menu.add_command(label="Add Robustness", command=lambda: self.add_node_of_type("Robustness Score"), accelerator="Ctrl+Shift+R")
        edit_menu.add_command(label="Add Gate", command=lambda: self.add_node_of_type("GATE"), accelerator="Ctrl+Shift+G")
        edit_menu.add_command(label="Add Basic Event", command=lambda: self.add_node_of_type("Basic Event"), accelerator="Ctrl+Shift+B")
        edit_menu.add_command(label="Add Triggering Condition", command=lambda: self.add_node_of_type("Triggering Condition"))
        edit_menu.add_command(label="Add Functional Insufficiency", command=lambda: self.add_node_of_type("Functional Insufficiency"))
        edit_menu.add_command(label="Add FMEA/FMEDA Event", command=self.add_basic_event_from_fmea)
        edit_menu.add_command(label="Edit Selected", command=self.edit_selected)
        edit_menu.add_command(label="Remove Connection", command=lambda: self.remove_connection(self.selected_node) if self.selected_node else None)
        edit_menu.add_command(label="Delete Node", command=lambda: self.delete_node_and_subtree(self.selected_node) if self.selected_node else None)
        edit_menu.add_command(label="Remove Node", command=self.remove_node)
        edit_menu.add_separator()
        edit_menu.add_command(label="Copy", command=self.copy_node, accelerator="Ctrl+C")
        edit_menu.add_command(label="Cut", command=self.cut_node, accelerator="Ctrl+X")
        edit_menu.add_command(label="Paste", command=self.paste_node, accelerator="Ctrl+V")
        edit_menu.add_separator()
        edit_menu.add_command(label="Edit User Name", command=self.edit_user_name, accelerator="Ctrl+U")
        edit_menu.add_command(label="Edit Description", command=self.edit_description, accelerator="Ctrl+D")
        edit_menu.add_command(label="Edit Rationale", command=self.edit_rationale, accelerator="Ctrl+L")
        edit_menu.add_command(label="Edit Value", command=self.edit_value)
        edit_menu.add_command(label="Edit Gate Type", command=self.edit_gate_type, accelerator="Ctrl+G")
        edit_menu.add_command(label="Edit Severity", command=self.edit_severity, accelerator="Ctrl+E")
        edit_menu.add_command(label="Edit Controllability", command=self.edit_controllability)
        edit_menu.add_command(label="Edit Page Flag", command=self.edit_page_flag)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        process_menu = tk.Menu(menubar, tearoff=0)
        process_menu.add_command(label="Calc Prototype Assurance Level (PAL)", command=self.calculate_overall, accelerator="Ctrl+R")
        process_menu.add_command(label="Calc PMHF", command=self.calculate_pmfh, accelerator="Ctrl+M")
        menubar.add_cascade(label="Process", menu=process_menu)
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Zoom In", command=self.zoom_in, accelerator="Ctrl++")
        view_menu.add_command(label="Zoom Out", command=self.zoom_out, accelerator="Ctrl+-")
        view_menu.add_command(label="Auto Arrange", command=self.auto_arrange, accelerator="Ctrl+A")
        view_menu.add_command(label="Requirements Matrix", command=self.show_requirements_matrix)
        view_menu.add_command(label="Requirements Editor", command=self.show_requirements_editor)
        view_menu.add_command(label="FTA-FMEA Traceability", command=self.show_traceability_matrix)
        view_menu.add_command(label="Safety Goals Matrix", command=self.show_safety_goals_matrix)
        view_menu.add_command(label="Safety Goals Editor", command=self.show_safety_goals_editor)
        view_menu.add_command(label="FTA Cut Sets", command=self.show_cut_sets)
        view_menu.add_command(label="Common Cause Toolbox", command=self.show_common_cause_view)
        menubar.add_cascade(label="View", menu=view_menu)
        review_menu = tk.Menu(menubar, tearoff=0)
        review_menu.add_command(label="Start Peer Review", command=self.start_peer_review)
        review_menu.add_command(label="Start Joint Review", command=self.start_joint_review)
        review_menu.add_command(label="Open Review Toolbox", command=self.open_review_toolbox)
        review_menu.add_command(label="Set Current User", command=self.set_current_user)
        review_menu.add_command(label="Merge Review Comments", command=self.merge_review_comments)
        review_menu.add_command(label="Compare Versions", command=self.compare_versions)
        menubar.add_cascade(label="Review", menu=review_menu)
        reliability_menu = tk.Menu(menubar, tearoff=0)
        reliability_menu.add_command(label="Mission Profiles", command=self.manage_mission_profiles)
        reliability_menu.add_command(label="Mechanism Libraries", command=self.manage_mechanism_libraries)
        reliability_menu.add_command(label="Reliability Analysis", command=self.open_reliability_window)
        reliability_menu.add_command(label="FMEDA Analysis", command=self.open_fmeda_window)
        reliability_menu.add_command(label="FMEDA Manager", command=self.show_fmeda_list)
        menubar.add_cascade(label="Reliability", menu=reliability_menu)

        hara_menu = tk.Menu(menubar, tearoff=0)
        hara_menu.add_command(label="HAZOP Analysis", command=self.open_hazop_window)
        hara_menu.add_command(label="HARA Analysis", command=self.open_hara_window)
        menubar.add_cascade(label="HARA", menu=hara_menu)
        sotif_menu = tk.Menu(menubar, tearoff=0)
        sotif_menu.add_command(label="Triggering Conditions", command=self.show_triggering_condition_list)
        sotif_menu.add_command(label="Functional Insufficiencies", command=self.show_functional_insufficiency_list)
        sotif_menu.add_separator()
        sotif_menu.add_command(label="FI2TC Analysis", command=self.open_fi2tc_window)
        sotif_menu.add_command(label="TC2FI Analysis", command=self.open_tc2fi_window)
        menubar.add_cascade(label="SOTIF", menu=sotif_menu)

        libs_menu = tk.Menu(menubar, tearoff=0)
        libs_menu.add_command(label="Scenario Libraries", command=self.manage_scenario_libraries)
        libs_menu.add_command(label="ODD Libraries", command=self.manage_odd_libraries)
        menubar.add_cascade(label="Libraries", menu=libs_menu)
        root.config(menu=menubar)
        root.bind("<Control-n>", lambda event: self.new_model())
        root.bind("<Control-s>", lambda event: self.save_model())
        root.bind("<Control-o>", lambda event: self.load_model())
        root.bind("<Control-r>", lambda event: self.calculate_overall())
        root.bind("<Control-m>", lambda event: self.calculate_pmfh())
        root.bind("<Control-=>", lambda event: self.zoom_in())
        root.bind("<Control-minus>", lambda event: self.zoom_out())
        root.bind("<Control-a>", lambda event: self.auto_arrange())
        root.bind("<Control-u>", lambda event: self.edit_user_name())
        root.bind("<Control-d>", lambda event: self.edit_description())
        root.bind("<Control-l>", lambda event: self.edit_rationale())
        root.bind("<Control-g>", lambda event: self.edit_gate_type())
        root.bind("<Control-e>", lambda event: self.edit_severity())
        root.bind("<Control-Shift-c>", lambda event: self.add_node_of_type("Confidence Level"))
        root.bind("<Control-Shift-r>", lambda event: self.add_node_of_type("Robustness Score"))
        root.bind("<Control-Shift-g>", lambda event: self.add_node_of_type("GATE"))
        root.bind("<Control-Shift-b>", lambda event: self.add_node_of_type("Basic Event"))
        root.bind("<Control-Shift-t>", lambda event: self.add_node_of_type("Triggering Condition"))
        root.bind("<Control-Shift-f>", lambda event: self.add_node_of_type("Functional Insufficiency"))
        root.bind("<Control-c>", lambda event: self.copy_node())
        root.bind("<Control-x>", lambda event: self.cut_node())
        root.bind("<Control-v>", lambda event: self.paste_node())
        root.bind("<Control-p>", lambda event: self.save_diagram_png())
        self.main_pane = tk.PanedWindow(root, orient=tk.HORIZONTAL)
        self.main_pane.pack(fill=tk.BOTH, expand=True)
        self.tree_frame = ttk.Frame(self.main_pane)

        self.top_event_controls = ttk.Frame(self.tree_frame)
        self.top_event_controls.pack(side=tk.TOP, fill=tk.X)

        self.move_up_btn = ttk.Button(self.top_event_controls, text="Move Up", command=self.move_top_event_up)
        self.move_up_btn.pack(side=tk.LEFT, padx=2)
        self.move_down_btn = ttk.Button(self.top_event_controls, text="Move Down", command=self.move_top_event_down)
        self.move_down_btn.pack(side=tk.LEFT, padx=2)

        
        self.treeview = ttk.Treeview(self.tree_frame)
        self.treeview.pack(fill=tk.BOTH, expand=True)
        self.treeview.bind("<Double-1>", lambda e: self.edit_selected())
        self.treeview.bind("<ButtonRelease-1>", self.on_treeview_click)
        self.pmhf_var = tk.StringVar(value="")
        self.pmhf_label = ttk.Label(self.tree_frame, textvariable=self.pmhf_var, foreground="blue")
        self.pmhf_label.pack(side=tk.BOTTOM, fill=tk.X, pady=2)
        self.main_pane.add(self.tree_frame, width=300)
        self.canvas_frame = ttk.Frame(self.main_pane)
        self.main_pane.add(self.canvas_frame, stretch="always")
        self.canvas = tk.Canvas(self.canvas_frame, bg="white")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.hbar = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.vbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.config(xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set,
                           scrollregion=(0, 0, 2000, 2000))
        self.canvas.bind("<ButtonPress-3>", self.on_right_mouse_press)
        self.canvas.bind("<B3-Motion>", self.on_right_mouse_drag)
        self.canvas.bind("<ButtonRelease-3>", self.show_context_menu)
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
        self.canvas.bind("<Control-MouseWheel>", self.on_ctrl_mousewheel)
        self.root_node = FaultTreeNode("", "TOP EVENT")
        self.root_node.x, self.root_node.y = 300, 200
        self.top_events = [self.root_node]
        self.fmea_entries = []
        self.fmeas = []  # list of FMEA documents
        self.selected_node = None
        self.dragging_node = None
        self.drag_offset_x = 0
        self.drag_offset_y = 0
        self.grid_size = 20
        self.update_views()
        # Track the last saved state so we can prompt on exit
        self.last_saved_state = json.dumps(self.export_model_data(), sort_keys=True)
        root.protocol("WM_DELETE_WINDOW", self.confirm_close)

    # --- Requirement Traceability Helpers used by reviews and matrix view ---
    def get_requirement_allocation_names(self, req_id):
        """Return a list of node or FMEA entry names where the requirement appears."""
        names = []
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                names.append(n.user_name or f"Node {n.unique_id}")
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    if isinstance(e, dict):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                    else:
                        name = getattr(e, "description", "") or getattr(e, "user_name", f"BE {getattr(e, 'unique_id', '')}")
                    names.append(f"{fmea['name']}:{name}")
        return names

    def _collect_goal_names(self, node, acc):
        if node.node_type.upper() == "TOP EVENT":
            acc.add(node.safety_goal_description or (node.user_name or f"SG {node.unique_id}"))
        for p in getattr(node, "parents", []):
            self._collect_goal_names(p, acc)

    def get_requirement_goal_names(self, req_id):
        """Return a list of safety goal names linked to the requirement."""
        goals = set()
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                self._collect_goal_names(n, goals)
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    parent_list = e.get("parents") if isinstance(e, dict) else getattr(e, "parents", None)
                    parent = parent_list[0] if parent_list else None
                    if isinstance(parent, dict) and "unique_id" in parent:
                        node = self.find_node_by_id_all(parent["unique_id"])
                    else:
                        node = parent if hasattr(parent, "unique_id") else None
                    if node:
                        self._collect_goal_names(node, goals)
        return sorted(goals)

    def format_requirement_with_trace(self, req):
        """Return requirement text including allocation and safety goal lists."""
        rid = req.get("id", "")
        alloc = ", ".join(self.get_requirement_allocation_names(rid))
        goals = ", ".join(self.get_requirement_goal_names(rid))
        return (
            f"[{rid}] [{req.get('req_type','')}] [{req.get('asil','')}] {req.get('text','')}" +
            f" (Alloc: {alloc}; SGs: {goals})"
        )

    def build_requirement_diff_html(self, review):
        """Return HTML highlighting requirement differences for the review."""
        if not self.versions:
            return ""
        base_data = self.versions[-1]["data"]
        current = self.export_model_data(include_versions=False)

        def filter_data(data):
            return {
                "top_events": [t for t in data.get("top_events", []) if t["unique_id"] in review.fta_ids],
                "fmeas": [f for f in data.get("fmeas", []) if f["name"] in review.fmea_names],
                "fmedas": [d for d in data.get("fmedas", []) if d.get("name") in getattr(review, "fmeda_names", [])],
            }

        data1 = filter_data(base_data)
        data2 = filter_data(current)
        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def collect_reqs(node_dict, target):
            for r in node_dict.get("safety_requirements", []):
                rid = r.get("id")
                if rid and rid not in target:
                    target[rid] = r
            for ch in node_dict.get("children", []):
                collect_reqs(ch, target)

        reqs1, reqs2 = {}, {}
        for nid in review.fta_ids:
            if nid in map1:
                collect_reqs(map1[nid], reqs1)
            if nid in map2:
                collect_reqs(map2[nid], reqs2)

        fmea1 = {f["name"]: f for f in data1.get("fmeas", [])}
        fmea2 = {f["name"]: f for f in data2.get("fmeas", [])}
        for name in review.fmea_names:
            for e in fmea1.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs1:
                        reqs1[rid] = r
            for e in fmea2.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs2:
                        reqs2[rid] = r
        for f in data1.get("fmedas", []):
            for e in f.get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs1:
                        reqs1[rid] = r
        for f in data2.get("fmedas", []):
            for e in f.get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs2:
                        reqs2[rid] = r
        for f in data1.get("fmedas", []):
            for e in f.get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs1:
                        reqs1[rid] = r
        for f in data2.get("fmedas", []):
            for e in f.get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs2:
                        reqs2[rid] = r

        import difflib, html

        def html_diff(a, b):
            matcher = difflib.SequenceMatcher(None, a, b)
            parts = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    parts.append(html.escape(a[i1:i2]))
                elif tag == "delete":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                elif tag == "insert":
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
                elif tag == "replace":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
            return "".join(parts)

        lines = []
        all_ids = sorted(set(reqs1) | set(reqs2))
        for rid in all_ids:
            r1 = reqs1.get(rid)
            r2 = reqs2.get(rid)
            if r1 and not r2:
                lines.append(f"Removed: {html.escape(self.format_requirement_with_trace(r1))}")
            elif r2 and not r1:
                lines.append(f"Added: {html.escape(self.format_requirement_with_trace(r2))}")
            else:
                if json.dumps(r1, sort_keys=True) != json.dumps(r2, sort_keys=True):
                    lines.append("Updated: " + html_diff(self.format_requirement_with_trace(r1), self.format_requirement_with_trace(r2)))

        for nid in review.fta_ids:
            n1 = map1.get(nid, {})
            n2 = map2.get(nid, {})
            sg_old = f"{n1.get('safety_goal_description','')} [{n1.get('safety_goal_asil','')}]"
            sg_new = f"{n2.get('safety_goal_description','')} [{n2.get('safety_goal_asil','')}]"
            label = n2.get('user_name') or n1.get('user_name') or f"Node {nid}"
            if sg_old != sg_new:
                lines.append(
                    f"Safety Goal for {html.escape(label)}: " + html_diff(sg_old, sg_new)
                )
            if n1.get('safe_state','') != n2.get('safe_state',''):
                lines.append(
                    f"Safe State for {html.escape(label)}: " + html_diff(n1.get('safe_state',''), n2.get('safe_state',''))
                )

        return "<br>".join(lines)

    # --- Requirement Traceability Helpers used by reviews and matrix view ---
    def get_requirement_allocation_names(self, req_id):
        """Return a list of node or FMEA entry names where the requirement appears."""
        names = []
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                names.append(n.user_name or f"Node {n.unique_id}")
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    if isinstance(e, dict):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                    else:
                        name = getattr(e, "description", "") or getattr(e, "user_name", f"BE {getattr(e, 'unique_id', '')}")
                    names.append(f"{fmea['name']}:{name}")
        return names

    def _collect_goal_names(self, node, acc):
        if node.node_type.upper() == "TOP EVENT":
            acc.add(node.safety_goal_description or (node.user_name or f"SG {node.unique_id}"))
        for p in getattr(node, "parents", []):
            self._collect_goal_names(p, acc)

    def get_requirement_goal_names(self, req_id):
        """Return a list of safety goal names linked to the requirement."""
        goals = set()
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                self._collect_goal_names(n, goals)
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    parent_list = e.get("parents", []) if isinstance(e, dict) else getattr(e, "parents", [])
                    parent = parent_list[0] if parent_list else None
                    if isinstance(parent, dict) and "unique_id" in parent:
                        node = self.find_node_by_id_all(parent["unique_id"])
                    else:
                        node = parent if hasattr(parent, "unique_id") else None
                    if node:
                        self._collect_goal_names(node, goals)
        return sorted(goals)

    def format_requirement_with_trace(self, req):
        """Return requirement text including allocation and safety goal lists."""
        rid = req.get("id", "")
        alloc = ", ".join(self.get_requirement_allocation_names(rid))
        goals = ", ".join(self.get_requirement_goal_names(rid))
        return (
            f"[{rid}] [{req.get('req_type','')}] [{req.get('asil','')}] {req.get('text','')}" +
            f" (Alloc: {alloc}; SGs: {goals})"
        )

    def build_requirement_diff_html(self, review):
        """Return HTML highlighting requirement differences for the review."""
        if not self.versions:
            return ""
        base_data = self.versions[-1]["data"]
        current = self.export_model_data(include_versions=False)

        def filter_data(data):
            return {
                "top_events": [t for t in data.get("top_events", []) if t["unique_id"] in review.fta_ids],
                "fmeas": [f for f in data.get("fmeas", []) if f["name"] in review.fmea_names],
                "fmedas": [d for d in data.get("fmedas", []) if d.get("name") in getattr(review, "fmeda_names", [])],
            }

        data1 = filter_data(base_data)
        data2 = filter_data(current)
        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def collect_reqs(node_dict, target):
            for r in node_dict.get("safety_requirements", []):
                rid = r.get("id")
                if rid and rid not in target:
                    target[rid] = r
            for ch in node_dict.get("children", []):
                collect_reqs(ch, target)

        reqs1, reqs2 = {}, {}
        for nid in review.fta_ids:
            if nid in map1:
                collect_reqs(map1[nid], reqs1)
            if nid in map2:
                collect_reqs(map2[nid], reqs2)

        fmea1 = {f["name"]: f for f in data1.get("fmeas", [])}
        fmea2 = {f["name"]: f for f in data2.get("fmeas", [])}
        for name in review.fmea_names:
            for e in fmea1.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs1:
                        reqs1[rid] = r
            for e in fmea2.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs2:
                        reqs2[rid] = r

        import difflib, html

        def html_diff(a, b):
            matcher = difflib.SequenceMatcher(None, a, b)
            parts = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    parts.append(html.escape(a[i1:i2]))
                elif tag == "delete":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                elif tag == "insert":
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
                elif tag == "replace":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
            return "".join(parts)

        lines = []
        all_ids = sorted(set(reqs1) | set(reqs2))
        for rid in all_ids:
            r1 = reqs1.get(rid)
            r2 = reqs2.get(rid)
            if r1 and not r2:
                lines.append(f"Removed: {html.escape(self.format_requirement_with_trace(r1))}")
            elif r2 and not r1:
                lines.append(f"Added: {html.escape(self.format_requirement_with_trace(r2))}")
            else:
                if json.dumps(r1, sort_keys=True) != json.dumps(r2, sort_keys=True):
                    lines.append("Updated: " + html_diff(self.format_requirement_with_trace(r1), self.format_requirement_with_trace(r2)))

        for nid in review.fta_ids:
            n1 = map1.get(nid, {})
            n2 = map2.get(nid, {})
            sg_old = f"{n1.get('safety_goal_description','')} [{n1.get('safety_goal_asil','')}]"
            sg_new = f"{n2.get('safety_goal_description','')} [{n2.get('safety_goal_asil','')}]"
            label = n2.get('user_name') or n1.get('user_name') or f"Node {nid}"
            if sg_old != sg_new:
                lines.append(
                    f"Safety Goal for {html.escape(label)}: " + html_diff(sg_old, sg_new)
                )
            if n1.get('safe_state','') != n2.get('safe_state',''):
                lines.append(
                    f"Safe State for {html.escape(label)}: " + html_diff(n1.get('safe_state',''), n2.get('safe_state',''))
                )

        return "<br>".join(lines)

    # --- Requirement Traceability Helpers used by reviews and matrix view ---
    def get_requirement_allocation_names(self, req_id):
        """Return a list of node or FMEA entry names where the requirement appears."""
        names = []
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                names.append(n.user_name or f"Node {n.unique_id}")
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    if isinstance(e, dict):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                    else:
                        name = getattr(e, "description", "") or getattr(e, "user_name", f"BE {getattr(e, 'unique_id', '')}")
                    names.append(f"{fmea['name']}:{name}")
        return names

    def _collect_goal_names(self, node, acc):
        if node.node_type.upper() == "TOP EVENT":
            acc.add(node.safety_goal_description or (node.user_name or f"SG {node.unique_id}"))
        for p in getattr(node, "parents", []):
            self._collect_goal_names(p, acc)

    def get_requirement_goal_names(self, req_id):
        """Return a list of safety goal names linked to the requirement."""
        goals = set()
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                self._collect_goal_names(n, goals)
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    parent_list = e.get("parents", []) if isinstance(e, dict) else getattr(e, "parents", [])
                    parent = parent_list[0] if parent_list else None
                    if isinstance(parent, dict) and "unique_id" in parent:
                        node = self.find_node_by_id_all(parent["unique_id"])
                    else:
                        node = parent if hasattr(parent, "unique_id") else None
                    if node:
                        self._collect_goal_names(node, goals)
        return sorted(goals)

    def format_requirement_with_trace(self, req):
        """Return requirement text including allocation and safety goal lists."""
        if isinstance(req, dict):
            rid = req.get("id", "")
            rtype = req.get("req_type", "")
            asil = req.get("asil", "")
            text = req.get("text", "")
        else:
            rid = getattr(req, "id", "")
            rtype = getattr(req, "req_type", "")
            asil = getattr(req, "asil", "")
            text = getattr(req, "text", "")
        alloc = ", ".join(self.get_requirement_allocation_names(rid))
        goals = ", ".join(self.get_requirement_goal_names(rid))
        return f"[{rid}] [{rtype}] [{asil}] {text} (Alloc: {alloc}; SGs: {goals})"

    def build_requirement_diff_html(self, review):
        """Return HTML highlighting requirement differences for the review."""
        if not self.versions:
            return ""
        base_data = self.versions[-1]["data"]
        current = self.export_model_data(include_versions=False)

        def filter_data(data):
            return {
                "top_events": [
                    t for t in data.get("top_events", []) if t["unique_id"] in review.fta_ids
                ],
                "fmeas": [
                    f for f in data.get("fmeas", []) if f["name"] in review.fmea_names
                ],
                "fmedas": [
                    d
                    for d in data.get("fmedas", [])
                    if d.get("name") in getattr(review, "fmeda_names", [])
                ],
                "hazops": [
                    d
                    for d in data.get("hazops", [])
                    if d.get("name") in getattr(review, "hazop_names", [])
                ],
                "haras": [
                    d
                    for d in data.get("haras", [])
                    if d.get("name") in getattr(review, "hara_names", [])
                ],
            }

        data1 = filter_data(base_data)
        data2 = filter_data(current)
        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def collect_reqs(node_dict, target):
            for r in node_dict.get("safety_requirements", []):
                rid = r.get("id")
                if rid and rid not in target:
                    target[rid] = r
            for ch in node_dict.get("children", []):
                collect_reqs(ch, target)

        reqs1, reqs2 = {}, {}
        for nid in review.fta_ids:
            if nid in map1:
                collect_reqs(map1[nid], reqs1)
            if nid in map2:
                collect_reqs(map2[nid], reqs2)

        fmea1 = {f["name"]: f for f in data1.get("fmeas", [])}
        fmea2 = {f["name"]: f for f in data2.get("fmeas", [])}
        for name in review.fmea_names:
            for e in fmea1.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs1:
                        reqs1[rid] = r
            for e in fmea2.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs2:
                        reqs2[rid] = r

        import difflib, html

        def html_diff(a, b):
            matcher = difflib.SequenceMatcher(None, a, b)
            parts = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    parts.append(html.escape(a[i1:i2]))
                elif tag == "delete":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                elif tag == "insert":
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
                elif tag == "replace":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
            return "".join(parts)

        lines = []
        all_ids = sorted(set(reqs1) | set(reqs2))
        for rid in all_ids:
            r1 = reqs1.get(rid)
            r2 = reqs2.get(rid)
            if r1 and not r2:
                lines.append(f"Removed: {html.escape(self.format_requirement_with_trace(r1))}")
            elif r2 and not r1:
                lines.append(f"Added: {html.escape(self.format_requirement_with_trace(r2))}")
            else:
                if json.dumps(r1, sort_keys=True) != json.dumps(r2, sort_keys=True):
                    lines.append("Updated: " + html_diff(self.format_requirement_with_trace(r1), self.format_requirement_with_trace(r2)))

        for nid in review.fta_ids:
            n1 = map1.get(nid, {})
            n2 = map2.get(nid, {})
            sg_old = f"{n1.get('safety_goal_description','')} [{n1.get('safety_goal_asil','')}]"
            sg_new = f"{n2.get('safety_goal_description','')} [{n2.get('safety_goal_asil','')}]"
            label = n2.get('user_name') or n1.get('user_name') or f"Node {nid}"
            if sg_old != sg_new:
                lines.append(
                    f"Safety Goal for {html.escape(label)}: " + html_diff(sg_old, sg_new)
                )
            if n1.get('safe_state','') != n2.get('safe_state',''):
                lines.append(
                    f"Safe State for {html.escape(label)}: " + html_diff(n1.get('safe_state',''), n2.get('safe_state',''))
                )

        return "<br>".join(lines)

    # --- Requirement Traceability Helpers used by reviews and matrix view ---
    def get_requirement_allocation_names(self, req_id):
        """Return a list of node or FMEA entry names where the requirement appears."""
        names = []
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                names.append(n.user_name or f"Node {n.unique_id}")
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    if isinstance(e, dict):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                    else:
                        name = getattr(e, "description", "") or getattr(e, "user_name", f"BE {getattr(e, 'unique_id', '')}")
                    names.append(f"{fmea['name']}:{name}")
        return names

    def _collect_goal_names(self, node, acc):
        if node.node_type.upper() == "TOP EVENT":
            acc.add(node.safety_goal_description or (node.user_name or f"SG {node.unique_id}"))
        for p in getattr(node, "parents", []):
            self._collect_goal_names(p, acc)

    def get_requirement_goal_names(self, req_id):
        """Return a list of safety goal names linked to the requirement."""
        goals = set()
        for n in self.get_all_nodes(self.root_node):
            reqs = getattr(n, "safety_requirements", [])
            if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                self._collect_goal_names(n, goals)
        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                reqs = e.get("safety_requirements", []) if isinstance(e, dict) else getattr(e, "safety_requirements", [])
                if any((r.get("id") if isinstance(r, dict) else getattr(r, "id", None)) == req_id for r in reqs):
                    parent_list = e.get("parents", []) if isinstance(e, dict) else getattr(e, "parents", [])
                    parent = parent_list[0] if parent_list else None
                    if isinstance(parent, dict) and "unique_id" in parent:
                        node = self.find_node_by_id_all(parent["unique_id"])
                    else:
                        node = parent if hasattr(parent, "unique_id") else None
                    if node:
                        self._collect_goal_names(node, goals)
        return sorted(goals)

    def format_requirement_with_trace(self, req):
        """Return requirement text including allocation and safety goal lists."""
        if isinstance(req, dict):
            rid = req.get("id", "")
            rtype = req.get("req_type", "")
            asil = req.get("asil", "")
            text = req.get("text", "")
        else:
            rid = getattr(req, "id", "")
            rtype = getattr(req, "req_type", "")
            asil = getattr(req, "asil", "")
            text = getattr(req, "text", "")
        alloc = ", ".join(self.get_requirement_allocation_names(rid))
        goals = ", ".join(self.get_requirement_goal_names(rid))
        return f"[{rid}] [{rtype}] [{asil}] {text} (Alloc: {alloc}; SGs: {goals})"

    def build_requirement_diff_html(self, review):
        """Return HTML highlighting requirement differences for the review."""
        if not self.versions:
            return ""
        base_data = self.versions[-1]["data"]
        current = self.export_model_data(include_versions=False)

        def filter_data(data):
            return {
                "top_events": [t for t in data.get("top_events", []) if t["unique_id"] in review.fta_ids],
                "fmeas": [f for f in data.get("fmeas", []) if f["name"] in review.fmea_names],
            }

        data1 = filter_data(base_data)
        data2 = filter_data(current)
        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def collect_reqs(node_dict, target):
            for r in node_dict.get("safety_requirements", []):
                rid = r.get("id")
                if rid and rid not in target:
                    target[rid] = r
            for ch in node_dict.get("children", []):
                collect_reqs(ch, target)

        reqs1, reqs2 = {}, {}
        for nid in review.fta_ids:
            if nid in map1:
                collect_reqs(map1[nid], reqs1)
            if nid in map2:
                collect_reqs(map2[nid], reqs2)

        fmea1 = {f["name"]: f for f in data1.get("fmeas", [])}
        fmea2 = {f["name"]: f for f in data2.get("fmeas", [])}
        for name in review.fmea_names:
            for e in fmea1.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs1:
                        reqs1[rid] = r
            for e in fmea2.get(name, {}).get("entries", []):
                for r in e.get("safety_requirements", []):
                    rid = r.get("id")
                    if rid and rid not in reqs2:
                        reqs2[rid] = r

        import difflib, html

        def html_diff(a, b):
            matcher = difflib.SequenceMatcher(None, a, b)
            parts = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    parts.append(html.escape(a[i1:i2]))
                elif tag == "delete":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                elif tag == "insert":
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
                elif tag == "replace":
                    parts.append(f"<span style='color:red'>{html.escape(a[i1:i2])}</span>")
                    parts.append(f"<span style='color:blue'>{html.escape(b[j1:j2])}</span>")
            return "".join(parts)

        lines = []
        all_ids = sorted(set(reqs1) | set(reqs2))
        for rid in all_ids:
            r1 = reqs1.get(rid)
            r2 = reqs2.get(rid)
            if r1 and not r2:
                lines.append(f"Removed: {html.escape(self.format_requirement_with_trace(r1))}")
            elif r2 and not r1:
                lines.append(f"Added: {html.escape(self.format_requirement_with_trace(r2))}")
            else:
                if json.dumps(r1, sort_keys=True) != json.dumps(r2, sort_keys=True):
                    lines.append("Updated: " + html_diff(self.format_requirement_with_trace(r1), self.format_requirement_with_trace(r2)))

        for nid in review.fta_ids:
            n1 = map1.get(nid, {})
            n2 = map2.get(nid, {})
            sg_old = f"{n1.get('safety_goal_description','')} [{n1.get('safety_goal_asil','')}]"
            sg_new = f"{n2.get('safety_goal_description','')} [{n2.get('safety_goal_asil','')}]"
            label = n2.get('user_name') or n1.get('user_name') or f"Node {nid}"
            if sg_old != sg_new:
                lines.append(
                    f"Safety Goal for {html.escape(label)}: " + html_diff(sg_old, sg_new)
                )
            if n1.get('safe_state','') != n2.get('safe_state',''):
                lines.append(
                    f"Safe State for {html.escape(label)}: " + html_diff(n1.get('safe_state',''), n2.get('safe_state',''))
                )

        return "<br>".join(lines)

    def generate_recommendations_for_top_event(self, node):
        # Determine the Prototype Assurance Level (PAL) based on the node’s quantitative score.
        level = AD_RiskAssessment_Helper.discretize_level(node.quant_value) if node.quant_value is not None else 1
        rec = dynamic_recommendations.get(level, {})
        rec_text = f"<b>Recommendations for Prototype Assurance Level (PAL) {level}:</b><br/>"
        for category in ["Testing Requirements", "IFTD Responsibilities", "Preventive Maintenance Actions", "Relevant AVSC Guidelines"]:
            if category in rec:
                rec_text += f"<b>{category}:</b><br/><ul><li>{rec[category]}</li></ul><br/>"
        return rec_text

    def back_all_pages(self):
        if self.page_history:
            # Jump to the very first page saved in history:
            first_page = self.page_history[0]
            # Clear the history so that subsequent back presses do not try to go further.
            self.page_history = []
            for widget in self.canvas_frame.winfo_children():
                widget.destroy()
            self.open_page_diagram(first_page)
        else:
            # No history: you could simply reinitialize the main diagram
            self.close_page_diagram()

    def move_top_event_up(self):
        sel = self.treeview.selection()
        if not sel:
            messagebox.showwarning("Move Up", "Select a top-level event to move.")
            return
        try:
            node_id = int(self.treeview.item(sel[0], "tags")[0])
        except Exception:
            return
        # Find the index in the top_events list.
        index = next((i for i, event in enumerate(self.top_events) if event.unique_id == node_id), None)
        if index is None:
            messagebox.showwarning("Move Up", "The selected node is not a top-level event.")
            return
        if index == 0:
            messagebox.showinfo("Move Up", "This event is already at the top.")
            return
        # Swap with the one above it.
        self.top_events[index], self.top_events[index - 1] = self.top_events[index - 1], self.top_events[index]
        self.update_views()

    def move_top_event_down(self):
        sel = self.treeview.selection()
        if not sel:
            messagebox.showwarning("Move Down", "Select a top-level event to move.")
            return
        try:
            node_id = int(self.treeview.item(sel[0], "tags")[0])
        except Exception:
            return
        index = next((i for i, event in enumerate(self.top_events) if event.unique_id == node_id), None)
        if index is None:
            messagebox.showwarning("Move Down", "The selected node is not a top-level event.")
            return
        if index == len(self.top_events) - 1:
            messagebox.showinfo("Move Down", "This event is already at the bottom.")
            return
        # Swap with the one below it.
        self.top_events[index], self.top_events[index + 1] = self.top_events[index + 1], self.top_events[index]
        self.update_views()

    def get_top_level_nodes(self):
        """Return a list of all nodes that have no parent."""
        all_nodes = self.get_all_nodes()
        top_level = [node for node in all_nodes if not node.parents]
        return top_level
        
    def find_node_by_id_all(self, unique_id):
        for top in self.top_events:
            result = self.find_node_by_id(top, unique_id)
            if result is not None:
                return result

        for entry in self.fmea_entries:
            if getattr(entry, "unique_id", None) == unique_id:
                return entry

        for fmea in self.fmeas:
            for e in fmea.get("entries", []):
                if getattr(e, "unique_id", None) == unique_id:
                    return e

        for d in self.fmedas:
            for e in d.get("entries", []):
                if getattr(e, "unique_id", None) == unique_id:
                    return e

        return None

    def get_hazop_by_name(self, name):
        for d in self.hazop_docs:
            if d.name == name:
                return d
        return None

    def get_hara_by_name(self, name):
        for d in self.hara_docs:
            if d.name == name:
                return d
        return None

    def get_safety_goal_asil(self, sg_name):
        """Return the highest ASIL level for a safety goal name across approved HARAs."""
        best = "QM"
        for doc in getattr(self, "hara_docs", []):
            if not getattr(doc, "approved", False):
                continue
            for e in doc.entries:
                if sg_name and sg_name == e.safety_goal and ASIL_ORDER.get(e.asil, 0) > ASIL_ORDER.get(best, 0):
                    best = e.asil
        for te in self.top_events:
            if sg_name and (sg_name == te.user_name or sg_name == te.safety_goal_description):
                if ASIL_ORDER.get(te.safety_goal_asil or "QM", 0) > ASIL_ORDER.get(best, 0):
                    best = te.safety_goal_asil or "QM"
        return best

    def sync_hara_to_safety_goals(self):
        """Propagate HARA values to safety goals when the HARA is approved."""
        sg_data = {}
        for doc in getattr(self, "hara_docs", []):
            if not getattr(doc, "approved", False):
                continue
            for e in doc.entries:
                if not e.safety_goal:
                    continue
                data = sg_data.setdefault(e.safety_goal, {"asil": "QM", "severity": 1, "cont": 1})
                if ASIL_ORDER.get(e.asil, 0) > ASIL_ORDER.get(data["asil"], 0):
                    data["asil"] = e.asil
                if e.severity > data["severity"]:
                    data["severity"] = e.severity
                if e.controllability > data["cont"]:
                    data["cont"] = e.controllability
        for te in self.top_events:
            name = te.safety_goal_description or (te.user_name or f"SG {te.unique_id}")
            if name in sg_data:
                te.safety_goal_asil = sg_data[name]["asil"]
                te.severity = sg_data[name]["severity"]
                te.controllability = sg_data[name]["cont"]

    def edit_selected(self):
        sel = self.treeview.selection()
        target = None
        if sel:
            try:
                node_id = int(self.treeview.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                return
            target = self.find_node_by_id_all(node_id)
        elif self.selected_node:
            target = self.selected_node
        if not target:
            messagebox.showwarning("No selection", "Select a node to edit.")
            return

        # If the node is a clone, resolve it to its original.
        if not target.is_primary_instance and hasattr(target, "original") and target.original:
            target = target.original

        EditNodeDialog(self.root, target, self)
        self.update_views()

    def add_top_level_event(self):
        new_event = FaultTreeNode("", "TOP EVENT")
        new_event.x, new_event.y = 300, 200
        new_event.is_top_event = True
        self.top_events.append(new_event)
        self.root_node = new_event
        self.update_views()

    def edit_project_properties(self):
        prop_win = tk.Toplevel(self.root)
        prop_win.title("Project Properties")
        prop_win.geometry("400x200")
        dialog_font = tkFont.Font(family="Arial", size=10)

        ttk.Label(prop_win, text="PDF Report Name:", font=dialog_font).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        pdf_entry = tk.Entry(prop_win, width=40, font=dialog_font)
        pdf_entry.insert(0, self.project_properties.get("pdf_report_name", "Autonomous Driving Risk Assessment PDF Report"))
        pdf_entry.grid(row=0, column=1, padx=10, pady=10)

        # New checkbox to choose between detailed formulas or score results only.
        # Default to showing detailed formulas.
        var_detailed = tk.BooleanVar(value=self.project_properties.get("pdf_detailed_formulas", True))
        chk = tk.Checkbutton(prop_win, text="Show Detailed Formulas in PDF Report", variable=var_detailed, font=dialog_font)
        chk.grid(row=1, column=0, columnspan=2, padx=10, pady=5, sticky="w")

        var_grid = tk.BooleanVar(value=self.project_properties.get("show_grid", True))
        chk_grid = tk.Checkbutton(prop_win, text="Show Grid", variable=var_grid, font=dialog_font)
        chk_grid.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="w")

        var_bw = tk.BooleanVar(value=self.project_properties.get("black_white", False))
        chk_bw = tk.Checkbutton(prop_win, text="Black and White FTA", variable=var_bw, font=dialog_font)
        chk_bw.grid(row=3, column=0, columnspan=2, padx=10, pady=5, sticky="w")

        def save_props():
            new_name = pdf_entry.get().strip()
            if new_name:
                self.project_properties["pdf_report_name"] = new_name
                self.project_properties["pdf_detailed_formulas"] = var_detailed.get()
                self.project_properties["show_grid"] = var_grid.get()
                self.project_properties["black_white"] = var_bw.get()
                messagebox.showinfo("Project Properties", "Project properties updated.")
            else:
                messagebox.showwarning("Project Properties", "PDF Report Name cannot be empty.")
            prop_win.destroy()

        def save_props():
            new_name = pdf_entry.get().strip()
            if new_name:
                self.project_properties["pdf_report_name"] = new_name
                self.project_properties["pdf_detailed_formulas"] = var_detailed.get()
                self.project_properties["show_grid"] = var_grid.get()
                self.project_properties["black_white"] = var_bw.get()
                messagebox.showinfo("Project Properties", "Project properties updated.")
            else:
                messagebox.showwarning("Project Properties", "PDF Report Name cannot be empty.")
            prop_win.destroy()

        def save_props():
            new_name = pdf_entry.get().strip()
            if new_name:
                self.project_properties["pdf_report_name"] = new_name
                self.project_properties["pdf_detailed_formulas"] = var_detailed.get()
                self.project_properties["show_grid"] = var_grid.get()
                self.project_properties["black_white"] = var_bw.get()
                messagebox.showinfo("Project Properties", "Project properties updated.")
            else:
                messagebox.showwarning("Project Properties", "PDF Report Name cannot be empty.")
            prop_win.destroy()

        save_btn = tk.Button(prop_win, text="Save", command=save_props, font=dialog_font)
        save_btn.grid(row=4, column=0, columnspan=2, pady=10)
        prop_win.transient(self.root)
        prop_win.grab_set()
        self.root.wait_window(prop_win)

    def create_diagram_image(self):
        self.canvas.update()
        bbox = self.canvas.bbox("all")
        if not bbox:
            return None
        x, y, w, h = bbox[0], bbox[1], bbox[2]-bbox[0], bbox[3]-bbox[1]
        ps = self.canvas.postscript(colormode="color", x=x, y=y, width=w, height=h)
        from io import BytesIO
        ps_bytes = BytesIO(ps.encode("utf-8"))
        img = Image.open(ps_bytes)
        img.load(scale=3)
        return img.convert("RGB")

    def get_page_nodes(self, node):
        result = []
        if node.is_page and node != self.root_node:
            result.append(node)
        for child in node.children:
            result.extend(self.get_page_nodes(child))
        return result

    def capture_page_diagram(self, page_node):
        """
        Create an off-screen Toplevel with a Canvas, draw the page diagram (using PageDiagram),
        and return a PIL Image of the diagram.
        """
        from io import BytesIO
        from PIL import Image

        # Create a temporary Toplevel window and canvas
        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()
        
        # Create and redraw the page diagram
        pd = PageDiagram(self, page_node, canvas)
        pd.redraw_canvas()
        
        # Remove grid if present and force an update
        canvas.delete("grid")
        canvas.update()
        
        # Get the bounding box; print debug info if empty.
        bbox = canvas.bbox("all")
        if not bbox:
            print(f"Debug: No drawing found for page node {page_node.unique_id} - bbox is empty.")
            temp.destroy()
            return None
        
        x, y, x2, y2 = bbox
        width, height = x2 - x, y2 - y
        print(f"Debug: Capturing page diagram for node {page_node.unique_id} with bbox=({x},{y},{x2},{y2})")
        
        # Get the PostScript output for the region.
        ps = canvas.postscript(colormode="color", x=x, y=y, width=width, height=height)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception as e:
            print(f"Debug: Error loading image for page node {page_node.unique_id}: {e}")
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            def req_lines(reqs):
                return "; ".join(
                    self.format_requirement_with_trace(r) for r in reqs
                )

            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(
                    old_data.get("description", ""), new_data.get("description", "")
                )
                rat_segments = [("Rationale: ", "black")] + diff_segments(
                    old_data.get("rationale", ""), new_data.get("rationale", "")
                )
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            req_segments = []

            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            def req_lines(reqs):
                return "; ".join(
                    self.format_requirement_with_trace(r) for r in reqs
                )

            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(
                    old_data.get("description", ""), new_data.get("description", "")
                )
                rat_segments = [("Rationale: ", "black")] + diff_segments(
                    old_data.get("rationale", ""), new_data.get("rationale", "")
                )
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            req_segments = []

            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            def req_lines(reqs):
                return "; ".join(
                    self.format_requirement_with_trace(r) for r in reqs
                )

            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(
                    old_data.get("description", ""), new_data.get("description", "")
                )
                rat_segments = [("Rationale: ", "black")] + diff_segments(
                    old_data.get("rationale", ""), new_data.get("rationale", "")
                )
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            req_segments = []

            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            def req_lines(reqs):
                return "; ".join(
                    self.format_requirement_with_trace(r) for r in reqs
                )

            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(
                    old_data.get("description", ""), new_data.get("description", "")
                )
                rat_segments = [("Rationale: ", "black")] + diff_segments(
                    old_data.get("rationale", ""), new_data.get("rationale", "")
                )
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            req_segments = []

            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            def req_lines(reqs):
                return "; ".join(
                    f"[{r.get('id','')}] [{r.get('req_type','')}] {r.get('text','')}" for r in reqs
                )

            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(
                    old_data.get("description", ""), new_data.get("description", "")
                )
                rat_segments = [("Rationale: ", "black")] + diff_segments(
                    old_data.get("rationale", ""), new_data.get("rationale", "")
                )
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            req_segments = []

            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(old_data.get("description", ""), new_data.get("description", ""))
                rat_segments = [("Rationale: ", "black")] + diff_segments(old_data.get("rationale", ""), new_data.get("rationale", ""))
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(old_data.get("description", ""), new_data.get("description", ""))
                rat_segments = [("Rationale: ", "black")] + diff_segments(old_data.get("rationale", ""), new_data.get("rationale", ""))
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(old_data.get("description", ""), new_data.get("description", ""))
                rat_segments = [("Rationale: ", "black")] + diff_segments(old_data.get("rationale", ""), new_data.get("rationale", ""))
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def capture_diff_diagram(self, top_event):
        """Return an image of the FTA with diff colouring versus last version."""
        if not self.versions:
            return self.capture_page_diagram(top_event)

        from io import BytesIO
        from PIL import Image
        import difflib

        current = self.export_model_data(include_versions=False)
        base_data = self.versions[-1]["data"]

        def filter_events(data):
            return [t for t in data.get("top_events", []) if t["unique_id"] == top_event.unique_id]

        data1 = {"top_events": filter_events(base_data)}
        data2 = {"top_events": filter_events(current)}

        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])

        def build_conn_set(events):
            conns = set()
            def visit(d):
                for ch in d.get("children", []):
                    conns.add((d["unique_id"], ch["unique_id"]))
                    visit(ch)
            for t in events:
                visit(t)
            return conns

        conns1 = build_conn_set(data1["top_events"])
        conns2 = build_conn_set(data2["top_events"])

        conn_status = {}
        for c in conns1 | conns2:
            if c in conns1 and c not in conns2:
                conn_status[c] = "removed"
            elif c in conns2 and c not in conns1:
                conn_status[c] = "added"
            else:
                conn_status[c] = "existing"

        status = {}
        for nid in set(map1) | set(map2):
            if nid in map1 and nid not in map2:
                status[nid] = "removed"
            elif nid in map2 and nid not in map1:
                status[nid] = "added"
            else:
                if json.dumps(map1[nid], sort_keys=True) != json.dumps(map2[nid], sort_keys=True):
                    status[nid] = "added"
                else:
                    status[nid] = "existing"

        module = sys.modules.get(self.__class__.__module__)
        FaultTreeNodeCls = getattr(module, 'FaultTreeNode', None)
        if not FaultTreeNodeCls and self.top_events:
            FaultTreeNodeCls = type(self.top_events[0])
        new_roots = [FaultTreeNodeCls.from_dict(t) for t in data2["top_events"]]
        removed_ids = [nid for nid, st in status.items() if st == "removed"]
        for rid in removed_ids:
            if rid in map1:
                nd = map1[rid]
                new_roots.append(FaultTreeNodeCls.from_dict(nd))

        allow_ids = set()
        def collect_ids(d):
            allow_ids.add(d["unique_id"])
            for ch in d.get("children", []):
                collect_ids(ch)
        if top_event.unique_id in map1:
            collect_ids(map1[top_event.unique_id])
        if top_event.unique_id in map2:
            collect_ids(map2[top_event.unique_id])

        node_objs = {}
        def collect_nodes(n):
            if n.unique_id not in node_objs:
                node_objs[n.unique_id] = n
            for ch in n.children:
                collect_nodes(ch)
        for r in new_roots:
            collect_nodes(r)

        def diff_segments(old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            segments = []
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    segments.append((old[i1:i2], "black"))
                elif tag == "delete":
                    segments.append((old[i1:i2], "red"))
                elif tag == "insert":
                    segments.append((new[j1:j2], "blue"))
                elif tag == "replace":
                    segments.append((old[i1:i2], "red"))
                    segments.append((new[j1:j2], "blue"))
            return segments

        def draw_segment_text(canvas, cx, cy, segments, font_obj):
            lines = [[]]
            for text, color in segments:
                parts = text.split("\n")
                for idx, part in enumerate(parts):
                    if idx > 0:
                        lines.append([])
                    lines[-1].append((part, color))
            line_height = font_obj.metrics("linespace")
            total_height = line_height * len(lines)
            start_y = cy - total_height / 2
            for line in lines:
                line_width = sum(font_obj.measure(part) for part, _ in line)
                start_x = cx - line_width / 2
                x = start_x
                for part, color in line:
                    if part:
                        canvas.create_text(x, start_y, text=part, anchor="nw", fill=color, font=font_obj)
                        x += font_obj.measure(part)
                start_y += line_height

        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()

        def draw_connections(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_connections(ch)
                return
            region_width = 60
            parent_bottom = (n.x, n.y + 20)
            for i, ch in enumerate(n.children):
                if ch.unique_id not in allow_ids:
                    continue
                parent_conn = (
                    n.x - region_width / 2 + (i + 0.5) * (region_width / len(n.children)),
                    parent_bottom[1],
                )
                child_top = (ch.x, ch.y - 25)
                edge_st = conn_status.get((n.unique_id, ch.unique_id), "existing")
                if status.get(n.unique_id) == "removed" or status.get(ch.unique_id) == "removed":
                    edge_st = "removed"
                color = "gray"
                if edge_st == "added":
                    color = "blue"
                elif edge_st == "removed":
                    color = "red"
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top, outline_color=color, line_width=1)
                draw_connections(ch)

        def draw_node(n):
            if n.unique_id not in allow_ids:
                for ch in n.children:
                    draw_node(ch)
                return
            st = status.get(n.unique_id, "existing")
            color = "dimgray"
            if st == "added":
                color = "blue"
            elif st == "removed":
                color = "red"

            source = n if getattr(n, "is_primary_instance", True) else getattr(n, "original", n)
            subtype_text = source.input_subtype if source.input_subtype else "N/A"
            display_label = source.display_label
            old_data = map1.get(n.unique_id)
            new_data = map2.get(n.unique_id)
            def req_lines(reqs):
                return "; ".join(
                    self.format_requirement_with_trace(r) for r in reqs
                )

            if old_data and new_data:
                desc_segments = [("Desc: ", "black")] + diff_segments(
                    old_data.get("description", ""), new_data.get("description", "")
                )
                rat_segments = [("Rationale: ", "black")] + diff_segments(
                    old_data.get("rationale", ""), new_data.get("rationale", "")
                )
            else:
                desc_segments = [("Desc: " + source.description, "black")]
                rat_segments = [("Rationale: " + source.rationale, "black")]
            req_segments = []

            segments = [
                (f"Type: {source.node_type}\n", "black"),
                (f"Subtype: {subtype_text}\n", "black"),
                (f"{display_label}\n", "black"),
            ] + desc_segments + [("\n\n", "black")] + rat_segments

            top_text = "".join(seg[0] for seg in segments)
            bottom_text = n.name
            fill = self.get_node_fill_color(n)
            eff_x, eff_y = n.x, n.y
            typ = n.node_type.upper()
            items_before = canvas.find_all()
            if typ in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if n.gate_type and n.gate_type.upper() == "OR":
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_or_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
                else:
                    if self.fta_drawing_helper:
                        self.fta_drawing_helper.draw_rotated_and_gate_shape(canvas, eff_x, eff_y, scale=40, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)
            else:
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45, top_text=top_text, bottom_text=bottom_text, fill=fill, outline_color=color, line_width=2)

            items_after = canvas.find_all()
            text_id = None
            for item in items_after:
                if item in items_before:
                    continue
                if canvas.type(item) == "text" and canvas.itemcget(item, "text") == top_text:
                    text_id = item
                    break

            if text_id and any(c in ("red", "blue") for _, c in segments):
                bbox = canvas.bbox(text_id)
                cx = (bbox[0] + bbox[2]) / 2
                cy = (bbox[1] + bbox[3]) / 2
                canvas.itemconfigure(text_id, state="hidden")
                draw_segment_text(canvas, cx, cy, segments, self.diagram_font)
            for ch in n.children:
                draw_node(ch)

        for r in new_roots:
            draw_connections(r)
            draw_node(r)

        existing_pairs = set()
        for p in node_objs.values():
            for ch in p.children:
                existing_pairs.add((p.unique_id, ch.unique_id))
        for (pid, cid), st in conn_status.items():
            if st != "removed":
                continue
            if (pid, cid) in existing_pairs:
                continue
            if pid in node_objs and cid in node_objs and pid in allow_ids and cid in allow_ids:
                parent = node_objs[pid]
                child = node_objs[cid]
                parent_pt = (parent.x, parent.y + 20)
                child_pt = (child.x, child.y - 25)
                if self.fta_drawing_helper:
                    self.fta_drawing_helper.draw_90_connection(canvas, parent_pt, child_pt, outline_color="red", line_width=1)

        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        ps = canvas.postscript(colormode="color", x=x, y=y, width=x2 - x, height=y2 - y)
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def metric_to_text(self, metric_type, value):
        if value is None:
            return "unknown"
        disc = AD_RiskAssessment_Helper.discretize_level(value)
        if metric_type == "maturity":
            return "high maturity" if disc == 5 else "low maturity" if disc == 1 else f"a maturity of {disc}"
        elif metric_type == "rigor":
            return "high rigor" if disc == 5 else "low rigor" if disc == 1 else f"a rigor of {disc}"
        elif metric_type == "severity":
            return "high severity" if disc >= 3 else "low severity" if disc == 1 else f"a severity of {disc}"
        else:
            return str(disc)

    def assurance_level_text(self, level):
        mapping = {1:"extra low",2:"low",3:"medium",4:"high",5:"high+"}
        return mapping.get(level, str(level))

    def calculate_cut_sets(self, node):
        if not node.children:
            return [{node.unique_id}]
        gate = (node.gate_type or "AND").upper() if node.node_type.upper() in ["TOP EVENT", "GATE", "RIGOR LEVEL"] else "AND"
        child_cut_sets = [self.calculate_cut_sets(child) for child in node.children]
        if gate == "OR":
            result = []
            for cuts in child_cut_sets:
                result.extend(cuts)
            return result
        elif gate == "AND":
            result = [set()]
            for cuts in child_cut_sets:
                temp = []
                for partial in result:
                    for cs in cuts:
                        temp.append(partial.union(cs))
                result = temp
            return result
        else:
            result = []
            for cuts in child_cut_sets:
                result.extend(cuts)
            return result

    def build_hierarchical_argumentation(self, node, indent=0):
        indent_str = "    " * indent
        node_name = node.user_name if node.user_name else f"Node {node.unique_id}"
        details = f"{node.node_type}"
        if node.input_subtype:
            details += f" ({node.input_subtype})"
        if node.description:
            details += f": {node.description}"
        metric_type = "maturity" if node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"] else "rigor"
        metric_descr = self.metric_to_text(metric_type, node.quant_value)
        line = f"{indent_str}- {node_name} ({details}) -> {metric_descr}"
        if node.rationale and node.node_type.upper() not in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            line += f" [Rationale: {node.rationale.strip()}]"
        child_lines = [self.build_hierarchical_argumentation(child, indent+1) for child in node.children]
        if child_lines:
            line += "\n" + "\n".join(child_lines)
        return line

    def build_hierarchical_argumentation_common(self, node, indent=0, described=None):
        if described is None:
            described = set()
        indent_str = "    " * indent
        node_name = node.user_name if node.user_name else f"Node {node.unique_id}"
        if node.unique_id not in described:
            details = f"{node.node_type}"
            if node.input_subtype:
                details += f" ({node.input_subtype})"
            if node.description:
                details += f": {node.description}"
            described.add(node.unique_id)
        else:
            details = f"{node.node_type} (see common cause: {node_name})"
        metric_type = "maturity" if node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"] else "rigor"
        metric_descr = self.metric_to_text(metric_type, node.quant_value)
        line = f"{indent_str}- {node_name} ({details}) -> {metric_descr}"
        if node.rationale and node.node_type.upper() not in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            line += f" [Rationale: {node.rationale.strip()}]"
        child_lines = [self.build_hierarchical_argumentation_common(child, indent+1, described) for child in node.children]
        if child_lines:
            line += "\n" + "\n".join(child_lines)
        return line

    def build_page_argumentation(self, page_node):
        return self.build_hierarchical_argumentation(page_node)

    def build_unified_recommendation_table(self):
        """
        Collect ALL nodes across ALL top-level events, group them by the
        recommendation(s) they trigger, and return a single LongTable.
        
        *Only primary nodes (originals) are used so that clones are not duplicated.
        Each node gets its own row, so large text can split across pages.
        """
        from reportlab.platypus import LongTable, Paragraph
        from reportlab.lib import colors
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import TableStyle

        style_sheet = getSampleStyleSheet()
        body_style = style_sheet["BodyText"]
        header_style = ParagraphStyle(
            name="RecHeader",
            parent=body_style,
            fontSize=5,
            leading=6,
            wordWrap='CJK',
            alignment=1
        )

        # 1) Gather ALL nodes from ALL top events.
        # Assumes get_all_nodes_in_model() is defined to merge all nodes.
        all_nodes = self.get_all_nodes_in_model()
        if not all_nodes:
            print("Debug: No nodes found in the entire model.")
            return None

        # 2) Filter out clones: only include primary instances.
        primary_nodes = [n for n in all_nodes if n.is_primary_instance]

        # 3) Build a mapping: recommendation text -> list of nodes that trigger it.
        rec_to_nodes = {}
        for node in primary_nodes:
            # Only consider nodes with a quant_value and a nonempty description.
            if node.quant_value is not None and node.description:
                discrete = AD_RiskAssessment_Helper.discretize_level(node.quant_value)
                extra_dict = dynamic_recommendations.get(discrete, {}).get("Extra Recommendations", {})
                desc_lower = node.description.lower()
                for keyword, rec_text in extra_dict.items():
                    if keyword.lower() in desc_lower:
                        rec_to_nodes.setdefault(rec_text, []).append(node)

        if not rec_to_nodes:
            print("Debug: No matching recommendations found for any node.")
            return None

        # 4) Build the table rows.
        # We use two columns: "Extra Recommendation" and "Metric Details"
        # For each recommendation, the first row shows the recommendation text and the details
        # for the first node; subsequent rows leave the recommendation column blank.
        data = [[
            Paragraph("<b>Extra Recommendation</b>", header_style),
            Paragraph("<b>Metric Details</b>", header_style)
        ]]

        for rec_text, nodes in rec_to_nodes.items():
            first_row = True
            for node in nodes:
                # Use the node's display_label if it does not fall back to "Node ..."; otherwise, use the quant_value.
                metric_str = (node.display_label 
                              if node.display_label and not node.display_label.startswith("Node")
                              else (f"{node.quant_value:.2f}" if node.quant_value is not None else "N/A"))
                desc = (node.description or "N/A").strip().replace("\n", "<br/>")
                rat = (node.rationale or "N/A").strip().replace("\n", "<br/>")
                node_details = (
                    f"{node.unique_id}: {node.name}"
                    f"<br/><b>Metric:</b> {metric_str}"
                    f"<br/><b>Desc:</b> {desc}"
                    f"<br/><b>Rationale:</b> {rat}"
                )
                if first_row:
                    data.append([
                        Paragraph(rec_text, body_style),
                        Paragraph(node_details, body_style)
                    ])
                    first_row = False
                else:
                    data.append([
                        "",
                        Paragraph(node_details, body_style)
                    ])

        # 5) Create and style the LongTable.
        col_widths = [200, 450]
        table = LongTable(data, colWidths=col_widths, repeatRows=1, splitByRow=True)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.orange),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        return table

    def build_base_events_table_html(self,root_node):
        """
        Traverse the fault tree starting from root_node and collect all base events (leaf nodes).
        Return an HTML string representing a table with columns:
        Node ID, Name, Score, Description, and Rationale.
        
        The table header cells are styled with an orange background (#FFCC99),
        and multiline descriptions/rationales are preserved by converting newlines to <br>.
        """
        base_events = []
        
        # Collect all leaf nodes (base events).
        def traverse(n):
            if not n.children:
                base_events.append(n)
            else:
                for child in n.children:
                    traverse(child)
        traverse(root_node)
        
        # Build the HTML table.
        html_lines = []
        html_lines.append('<table style="border-collapse: collapse; width: 100%;">')
        
        # Table header row with orange background.
        html_lines.append(
            '  <thead>'
            '    <tr style="background-color: #FFCC99;">'
            '      <th style="border: 1px solid #ccc; padding: 8px;">Node ID</th>'
            '      <th style="border: 1px solid #ccc; padding: 8px;">Name</th>'
            '      <th style="border: 1px solid #ccc; padding: 8px;">Score</th>'
            '      <th style="border: 1px solid #ccc; padding: 8px;">Description</th>'
            '      <th style="border: 1px solid #ccc; padding: 8px;">Rationale</th>'
            '    </tr>'
            '  </thead>'
        )
        html_lines.append('  <tbody>')
        
        for node in base_events:
            node_id = str(node.unique_id)
            name = node.name or f"Node {node.unique_id}"
            score = f"{node.quant_value:.2f}" if node.quant_value is not None else "N/A"
            
            # Convert newlines to <br> for multiline display.
            desc = node.description.strip().replace('\n', '<br>') if node.description else "N/A"
            rat = node.rationale.strip().replace('\n', '<br>') if node.rationale else "N/A"
            
            row_html = (
                '    <tr>'
                f'      <td style="border: 1px solid #ccc; padding: 8px; vertical-align: top;">{node_id}</td>'
                f'      <td style="border: 1px solid #ccc; padding: 8px; vertical-align: top;">{name}</td>'
                f'      <td style="border: 1px solid #ccc; padding: 8px; vertical-align: top;">{score}</td>'
                f'      <td style="border: 1px solid #ccc; padding: 8px; vertical-align: top;">{desc}</td>'
                f'      <td style="border: 1px solid #ccc; padding: 8px; vertical-align: top;">{rat}</td>'
                '    </tr>'
            )
            html_lines.append(row_html)
        
        html_lines.append('  </tbody>')
        html_lines.append('</table>')
        
        return "\n".join(html_lines)

    def build_argumentation(self, node):
        if not node.children:
            return ""
        header = ""
        if node.node_type.upper() == "TOP EVENT":
            disc = AD_RiskAssessment_Helper.discretize_level(node.quant_value)
            assurance_descr = self.assurance_level_text(disc)
            severity_str = f"{node.severity}" if node.severity is not None else "N/A"
            controllability_str = f"{node.controllability}" if node.controllability is not None else "N/A"
            header += (
                f"Prototype Assurance Level (PAL) Explanation:<br/>"
                f"Based on the aggregated scores of its child nodes, this top event has been assigned an Prototype Assurance Level (PAL) of <b>{assurance_descr}</b> "
                f"with a severity rating of <b>{severity_str}</b> and controllability <b>{controllability_str}</b>.<br/><br/>"
            )
            # Append the dynamically generated recommendations.
            header += self.generate_recommendations_for_top_event(node) + "<br/>"
        # Now generate the cut-set table.
        nodes_by_id = {}
        def map_nodes(n):
            nodes_by_id[n.unique_id] = n
            for child in n.children:
                map_nodes(child)
        map_nodes(node)
        cut_sets = self.calculate_cut_sets(node)
        cut_set_table = "Cut Set Table:<br/>"
        for i, cs in enumerate(cut_sets, start=1):
            cs_ids = ", ".join(f"Node {uid}" for uid in sorted(cs))
            cut_set_table += f"Cut Set {i}: {cs_ids}<br/>"
        node_definitions = "Node Definitions:<br/>"
        unique_ids = set()
        for cs in cut_sets:
            unique_ids.update(cs)
        for uid in sorted(unique_ids):
            n = nodes_by_id.get(uid)
            if n is None:
                continue
            subtype = n.input_subtype if n.input_subtype is not None else (
                VALID_SUBTYPES["Confidence"][0] if n.node_type.upper() == "CONFIDENCE LEVEL"
                else VALID_SUBTYPES.get("Prototype Assurance Level (PAL)", ["Default"])[0]
            )
            desc = n.description.strip() if n.description else "No description provided."
            node_definitions += f"Node {uid}: {n.name}<br/>"
            node_definitions += f"Type: {n.node_type}, Subtype: {subtype}<br/>"
            node_definitions += f"Description: {desc}<br/><br/>"
        diagram_note = "Cause-and-Effect Diagram is generated below.<br/>"
        return header + cut_set_table + "<br/>" + node_definitions + "<br/>" + diagram_note
       
    def auto_create_argumentation(self, node, suppress_top_event_recommendations=False):
        """
        Generate qualitative argumentation text for a given node.
        For a TOP EVENT (unless suppressed), include dynamic recommendations from the dictionary
        filtered by the node’s description. For non–top-level nodes, simply display the node's input score,
        description, and rationale.
        """
        level = AD_RiskAssessment_Helper.discretize_level(node.quant_value) if node.quant_value is not None else 1

        if node.node_type.upper() == "TOP EVENT" and not suppress_top_event_recommendations:
            assurance_descr = self.assurance_level_text(level)
            severity_str = f"{node.severity}" if node.severity is not None else "N/A"
            controllability_str = f"{node.controllability}" if node.controllability is not None else "N/A"
            header = (
                f"Prototype Assurance Level (PAL) Explanation:\n"
                f"This top event is assigned an Prototype Assurance Level (PAL) of '{assurance_descr}' with a severity rating of {severity_str} and controllability {controllability_str}.\n\n"
            )
            # Instead of showing all dynamic recommendations, select only those triggered by the description.
            rec_from_desc = self.get_recommendation_from_description(node.description, level)
            if rec_from_desc:
                base_arg = header + "Dynamic Recommendation:\n" + rec_from_desc
            else:
                # If no keyword found, show the full recommendations.
                rec = dynamic_recommendations.get(level, {})
                rec_lines = []
                for category in ["Testing Requirements", "IFTD Responsibilities", "Preventive Maintenance Actions", "Relevant AVSC Guidelines"]:
                    if category in rec:
                        rec_lines.append(f"{category}: {rec[category]}")
                rec_text = "\n".join(rec_lines)
                base_arg = header + "Recommendations:\n" + rec_text
        elif node.node_type.upper() == "TOP EVENT" and suppress_top_event_recommendations:
            base_arg = f"Top Event: Input score: {node.quant_value:.2f}" if node.quant_value is not None else "Top Event: No input score provided."
        else:
            base_arg = f"Input score: {node.quant_value:.2f}" if node.quant_value is not None else "No input score provided."

        own_text = ""
        if node.description:
            own_text += f"Description: {node.description}\n"
        if node.rationale:
            own_text += f"Rationale: {node.rationale}\n"
        if not own_text:
            own_text = "No additional details provided."
            
        return base_arg + "\n\n" + own_text

    def generate_argumentation_report(self, event):
        """
        Generate dynamic assurance-level argumentation for a top-level event.
        In this version, the event’s description is added at the very beginning,
        followed by the Prototype Assurance Level (PAL) explanation (including the rationale behind its severity)
        and the dynamic recommendations.
        """
        # Ensure a quant_value exists; default to 1.
        quant = event.quant_value if event.quant_value is not None else 1
        level = AD_RiskAssessment_Helper.discretize_level(quant)
        assurance_level = self.assurance_level_text(level)
        severity = event.severity if event.severity is not None else "N/A"
        controllability = event.controllability if event.controllability is not None else "N/A"
        
        # Get dynamic recommendations from the dictionary.
        rec = dynamic_recommendations.get(level, {})
        test_req = rec.get("Testing Requirements", "N/A")
        iftd_resp = rec.get("IFTD Responsibilities", "N/A")
        maint_act = rec.get("Preventive Maintenance Actions", "N/A")
        avsc_guid = rec.get("Relevant AVSC Guidelines", "N/A")
        
        # Get and clean up the top event’s description and rationale.
        top_description = event.description.strip() if event.description and event.description.strip() else "N/A"
        top_rationale = event.rationale.strip() if event.rationale and event.rationale.strip() else "N/A"
        
        text = (
            f"Description:<br/>{top_description}<br/><br/>"
            f"Prototype Assurance Level (PAL) Explanation:<br/>"
            f"This top event is assigned an Prototype Assurance Level (PAL) of <b>{assurance_level}</b> "
            f"with a severity rating of <b>{severity}</b> and controllability <b>{controllability}</b>.<br/>"
            f"Rationale for Severity: {top_rationale}<br/><br/>"
            #"Dynamic Recommendations:<br/>"
            #f"<b>Testing Requirements:</b> {test_req}<br/>"
            #f"<b>IFTD Responsibilities:</b> {iftd_resp}<br/>"
            #f"<b>Preventive Maintenance Actions:</b> {maint_act}<br/>"
            #f"<b>Relevant AVSC Guidelines:</b> {avsc_guid}<br/>"
        )
        return text

    def get_extra_recommendations_list(self, description, level):
        """
        Given a node's description and its Prototype Assurance Level (PAL), return a list of extra recommendations.
        This function iterates over all keys in the level's "Extra Recommendations" dictionary and
        collects the recommendation text for every keyword found in the description.
        """
        if not description:
            return []
        desc = description.lower()
        level_extras = dynamic_recommendations.get(level, {}).get("Extra Recommendations", {})
        rec_list = []
        for keyword, rec in level_extras.items():
            if keyword.lower() in desc:
                rec_list.append(rec)
        return rec_list

    def get_extra_recommendations_from_level(self,description, level):
        """
        Given a node's description and its Prototype Assurance Level (PAL) (1-5), look up keywords from the level's 
        "Extra Recommendations" in the dynamic_recommendations dictionary. If any keyword is found in the description
        (within a proximity of malfunction words), return the extra recommendations.
        """
        if not description:
            return ""
        desc = description.lower()
        level_extras = dynamic_recommendations.get(level, {}).get("Extra Recommendations", {})
        malfunction_words = ["unintended", "no", "not", "excessive", "incorrect"]
        
        recommendations = []
        for keyword, rec in level_extras.items():
            # Check if the keyword is present.
            if re.search(r'\b' + re.escape(keyword) + r'\b', desc):
                # Look for a malfunction word within 5 words of the keyword.
                pattern = r'\b' + re.escape(keyword) + r'\b(?:\W+\w+){0,5}\W+(?:' + "|".join(malfunction_words) + r')\b'
                if re.search(pattern, desc):
                    recommendations.append(rec)
        if recommendations:
            return "\nExtra Testing Recommendations:\n" + "\n".join(f"- {r}" for r in recommendations)
        return ""

    def get_recommendation_from_description(self, description, level):
        """
        Given a node's description and its Prototype Assurance Level (PAL), this function iterates over all keys 
        in the corresponding level's "Extra Recommendations" dictionary. It checks if each keyword 
        appears in the description (in a case-insensitive manner) and concatenates all matching recommendations.
        """
        if not description:
            return ""
        desc = description.lower()
        level_extras = dynamic_recommendations.get(level, {}).get("Extra Recommendations", {})
        rec_list = []
        for keyword, rec in level_extras.items():
            if keyword.lower() in desc:
                rec_list.append(rec)
        return " ".join(rec_list)
    
    def analyze_common_causes(self, node):
        occurrence = {}
        def traverse(n):
            if n.unique_id in occurrence:
                occurrence[n.unique_id]["count"] += 1
            else:
                occurrence[n.unique_id] = {"node": n, "count": 1}
            for child in n.children:
                traverse(child)
        traverse(node)
        report_lines = ["Common Cause Analysis:"]
        for uid, info in occurrence.items():
            if info["count"] > 1:
                n = info["node"]
                name = n.user_name if n.user_name else f"Node {n.unique_id}"
                report_lines.append(f" - {name} (Type: {n.node_type}) appears {info['count']} times. Description: {n.description or 'No description'}")
        if len(report_lines) == 1:
            report_lines.append(" None found.")
        return "\n".join(report_lines)

    def build_text_report(self, node, indent=0):
        report = "    " * indent + f"{node.name} ({node.node_type}"
        if node.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
            report += f", {node.gate_type}"
        report += ")"
        if node.display_label:
            report += f" => {node.display_label}"
        arg_text = self.build_argumentation(node)
        if arg_text:
            report += f"\n{'    ' * (indent+1)}Argumentation: {arg_text}"
        report += "\n\n"
        for child in node.children:
            report += self.build_text_report(child, indent+1)
        return report

    def all_children_are_base_events(self,node):
        """
        Return True if *every* child of 'node' is a base event 
        (i.e. Confidence Level or Robustness Score).
        """
        # If node has no children, we treat it as "False" (it’s effectively a leaf, not a gate).
        if not node.children:
            return False

        for child in node.children:
            t = child.node_type.upper()
            if t not in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
                return False
        return True

    def build_simplified_fta_model(self, top_event):
        """
        Build a simplified FTA model from the fault tree by including only the gate-level nodes.
        If a node is a GATE, RIGOR LEVEL, or TOP EVENT but all its children are base events,
        we will skip showing its gate_type.
        """
        nodes = []
        edges = []
        
        def traverse(node):
            node_type_up = node.node_type.upper()
            if node_type_up in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                final_gate_type = node.gate_type  # e.g. "AND" or "OR"

                node_info = {
                    "id": str(node.unique_id),
                    "label": node.name,
                    "gate_type": final_gate_type,   # store it only if not all children are base
                }
                if node.input_subtype:
                    node_info["subtype"] = node.input_subtype
                
                nodes.append(node_info)

                # Only traverse children that are also gates or top events
                for child in node.children:
                    child_type = child.node_type.upper()
                    if child_type in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                        edges.append({"source": str(node.unique_id), "target": str(child.unique_id)})
                        traverse(child)

        traverse(top_event)
        return {"nodes": nodes, "edges": edges}

    @staticmethod
    def auto_generate_fta_diagram(fta_model, output_path):
        """
        Generate a cause-and-effect diagram with a layered (hierarchical) layout,
        but draw the arrows in reverse (child -> parent).
        """
        import networkx as nx
        import matplotlib.pyplot as plt
        try:
            from adjustText import adjust_text
        except ImportError:
            adjust_text = None
        import numpy as np

        # --- 1) Build the directed graph (parent->child) ---
        G = nx.DiGraph()
        node_labels = {}
        node_colors = {}

        for node in fta_model["nodes"]:
            node_id = node["id"]
            label   = node.get("label", f"Node {node_id}")

            # If there's a gate_type, append it to the label
            gate_type = node.get("gate_type", "")
            if gate_type:
                # e.g. label = "Node 1\n(AND)"
                label += f"\n({gate_type.upper()})"

            G.add_node(node_id)
            node_labels[node_id] = label

            # Keep your color logic based on "subtype":
            subtype = node.get("subtype", "").lower()
            if "vehicle level function" in subtype:
                node_colors[node_id] = "lightcoral"
            elif "safety mechanism" in subtype:
                node_colors[node_id] = "lightyellow"
            elif "capability" in subtype:
                node_colors[node_id] = "lightblue"
            else:
                node_colors[node_id] = "white"  # clone

        # Add edges
        for edge in fta_model["edges"]:
            src = edge["source"]
            tgt = edge["target"]
            if not G.has_node(src) or not G.has_node(tgt):
                continue
            G.add_edge(src, tgt)

        # --- 2) Identify the top event as 'root' (layer 0) ---
        if fta_model["nodes"]:
            top_event_id = fta_model["nodes"][0]["id"]
        else:
            # If empty, just bail out
            plt.figure(figsize=(8,6))
            plt.title("No nodes to display")
            plt.savefig(output_path)
            plt.close()
            return

        # --- 3) BFS layering from top_event to find each node's layer ---
        layers = {}
        layers[top_event_id] = 0
        queue = [top_event_id]
        visited = set([top_event_id])

        while queue:
            current = queue.pop(0)
            current_layer = layers[current]
            for child in G.successors(current):
                if child not in visited:
                    visited.add(child)
                    layers[child] = current_layer + 1
                    queue.append(child)

        # Any node not reached gets placed in a higher layer
        max_layer = max(layers.values()) if layers else 0
        for n in G.nodes():
            if n not in layers:
                max_layer += 1
                layers[n] = max_layer

        # Group nodes by layer
        layer_dict = {}
        for node_id, layer in layers.items():
            layer_dict.setdefault(layer, []).append(node_id)

        # --- 4) Assign (x, y) by layer ---
        horizontal_gap = 2.0
        vertical_gap   = 1.0
        pos = {}

        for layer in sorted(layer_dict.keys()):
            node_list = layer_dict[layer]

            # Sort siblings by average parent index (optional)
            def avg_parent_position(n):
                parents = list(G.predecessors(n))
                if not parents:
                    return 0
                # we assume all parents are in a smaller layer
                return sum(layer_dict[layers[p]].index(p) for p in parents) / len(parents)

            node_list.sort(key=avg_parent_position)

            # Place them at x = layer*gap, y around 0
            middle = (len(node_list) - 1) / 2.0
            for i, n in enumerate(node_list):
                x = layer * horizontal_gap
                y = (i - middle) * vertical_gap
                pos[n] = (x, y)

        # --- 5) Light collision-avoidance pass (optional) ---
        def get_node_bbox(p, box_size=0.3):
            return (p[0] - box_size, p[1] - box_size, p[0] + box_size, p[1] + box_size)

        def bboxes_overlap(b1, b2):
            return not (b1[2] < b2[0] or b1[0] > b2[2] or b1[3] < b2[1] or b1[1] > b2[3])

        for _ in range(10):
            for n1 in G.nodes():
                for n2 in G.nodes():
                    if n1 == n2:
                        continue
                    b1 = get_node_bbox(pos[n1])
                    b2 = get_node_bbox(pos[n2])
                    if bboxes_overlap(b1, b2):
                        p1 = np.array(pos[n1])
                        p2 = np.array(pos[n2])
                        delta = p1 - p2
                        dist = np.linalg.norm(delta) + 1e-9
                        push = 0.02
                        shift = (delta/dist)*push
                        pos[n1] = tuple(p1 + shift)
                        pos[n2] = tuple(p2 - shift)

        # --- 6) Draw the diagram with REVERSED edges (child->parent) ---
        plt.figure(figsize=(12,8))

        # Build reversed edge list for drawing:
        reversed_edges = [(t, s) for (s, t) in G.edges()]

        nx.draw_networkx_edges(
            G, pos,
            edgelist=reversed_edges,    # <--- Inverted direction
            arrowstyle='-|>',
            arrowsize=20,
            edge_color='gray',
            connectionstyle="arc3,rad=0.0",
            min_source_margin=15,
            min_target_margin=15
        )

        # Keep your node colors from node_colors
        node_color_list = [node_colors[n] for n in G.nodes()]
        nx.draw_networkx_nodes(G, pos,
                               node_color=node_color_list,
                               node_size=1200,
                               edgecolors="black")

        # Draw labels
        text_items = []
        for n, (x, y) in pos.items():
            lbl = node_labels.get(n, str(n))
            txt = plt.text(x, y, lbl, fontsize=9, ha='center', va='center', wrap=True)
            text_items.append(txt)

        # Optionally adjust text to avoid overlap
        if adjust_text:
            adjust_text(text_items, arrowprops=dict(arrowstyle="-", color="gray"))

        plt.title("Cause and Effect Diagram")
        plt.axis('off')
        plt.tight_layout()
        plt.savefig(output_path, dpi=150)
        plt.close()

    def build_dynamic_recommendations_table(events, app):
        """
        (Optional) If you still want to have a compact table of per-event recommendations,
        this function builds a multiline LongTable with columns:
        [Event Name, Prototype Assurance Level (PAL), Severity, Controllability, Description, Rationale, Dynamic Recommendations].
        (Not used in the final report if you prefer only the consolidated argumentation.)
        """
        style_sheet = getSampleStyleSheet()
        header_style = ParagraphStyle(
            name="CompactHeader",
            parent=style_sheet["BodyText"],
            fontSize=4,
            leading=5,
            wordWrap='CJK',
            alignment=1
        )
        body_style = ParagraphStyle(
            name="CompactBody",
            parent=style_sheet["BodyText"],
            fontSize=5,
            leading=6,
            wordWrap='CJK'
        )
        
        data = [[
            Paragraph("<b>Event Name</b>", header_style),
            Paragraph("<b>Prototype Assurance Level (PAL)</b>", header_style),
            Paragraph("<b>Severity</b>", header_style),
            Paragraph("<b>Controllability</b>", header_style),
            Paragraph("<b>Description</b>", header_style),
            Paragraph("<b>Rationale</b>", header_style),
            Paragraph("<b>Recommendations</b>", header_style),
        ]]
        
        for event in events:
            event_name = event.name if event.name else f"Node {event.unique_id}"
            if event.quant_value is not None:
                disc_level = AD_RiskAssessment_Helper.discretize_level(event.quant_value)
                assurance_str = app.assurance_level_text(disc_level)
            else:
                assurance_str = "N/A"
            severity_str = str(event.severity) if event.severity is not None else "N/A"
            controllability_str = str(event.controllability) if event.controllability is not None else "N/A"
            desc_text = (event.description or "N/A").strip().replace("\n", "<br/>")
            rat_text = (event.rationale or "N/A").strip().replace("\n", "<br/>")
            rec_text = app.generate_argumentation_report(event)
            if isinstance(rec_text, list):
                rec_text = "\n".join(str(x) for x in rec_text)
            rec_text = rec_text.strip().replace("\n", "<br/>")
            data.append([
                Paragraph(event_name, body_style),
                Paragraph(assurance_str, body_style),
                Paragraph(severity_str, body_style),
                Paragraph(controllability_str, body_style),
                Paragraph(desc_text, body_style),
                Paragraph(rat_text, body_style),
                Paragraph(rec_text, body_style),
            ])
        
        col_widths = [100, 60, 40, 40, 150, 150, 200]
        table = LongTable(data, colWidths=col_widths, repeatRows=1, splitByRow=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.orange),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ROWHEIGHT', (0,0), (-1,0), 12),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('FONTSIZE', (0,1), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 1),
            ('RIGHTPADDING', (0,0), (-1,-1), 1),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('SPLITTABLE', (0,0), (-1,-1), True),
        ]))
        return table

    def get_all_nodes_no_filter(self,node):
        nodes = [node]
        for child in node.children:
            nodes.extend(self.get_all_nodes_no_filter(child))
        return nodes
        
    def derive_requirements_for_event(self, event):
        req_set = set()
        for node in self.get_all_nodes(event):
            if hasattr(node, "safety_requirements"):
                for req in node.safety_requirements:
                    req_set.add(f"[{req['id']}] [{req['req_type']}] {req['text']}")
        return req_set

    def get_combined_safety_requirements(self,node):
        """
        Returns a list of safety requirement dicts for the given node.
        If the node is a clone, it also combines the original node's safety_requirements.
        """
        req_list = []
        # Always take the node's own requirements if they exist.
        if hasattr(node, "safety_requirements") and node.safety_requirements:
            req_list.extend(node.safety_requirements)
        # If node is a clone, also add requirements from its original.
        if not node.is_primary_instance and hasattr(node, "original") and node.original.safety_requirements:
            req_list.extend(node.original.safety_requirements)
        return req_list

    def get_top_event(self, node):
        """
        Walk up the parent chain until a node whose node_type is 'TOP EVENT' is found.
        If none is found, return the node itself.
        """
        current = node
        while current.parents:
            for parent in current.parents:
                if parent.node_type.upper() == "TOP EVENT":
                    print(f"DEBUG: Found TOP EVENT for node {node.unique_id}: {parent.name}")
                    return parent
            current = current.parents[0]
        print(f"DEBUG: No TOP EVENT found for node {node.unique_id}; returning self")
        return node

    def aggregate_safety_requirements(self, node, all_nodes):
        aggregated = set()
        # Always add the node’s own safety requirements.
        for req in node.get("safety_requirements", []):
            aggregated.add(req["id"])
        
        # If this node is a clone, also add the original’s aggregated safety requirements.
        if node.get("original_id"):
            original = all_nodes.get(node["original_id"])
            if original:
                aggregated.update(self.aggregate_safety_requirements(original, all_nodes))
        
        # NEW: Also add safety requirements from the node’s immediate parents.
        for parent in node.get("parents", []):
            for req in parent.get("safety_requirements", []):
                aggregated.add(req["id"])
        
        # Recurse into children.
        for child in node.get("children", []):
            aggregated.update(self.aggregate_safety_requirements(child, all_nodes))
        
        node["aggregated_safety_requirements"] = sorted(aggregated)
        return aggregated

    def generate_top_event_summary(self, top_event):
        """
        Generates a structured, easy-to-read summary for a top-level event.
        
        It recursively collects all base nodes (nodes with type "CONFIDENCE LEVEL" or "ROBUSTNESS SCORE")
        from the event’s entire subtree (using originals for clones) and then constructs a multi-line summary
        that includes:
          - The top-level event name.
          - The required Prototype Assurance Level (PAL) (with numeric score) and the severity rating.
          - A bullet-point list of base nodes with their scores and rationales.
        """
        # Retrieve all nodes from the entire subtree (including originals for clones)
        all_nodes = self.get_all_nodes_no_filter(top_event)
        
        # Filter base nodes (confidence or robustness)
        base_nodes = [n for n in all_nodes if n.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]]
        
        # Build a bullet list for base nodes
        bullet_lines = []
        for bn in base_nodes:
            # Use the original's details for clones
            orig = bn if bn.is_primary_instance else bn.original
            identifier = orig.name if orig.name else f"Node {orig.unique_id}"
            score = f"{orig.quant_value:.2f}" if orig.quant_value is not None else "N/A"
            rationale = orig.rationale.strip() if orig.rationale and orig.rationale.strip() != "" else "No rationale provided"
            bullet_lines.append(f"• {identifier}: Score = {score}, Rationale: {rationale}")
        base_summary = "\n".join(bullet_lines) if bullet_lines else "No base nodes available."
        
        # Map overall assurance value to a descriptive level
        overall_assurance = top_event.quant_value if top_event.quant_value is not None else 1.0
        if overall_assurance >= 4.5:
            assurance_descr = "High+"
        elif overall_assurance >= 3.5:
            assurance_descr = "High"
        elif overall_assurance >= 2.5:
            assurance_descr = "Moderate"
        elif overall_assurance >= 1.5:
            assurance_descr = "Low"
        else:
            assurance_descr = "Very Low"
        
        # Use the top event's severity and controllability (defaults if missing)
        try:
            overall_severity = float(top_event.severity) if top_event.severity is not None else 3.0
        except Exception:
            overall_severity = 3.0
        try:
            overall_cont = float(top_event.controllability) if top_event.controllability is not None else 3.0
        except Exception:
            overall_cont = 3.0
        
        # Build the structured summary sentence
        summary_sentence = (
            f"Top-Level Event: {top_event.name}\n\n"
            f"Assurance Requirement:\n"
            f"  - Required Prototype Assurance Level (PAL): {assurance_descr} (Score: {overall_assurance:.2f})\n"
            f"  - Severity Rating: {overall_severity:.2f}\n"
            f"  - Controllability: {overall_cont:.2f}\n\n"
            f"Rationale:\n"
            f"  Based on analysis of its base nodes, the following factors contributed to this level:\n"
            f"{base_summary}"
        )
        return summary_sentence

    def _generate_pdf_report(self, include_assurance=True):
        report_title = self.project_properties.get("pdf_report_name", "Autonomous Driving Risk Assessment PDF Report")
        path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not path:
            return

        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                Paragraph,
                Spacer,
                PageBreak,
                SimpleDocTemplate,
                Image as RLImage,
                Table,
                TableStyle,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from io import BytesIO
            import PIL.Image as PILImage
        except ImportError:
            messagebox.showerror(
                "Report",
                "Reportlab package is required to generate PDF reports. "
                "Please install it and try again.",
            )
            return

        # Build a dictionary of all nodes (using each node’s to_dict())
        all_nodes = {}
        for node in self.get_all_nodes_in_model():
            node_dict = node.to_dict()
            all_nodes[node.unique_id] = node_dict

        # Now, for each node in the model, aggregate its safety requirements recursively.
        for node_dict in all_nodes.values():
            self.aggregate_safety_requirements(node_dict, all_nodes)

        # Define document with extra margins.
        doc = SimpleDocTemplate(
            path,
            pagesize=landscape(letter),
            leftMargin=0.8 * inch,
            rightMargin=0.8 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch
        )

        pdf_styles = getSampleStyleSheet()
        preformatted_style = ParagraphStyle(name="Preformatted", fontName="Courier", fontSize=10)
        pdf_styles.add(preformatted_style)
        Story = []

        # -------------------------------------------------------------
        # Executive Summary Page (First Page)
        # -------------------------------------------------------------
        Story.append(Paragraph(report_title, pdf_styles["Title"]))
        Story.append(Spacer(1, 12))

        if include_assurance:
            exec_summary_text = (
            "<b>Executive Summary: Manual Calculation of Prototype Assurance Level (PAL)</b><br/><br/>"
                "This document provides a step-by-step procedure to manually calculate the Prototype Assurance Level (PAL) for a subsystem in an "
                "autonomous system. The Prototype Assurance Level (PAL) is a single metric ranging from 1 to 5 (mapped to qualitative labels: "
                "Extra Low, Low, Moderate, High, High+). Follow these instructions using the provided tables.<br/><br/>"
                
                "<b>Calculation Instructions:</b><br/>"
                "1. <u>Base Assurance Derivation</u>:<br/>"
                " a. Assign a Confidence Level (CL) and a Robustness Score (RS) to the component, each on a scale from 1 (Extra Low) to 5 (High+).<br/>"
                " b. Using Table 1 (Base Assurance Inversion Matrix), locate the cell at the intersection of the CL (row) and RS (column).<br/>"
                "  For example, a CL of 1 and an RS of 1 yields a base assurance value of 5, indicating a very high requirement for additional safety measures.<br/><br/>"
                "2. <u>Combining Multiple Components</u>:<br/>"
                " a. If the subsystem consists of multiple components, first compute the base assurance value for each component individually as described above.<br/>"
                " b. Then, combine these values based on how the components interact:<br/>"
                "  - If the components must all perform reliably (an AND configuration), use a complement-product method as outlined in Table 3 (AND Decomposition Guidelines).<br/>"
                "  - If the components function as alternative options (an OR configuration), simply compute the average of their assurance values (see Table 4 for OR Decomposition Guidelines).<br/>"
                " c. When both types of inputs are present, average the base-derived values with the aggregated values to obtain a combined score.<br/><br/>"
                "3. <u>Severity Adjustment</u>:<br/>"
                " a. Adjust the combined assurance value to reflect hazard severity.<br/>"
                " b. For most subsystems, take the highest severity rating from the related elements and compute the average with the combined assurance score.<br/>"
                " c. For vehicle-level functions, use the formula: <br/>"
                "  Final Assurance = (Combined Value + Severity) / 2 <br/>"
                " Ensure the final score remains within the 1 to 5 range.<br/><br/>"
                "4. <u>Final Discretization</u>:<br/>"
                " a. Round the adjusted assurance value to the nearest 0.5.<br/>"
                " b. Refer to Table 2 (Output Discretization Mapping) to map the rounded value to one of the five discrete Prototype Assurance Levels (PAL), "
                "which correspond to the qualitative labels (Extra Low, Low, Moderate, High, High+).<br/><br/>"
                "By following these steps—deriving a base assurance from individual Confidence and Robustness ratings, combining multiple values "
                "through averaging or using complement-product methods (depending on the configuration), adjusting for hazard severity, and finally "
                "discretizing the result—you can manually calculate the Prototype Assurance Level (PAL) for any subsystem in a clear and systematic manner."
            )
            Story.append(Paragraph(exec_summary_text, pdf_styles["Normal"]))
            Story.append(Spacer(1, 12))
            
            # --- Table 1: Base Assurance Inversion Matrix ---
            header_style = ParagraphStyle(name="SafetyGoalsHeader", parent=pdf_styles["Normal"], fontSize=10, leading=12, alignment=1)
            base_matrix_data = [
                [Paragraph("<b>Robustness \\ Confidence</b>", header_style),
                 Paragraph("<b>1 (Extra Low)</b>", header_style),
                 Paragraph("<b>2 (Low)</b>", header_style),
                 Paragraph("<b>3 (Moderate)</b>", header_style),
                 Paragraph("<b>4 (High)</b>", header_style),
                 Paragraph("<b>5 (High+)</b>", header_style)],
                [Paragraph("<b>1 (Extra Low)</b>", header_style),
                 Paragraph("High+", pdf_styles["Normal"]),
                 Paragraph("High+", pdf_styles["Normal"]),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("High", pdf_styles["Normal"])],
                [Paragraph("<b>2 (Low)</b>", header_style),
                 Paragraph("High+", pdf_styles["Normal"]),
                 Paragraph("High+", pdf_styles["Normal"]),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"])],
                [Paragraph("<b>3 (Moderate)</b>", header_style),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"]),
                 Paragraph("Extra Low", pdf_styles["Normal"])],
                [Paragraph("<b>4 (High)</b>", header_style),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"]),
                 Paragraph("Extra Low", pdf_styles["Normal"]),
                 Paragraph("Extra Low", pdf_styles["Normal"])],
                [Paragraph("<b>5 (High+)</b>", header_style),
                 Paragraph("High", pdf_styles["Normal"]),
                 Paragraph("Moderate", pdf_styles["Normal"]),
                 Paragraph("Extra Low", pdf_styles["Normal"]),
                 Paragraph("Extra Low", pdf_styles["Normal"]),
                 Paragraph("Extra Low", pdf_styles["Normal"])]
            ]
            base_matrix_table = Table(base_matrix_data, colWidths=[80, 70, 70, 70, 70, 70])
            base_matrix_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
                ('BACKGROUND', (0,0), (0,-1), colors.lightblue),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTSIZE', (0,0), (-1,-1), 8)
            ]))
            Story.append(Paragraph("Table 1: Base Assurance Inversion Matrix", pdf_styles["Heading3"]))
            Story.append(Spacer(1, 6))
            Story.append(base_matrix_table)
            Story.append(Spacer(1, 12))
            
            # --- Table 2: Output Discretization Mapping ---
            discretization_data = [
                [Paragraph("<b>Continuous Value (Rounded)</b>", header_style),
                 Paragraph("<b>Prototype Assurance Level (PAL)</b>", header_style)],
                [Paragraph("< 1.5", header_style), Paragraph("Level 1 (Extra Low)", pdf_styles["Normal"])],
                [Paragraph("1.5 – < 2.5", header_style), Paragraph("Level 2 (Low)", pdf_styles["Normal"])],
                [Paragraph("2.5 – < 3.5", header_style), Paragraph("Level 3 (Moderate)", pdf_styles["Normal"])],
                [Paragraph("3.5 – < 4.5", header_style), Paragraph("Level 4 (High)", pdf_styles["Normal"])],
                [Paragraph("≥ 4.5", header_style), Paragraph("Level 5 (High+)", pdf_styles["Normal"])]
            ]
            discretization_table = Table(discretization_data, colWidths=[150, 200])
            discretization_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTSIZE', (0,0), (-1,-1), 8)
            ]))
            Story.append(Paragraph("Table 2: Output Discretization Mapping", pdf_styles["Heading3"]))
            Story.append(Spacer(1, 6))
            Story.append(discretization_table)
            Story.append(Spacer(1, 12))
            
            # Define mapping from numeric level to qualitative label.
            level_labels = {1: "Extra Low", 2: "Low", 3: "Moderate", 4: "High", 5: "High+"}
    
            # ------------------------------------------------------------------
            # Helper: Get the highest Prototype Assurance Level (PAL) from immediate parents.
            # For a given node (or its clone), this returns the maximum assurance (as an integer 1-5)
            # among all its immediate parents. If no parent exists, it returns the node's own assurance.
            def get_immediate_parent_assurance(node):
                if node.parents:
                    assurances = []
                    for p in node.parents:
                        # For clones, use the original parent's assurance value.
                        parent = p if p.is_primary_instance else p.original
                        try:
                            val = int(parent.quant_value)
                        except (TypeError, ValueError):
                            val = 1
                        assurances.append(val)
                    return max(assurances) if assurances else int(node.quant_value if node.quant_value is not None else 1)
                else:
                    return int(node.quant_value if node.quant_value is not None else 1)
            # ------------------------------------------------------------------
    
            # --- Safety Goals Summary Table ---
            safety_goals_data = []
            header_style = ParagraphStyle(name="SafetyGoalsHeader", parent=pdf_styles["Normal"], fontSize=10, leading=12, alignment=1)
            safety_goals_data.append([
                Paragraph("<b>Safety Goal</b>", header_style),
                Paragraph("<b>Highest Immediate Parent Assurance</b>", header_style),
                Paragraph("<b>Linked Recommendations</b>", header_style)
            ])
                    
            # Instead of iterating over only top-level events,
            # we iterate over all nodes that have safety requirements.
            grouped_by_linked = {}
            for node in self.get_all_nodes_in_model():
                if hasattr(node, "safety_requirements") and node.safety_requirements:
                    # Determine the safety goal from the node.
                    safety_goal = node.safety_goal_description.strip() if node.safety_goal_description.strip() != "" else node.name
                    # Get the highest assurance from its immediate parent(s)
                    parent_assur = get_immediate_parent_assurance(node)
                    assurance_str = f"Level {parent_assur} ({level_labels.get(parent_assur, 'N/A')})"
                    # Use the node's description to generate a linked recommendation.
                    # (You can adjust this method as needed.)
                    linked_rec = self.generate_recommendations_for_top_event(node)
                    extra_recs = self.get_extra_recommendations_list(node.description,
                                                                      AD_RiskAssessment_Helper.discretize_level(node.quant_value))
                    if not extra_recs:
                        extra_recs = ["No Extra Recommendation"]
                    # Group by the linked recommendation text.
                    grouped_by_linked.setdefault(linked_rec, {})
                    for extra in extra_recs:
                        grouped_by_linked[linked_rec].setdefault(extra, [])
                        grouped_by_linked[linked_rec][extra].append(f"- {safety_goal} (Assurance: {assurance_str})")
    
            sg_data = []
            sg_data.append([
                Paragraph("<b>Linked Recommendation</b>", header_style),
                Paragraph("<b>Safety Goals Grouped by Extra Recommendation</b>", header_style)
            ])
            for linked_rec, extra_groups in grouped_by_linked.items():
                nested_text = ""
                for extra_rec, goals in extra_groups.items():
                    nested_text += f"<b>{extra_rec}:</b><br/>" + "<br/>".join(goals) + "<br/><br/>"
                sg_data.append([
                    Paragraph(linked_rec, pdf_styles["Normal"]),
                    Paragraph(nested_text, pdf_styles["Normal"])
                ])
            if len(sg_data) > 1:
                sg_table = Table(sg_data, colWidths=[200, 400])
                sg_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('FONTSIZE', (0,0), (-1,-1), 10),
                    ('ALIGN', (0,0), (-1,0), 'CENTER')
                ]))
                Story.append(Paragraph("Safety Goals Summary:", pdf_styles["Heading2"]))
                Story.append(Spacer(1, 12))
                Story.append(sg_table)
                Story.append(Spacer(1, 12))
            Story.append(PageBreak())
            
        # --- Per-Top-Level-Event Content (Diagrams and Argumentation) ---
        def scale_image(pil_img):
            """Scale images so they fit within the doc page nicely."""
            orig_width, orig_height = pil_img.size
            scale_factor = 0.95 * min(doc.width / orig_width, doc.height / orig_height, 1)
            desired_width = orig_width * scale_factor
            desired_height = orig_height * scale_factor
            return desired_width, desired_height

        processed_ids = set()
        for idx, event in enumerate(self.top_events, start=1):
            if event.unique_id in processed_ids:
                continue
            processed_ids.add(event.unique_id)
            
            Story.append(Paragraph(f"Top-Level Event #{idx}: {event.name}", pdf_styles["Heading2"]))
            Story.append(Spacer(1, 12))
            
            # Argumentation text
            argumentation_text = self.generate_argumentation_report(event)
            if isinstance(argumentation_text, list):
                argumentation_text = "\n".join(str(x) for x in argumentation_text)
            argumentation_text = argumentation_text.replace("\n", "<br/>")
            Story.append(Paragraph(argumentation_text, preformatted_style))
            Story.append(Spacer(1, 12))

            # (A) "Detailed" event diagram (the subtree as captured in code)
            event_img = self.capture_event_diagram(event)
            if event_img is not None:
                buf = BytesIO()
                event_img.save(buf, format="PNG")
                buf.seek(0)
                desired_width, desired_height = scale_image(event_img)
                rl_img = RLImage(buf, width=desired_width, height=desired_height)
                Story.append(Paragraph("Detailed Diagram (Subtree):", pdf_styles["Heading3"]))
                Story.append(Spacer(1, 12))
                Story.append(rl_img)
                Story.append(Spacer(1, 12))

            # (B) "Cause and Effect" BFS diagram: build a single-root model for THIS event only
            fta_model = self.build_simplified_fta_model(event)
            temp_diagram_path = f"temp_fta_diagram_{idx}.png"
            self.auto_generate_fta_diagram(fta_model, temp_diagram_path)
            try:
                with open(temp_diagram_path, "rb") as img_file:
                    buf = BytesIO(img_file.read())
                pil_img = PILImage.open(buf)
                desired_width, desired_height = scale_image(pil_img)
                buf.seek(0)
                rl_img2 = RLImage(buf, width=desired_width, height=desired_height)
                Story.append(Paragraph("Cause and Effect Diagram (Single Root):", pdf_styles["Heading3"]))
                Story.append(Spacer(1, 12))
                Story.append(rl_img2)
                Story.append(Spacer(1, 12))
            except Exception as e:
                Story.append(Paragraph(f"Error generating BFS diagram for {event.name}: {e}", pdf_styles["Normal"]))
            Story.append(PageBreak())
        
        # --- Insert Page Diagrams (for 'page gates') ---
        unique_page_nodes = {}
        for evt in self.top_events:
            for pg in self.get_page_nodes(evt):
                if pg.is_primary_instance:
                    unique_page_nodes[pg.unique_id] = pg

        if unique_page_nodes:
            Story.append(Paragraph("Page Diagrams:", pdf_styles["Heading2"]))
            Story.append(Spacer(1, 12))

        for page_node in unique_page_nodes.values():
            page_img = self.capture_page_diagram(page_node)
            if page_img is not None:
                buf = BytesIO()
                page_img.save(buf, format="PNG")
                buf.seek(0)
                desired_width, desired_height = scale_image(page_img)
                rl_page_img = RLImage(buf, width=desired_width, height=desired_height)
                Story.append(Paragraph(f"Page Diagram for: {page_node.name}", pdf_styles["Heading3"]))
                Story.append(Spacer(1, 12))
                Story.append(rl_page_img)
                Story.append(Spacer(1, 12))
            else:
                Story.append(Paragraph("A page diagram could not be captured.", pdf_styles["Normal"]))
                Story.append(Spacer(1, 12))

        # --- FMEA Tables ---
        if self.fmeas:
            Story.append(PageBreak())
            Story.append(Paragraph("FMEA Tables", pdf_styles["Heading2"]))
            Story.append(Spacer(1, 12))
            for fmea in self.fmeas:
                Story.append(Paragraph(fmea['name'], pdf_styles["Heading3"]))
                data = [["Component", "Parent", "Failure Mode", "Failure Effect", "Cause", "S", "O", "D", "RPN", "Requirements"]]
                for be in fmea['entries']:
                    parent = be.parents[0] if be.parents else None
                    if parent:
                        comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                        parent_name = comp
                    else:
                        comp = getattr(be, "fmea_component", "") or "N/A"
                        parent_name = ""
                    req_ids = "; ".join([r.get("id") for r in getattr(be, 'safety_requirements', [])])
                    rpn = be.fmea_severity * be.fmea_occurrence * be.fmea_detection
                    failure_mode = be.description or (be.user_name or f"BE {be.unique_id}")
                    row = [comp, parent_name, failure_mode, be.fmea_effect, getattr(be, 'fmea_cause', ''), be.fmea_severity, be.fmea_occurrence, be.fmea_detection, rpn, req_ids]
                    data.append(row)
                table = Table(data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('FONTSIZE', (0,0), (-1,-1), 8)
                ]))
                Story.append(table)
                Story.append(Spacer(1, 12))

        # --- FTA-FMEA Traceability Matrix ---
        basic_events = [n for n in self.get_all_nodes(self.root_node) if n.node_type.upper() == "BASIC EVENT"]
        if basic_events:
            Story.append(PageBreak())
            Story.append(Paragraph("FTA-FMEA Traceability", pdf_styles["Heading2"]))
            data = [["Basic Event", "Component"]]
            for be in basic_events:
                parent = be.parents[0] if be.parents else None
                comp = parent.user_name if parent and parent.user_name else (f"Node {parent.unique_id}" if parent else "N/A")
                data.append([be.user_name or f"BE {be.unique_id}", comp])
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('FONTSIZE', (0,0), (-1,-1), 8)
            ]))
            Story.append(table)
            Story.append(Spacer(1, 12))

        # --- Final Build ---
        try:
            doc.build(Story)
        except Exception as e:
            messagebox.showerror("Report", f"Failed to generate PDF: {e}")
            return

        messagebox.showinfo(
            "Report",
            "PDF report generated!",
        )

    def generate_pdf_report(self):
        self._generate_pdf_report(include_assurance=True)

    def generate_pdf_without_assurance(self):
        """Generate a PDF report without the Prototype Assurance Level (PAL) pages."""
        self._generate_pdf_report(include_assurance=False)

    def capture_event_diagram(self, event_node):
        temp = tk.Toplevel(self.root)
        temp.withdraw()
        canvas = tk.Canvas(temp, bg="white", width=2000, height=2000)
        canvas.pack()
        self.draw_subtree_with_filter(canvas, event_node, self.get_all_nodes(event_node))
        canvas.delete("grid")
        canvas.update()
        bbox = canvas.bbox("all")
        if not bbox:
            temp.destroy()
            return None
        x, y, x2, y2 = bbox
        margin_left = 100
        margin_top  = 30
        new_x = x - margin_left
        new_y = y - margin_top
        new_width = (x2 - x) + 2 * margin_left
        new_height = (y2 - y) + 2 * margin_top
        ps = canvas.postscript(colormode="color", x=new_x, y=new_y, width=new_width, height=new_height)
        from io import BytesIO
        ps_bytes = BytesIO(ps.encode("utf-8"))
        try:
            img = Image.open(ps_bytes)
            img.load(scale=3)
        except Exception:
            img = None
        temp.destroy()
        return img.convert("RGB") if img else None

    def draw_subtree_with_filter(self, canvas, root_event, visible_nodes):
        self.draw_connections_subtree(canvas, root_event, set())
        for n in visible_nodes:
            self.draw_node_on_canvas_pdf(canvas, n)

    def draw_subtree(self, canvas, root_event):
        canvas.delete("all")
        self.draw_connections_subtree(canvas, root_event, set())
        for n in self.get_all_nodes(root_event):
            self.draw_node_on_canvas(canvas, n)
        canvas.config(scrollregion=canvas.bbox("all"))

    def draw_connections_subtree(self, canvas, node, drawn_ids):
        if id(node) in drawn_ids:
            return
        drawn_ids.add(id(node))
        if node.is_page and node.node_type.upper() != "TOP EVENT":
            return
        region_width = 100 * self.zoom
        parent_bottom = (node.x * self.zoom, node.y * self.zoom + 40 * self.zoom)
        N = len(node.children)
        for i, child in enumerate(node.children):
            parent_conn = (node.x * self.zoom - region_width/2 + (i+0.5)*(region_width/N), parent_bottom[1])
            child_top = (child.x * self.zoom, child.y * self.zoom - 45 * self.zoom)
            # Call the helper’s method instead of a global function.
            fta_drawing_helper.draw_90_connection(canvas, parent_conn, child_top,
                                                  outline_color="dimgray", line_width=1)
        for child in node.children:
            self.draw_connections_subtree(canvas, child, drawn_ids)
            
    def draw_node_on_canvas_pdf(self, canvas, node):
        # For cloned nodes, use the original's values.
        if not node.is_primary_instance and hasattr(node, "original") and node.original:
            base_label = node.original.display_label
            subtype = node.original.input_subtype or "N/A"
            equation_text = node.original.equation
            detailed_eq = node.original.detailed_equation
        else:
            base_label = node.display_label
            subtype = node.input_subtype or "N/A"
            equation_text = node.equation
            detailed_eq = node.detailed_equation

        # Extract the score type from the base label.
        # For example, if the base label is "Required Rigor [4]", score_type becomes "Required Rigor".
        score_type = base_label.split('[')[0].strip()

        fill_color = self.get_node_fill_color(node)
        eff_x = node.x * self.zoom
        eff_y = node.y * self.zoom

        # Decide what to show in the top text based on the configuration.
        if self.project_properties.get("pdf_detailed_formulas", True):
            # Detailed mode: show the score type and the node description.
            top_text = (f"Type: {node.node_type}\n"
                        f"Score: {score_type}\n"
                        f"Subtype: {subtype}\n"
                        f"Desc: {node.description}")
        else:
            # Score-only mode: show the discretized metric (as an integer, without decimals).
            if node.quant_value is not None:
                # Convert quant_value to float and discretize
                score_value = float(node.quant_value)
                discrete = AD_RiskAssessment_Helper.discretize_level(score_value)
            else:
                discrete = "N/A"
            top_text = (f"Type: {node.node_type}\n"
                        f"{score_type} = {discrete}\n"
                        f"Subtype: {subtype}")

        bottom_text = node.name
        node_type_upper = node.node_type.upper()

        if node.is_page:
            fta_drawing_helper.draw_triangle_shape(canvas, eff_x, eff_y, scale=40 * self.zoom,
                                                   top_text=top_text,
                                                   bottom_text=bottom_text,
                                                   fill=fill_color,
                                                   outline_color="dimgray",
                                                   line_width=1,
                                                   font_obj=self.diagram_font)
        elif node_type_upper in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45 * self.zoom,
                                                       top_text=top_text,
                                                       bottom_text=bottom_text,
                                                       fill=fill_color,
                                                       outline_color="dimgray",
                                                       line_width=1,
                                                       font_obj=self.diagram_font)
        elif node_type_upper in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
            if node.gate_type.upper() == "OR":
                fta_drawing_helper.draw_rotated_or_gate_clone_shape(canvas, eff_x, eff_y,
                                                                    scale=40 * self.zoom,
                                                                    top_text=top_text,
                                                                    bottom_text=bottom_text,
                                                                    fill=fill_color,
                                                                    outline_color="dimgray",
                                                                    line_width=1,
                                                                    font_obj=self.diagram_font)
            else:
                fta_drawing_helper.draw_rotated_and_gate_clone_shape(canvas, eff_x, eff_y,
                                                                     scale=40 * self.zoom,
                                                                     top_text=top_text,
                                                                     bottom_text=bottom_text,
                                                                     fill=fill_color,
                                                                     outline_color="dimgray",
                                                                     line_width=1,
                                                                     font_obj=self.diagram_font)
        else:
            fta_drawing_helper.draw_circle_event_shape(canvas, eff_x, eff_y, 45 * self.zoom,
                                                       top_text=top_text,
                                                       bottom_text=bottom_text,
                                                       fill=fill_color,
                                                       outline_color="dimgray",
                                                       line_width=1,
                                                       font_obj=self.diagram_font)

        # In detailed mode, also draw the equations.
        if self.project_properties.get("pdf_detailed_formulas", True):
            canvas.create_text(eff_x - 80 * self.zoom, eff_y - 15 * self.zoom,
                               text=equation_text, anchor="e", fill="gray",
                               font=self.diagram_font)
            canvas.create_text(eff_x - 80 * self.zoom, eff_y + 15 * self.zoom,
                               text=detailed_eq, anchor="e", fill="gray",
                               font=self.diagram_font)

    def save_diagram_png(self):
        margin = 50
        all_nodes = self.get_all_nodes(self.root_node)
        if not all_nodes:
            messagebox.showerror("Error", "No nodes to export.")
            return
        min_x = min(n.x for n in all_nodes) - margin
        min_y = min(n.y for n in all_nodes) - margin
        max_x = max(n.x for n in all_nodes) + margin
        max_y = max(n.y for n in all_nodes) + margin
        scale_factor = 4
        width = int((max_x - min_x) * scale_factor)
        height = int((max_y - min_y) * scale_factor)
        img = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(img)
        grid_size = self.grid_size
        for x in range(0, int(max_x - min_x) + 1, grid_size):
            x_pos = int(x * scale_factor)
            draw.line([(x_pos, 0), (x_pos, height)], fill="#ddd")
        for y in range(0, int(max_y - min_y) + 1, grid_size):
            y_pos = int(y * scale_factor)
            draw.line([(0, y_pos), (width, y_pos)], fill="#ddd")
        try:
            font = ImageFont.truetype("arial.ttf", 10 * scale_factor)
        except IOError:
            font = ImageFont.load_default()
        for node in all_nodes:
            eff_x = int((node.x - min_x) * scale_factor)
            eff_y = int((node.y - min_y) * scale_factor)
            radius = int(45 * scale_factor)
            bbox = [eff_x - radius, eff_y - radius, eff_x + radius, eff_y + radius]
            node_color = self.get_node_fill_color(node)
            draw.ellipse(bbox, outline="dimgray", fill=node_color)
            text = node.name
            text_size = draw.textsize(text, font=font)
            text_x = eff_x - text_size[0] // 2
            text_y = eff_y - text_size[1] // 2
            draw.text((text_x, text_y), text, fill="black", font=font)
        file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 filetypes=[("PNG files", "*.png")])
        if file_path:
            try:
                img.save(file_path, "PNG")
                messagebox.showinfo("Saved", "High-resolution diagram exported as PNG.")
            except Exception as e:
                messagebox.showerror("Save Error", f"An error occurred: {e}")

    def on_treeview_click(self, event):
        sel = self.treeview.selection()
        if not sel:
            return
        try:
            node_id = int(self.treeview.item(sel[0], "tags")[0])
        except (IndexError, ValueError):
            return
        node = self.find_node_by_id_all(node_id)
        if node:
            self.open_page_diagram(node)

    def on_ctrl_mousewheel(self, event):
        if event.delta > 0:
            self.zoom_in()
        else:
            self.zoom_out()

    def new_model(self):
        if not messagebox.askyesno("New Model", "This will close the current project and start a new one. Continue?"):
            return

        AD_RiskAssessment_Helper.unique_node_id_counter = 1
        self.zoom = 1.0
        self.diagram_font.config(size=int(8 * self.zoom))
        self.scenario_libraries = []
        self.odd_libraries = []
        self.update_odd_elements()

        # Close any open page diagrams.
        if hasattr(self, "page_diagram") and self.page_diagram is not None:
            self.close_page_diagram()

        # Clear the tree view.
        self.treeview.delete(*self.treeview.get_children())

        # Destroy the old canvas_frame and re-create it.
        if self.canvas_frame:
            self.canvas_frame.destroy()
        self.canvas_frame = ttk.Frame(self.main_pane)
        self.main_pane.add(self.canvas_frame, stretch="always")

        # Use grid (instead of pack) to add the canvas and scrollbars.
        self.canvas = tk.Canvas(self.canvas_frame, bg="white")
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.hbar = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.hbar.grid(row=1, column=0, sticky="ew")
        self.vbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.vbar.grid(row=0, column=1, sticky="ns")
        self.canvas.config(xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set,
                           scrollregion=(0, 0, 2000, 2000))
        
        # Configure grid weights so that the canvas expands.
        self.canvas_frame.rowconfigure(0, weight=1)
        self.canvas_frame.columnconfigure(0, weight=1)

        # Rebind canvas events.
        self.canvas.bind("<ButtonPress-3>", self.on_right_mouse_press)
        self.canvas.bind("<B3-Motion>", self.on_right_mouse_drag)
        self.canvas.bind("<ButtonRelease-3>", self.show_context_menu)
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
        self.canvas.bind("<Control-MouseWheel>", self.on_ctrl_mousewheel)

        global unique_node_id_counter
        unique_node_id_counter = 1
        self.top_events = []
        self.root_node = None
        self.selected_node = None

        new_root = FaultTreeNode("Vehicle Level Function", "TOP EVENT")
        new_root.x, new_root.y = 300, 200
        self.top_events.append(new_root)
        self.root_node = new_root
        self.fmea_entries = []
        self.fmeas = []
        self.update_views()
        self.set_last_saved_state()
        self.canvas.update()

    def compute_occurrence_counts(self):
        counts = {}
        visited = set()
        def rec(node):
            if node.unique_id in visited:
                counts[node.unique_id] += 1
            else:
                visited.add(node.unique_id)
                counts[node.unique_id] = 1
            for child in node.children:
                rec(child)
        rec(self.root_node)
        return counts

    def get_node_fill_color(self, node):
        # Use the original node's properties for clones.
        base_node = node if node.is_primary_instance else node.original
        if self.project_properties.get("black_white", False):
            return "white"
        label = base_node.display_label  # use original's display label
        if "Prototype Assurance Level (PAL)" in label:
            base_type = "Prototype Assurance Level (PAL)"
        elif "Maturity" in label:
            base_type = "Maturity"
        elif "Rigor" in label:
            base_type = "Rigor"
        elif "Confidence" in label:
            base_type = "Confidence"
        elif "Robustness" in label:
            base_type = "Robustness"
        else:
            base_type = "Other"
        subtype = base_node.input_subtype if base_node.input_subtype else "Default"
        color_mapping = {
            "Confidence": {"Function": "lightpink", "Human Task": "lightgreen", "Default": "lightpink"},
            "Robustness": {"Function": "orange", "Human Task": "pink", "Default": "orange"},
            "Maturity": {"Functionality": "lightyellow", "Default": "lightyellow"},
            "Rigor": {"Capability": "turquoise", "Safety Mechanism": "yellow", "Default": "turquoise"},
            "Prototype Assurance Level (PAL)": {"Vehicle Level Function": "pink", "Functionality": "lightyellow","Capability": "turquoise", "Safety Mechanism": "yellow"},
            "Other": {"Default": "lightblue"}
        }
        return color_mapping.get(base_type, {}).get(subtype, color_mapping.get(base_type, {}).get("Default", "lightblue"))

    def on_right_mouse_press(self, event):
        self.canvas.scan_mark(event.x, event.y)

    def on_right_mouse_drag(self, event):
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def show_context_menu(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        clicked_node = None
        for n in self.get_all_nodes(self.root_node):
            radius = 60 if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else 45
            if (x - n.x)**2 + (y - n.y)**2 < radius**2:
                clicked_node = n
                break
        if not clicked_node:
            return
        self.selected_node = clicked_node
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label="Edit", command=lambda: self.edit_selected())
        menu.add_command(label="Remove Connection", command=lambda: self.remove_connection(clicked_node))
        menu.add_command(label="Delete Node", command=lambda: self.delete_node_and_subtree(clicked_node))
        menu.add_command(label="Remove Node", command=lambda: self.remove_node())
        menu.add_command(label="Copy", command=lambda: self.copy_node())
        menu.add_command(label="Cut", command=lambda: self.cut_node())
        menu.add_command(label="Paste", command=lambda: self.paste_node())
        menu.add_separator()
        menu.add_command(label="Edit User Name", command=lambda: self.edit_user_name())
        menu.add_command(label="Edit Description", command=lambda: self.edit_description())
        menu.add_command(label="Edit Rationale", command=lambda: self.edit_rationale())
        menu.add_command(label="Edit Value", command=lambda: self.edit_value())
        menu.add_command(label="Edit Gate Type", command=lambda: self.edit_gate_type())
        menu.add_command(label="Edit Severity", command=lambda: self.edit_severity())
        menu.add_command(label="Edit Controllability", command=lambda: self.edit_controllability())
        menu.add_command(label="Edit Page Flag", command=lambda: self.edit_page_flag())
        menu.add_separator()
        menu.add_command(label="Add Confidence", command=lambda: self.add_node_of_type("Confidence Level"))
        menu.add_command(label="Add Robustness", command=lambda: self.add_node_of_type("Robustness Score"))
        menu.add_command(label="Add Gate", command=lambda: self.add_node_of_type("GATE"))
        menu.tk_popup(event.x_root, event.y_root)

    def on_canvas_click(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        clicked_node = None
        for n in self.get_all_nodes(self.root_node):
            radius = 60 if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else 45
            if (x - n.x)**2 + (y - n.y)**2 < radius**2:
                clicked_node = n
                break
        self.selected_node = clicked_node
        if clicked_node:
            self.dragging_node = clicked_node
            self.drag_offset_x = x - clicked_node.x
            self.drag_offset_y = y - clicked_node.y
        else:
            self.dragging_node = None
        self.redraw_canvas()

    def on_canvas_double_click(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        clicked_node = None
        for n in self.get_all_nodes(self.root_node):
            radius = 60 if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else 45
            if (x - n.x)**2 + (y - n.y)**2 < radius**2:
                clicked_node = n
                break
        if clicked_node:
            if not clicked_node.is_primary_instance:
                self.open_page_diagram(getattr(clicked_node, "original", clicked_node))
            else:
                if clicked_node.is_page:
                    self.open_page_diagram(clicked_node)
                else:
                    EditNodeDialog(self.root, clicked_node, self)
            self.update_views()

    def on_canvas_drag(self, event):
        if self.dragging_node:
            x = self.canvas.canvasx(event.x) / self.zoom
            y = self.canvas.canvasy(event.y) / self.zoom
            new_x = x - self.drag_offset_x
            new_y = y - self.drag_offset_y
            dx = new_x - self.dragging_node.x
            dy = new_y - self.dragging_node.y
            self.dragging_node.x = new_x
            self.dragging_node.y = new_y
            if self.dragging_node.is_primary_instance:
                self.move_subtree(self.dragging_node, dx, dy)
            self.sync_nodes_by_id(self.dragging_node)
            self.redraw_canvas()

    def on_canvas_release(self, event):
        if self.dragging_node:
            self.dragging_node.x = round(self.dragging_node.x/self.grid_size)*self.grid_size
            self.dragging_node.y = round(self.dragging_node.y/self.grid_size)*self.grid_size
        self.dragging_node = None
        self.drag_offset_x = 0
        self.drag_offset_y = 0

    def move_subtree(self, node, dx, dy):
        for child in node.children:
            child.x += dx
            child.y += dy
            self.move_subtree(child, dx, dy)

    def zoom_in(self):
        self.zoom *= 1.2
        self.diagram_font.config(size=int(8 * self.zoom))
        self.redraw_canvas()

    def zoom_out(self):
        self.zoom /= 1.2
        self.diagram_font.config(size=int(8 * self.zoom))
        self.redraw_canvas()

    def auto_arrange(self):
        horizontal_gap = 150
        vertical_gap = 100
        next_y = [100]
        def layout(node, depth):
            node.x = depth * horizontal_gap + 100
            if not node.children:
                node.y = next_y[0]
                next_y[0] += vertical_gap
            else:
                for child in node.children:
                    layout(child, depth+1)
                node.y = (node.children[0].y + node.children[-1].y) / 2
        layout(self.root_node, 0)
        self.update_views()

    def get_all_nodes_table(self,root_node):
        """
        Recursively traverse the entire fault tree starting from root_node without any filtering.
        Returns a list of all nodes.
        """
        collector = []
        def rec(n):
            collector.append(n)
            for child in n.children:
                rec(child)
        rec(root_node)
        return collector

    def get_all_nodes_in_model(self):
        """
        Return a list of *all* nodes across *all* top-level events in self.top_events.
        """
        all_nodes = []
        for te in self.top_events:
            nodes = self.get_all_nodes_table(te)  # your existing method for one root
            all_nodes.extend(nodes)
        return all_nodes

    def get_all_basic_events(self):
        """Return a list of all basic events across all top-level trees."""
        return [n for n in self.get_all_nodes_in_model() if n.node_type.upper() == "BASIC EVENT"]

    def get_all_triggering_conditions(self):
        """Return all triggering condition nodes."""
        return [n for n in self.get_all_nodes_in_model() if n.node_type.upper() == "TRIGGERING CONDITION"]

    def get_all_functional_insufficiencies(self):
        """Return all functional insufficiency nodes."""
        return [n for n in self.get_all_nodes_in_model() if n.node_type.upper() == "FUNCTIONAL INSUFFICIENCY"]

    def get_all_scenario_names(self):
        """Return the list of scenario names from all scenario libraries."""
        names = []
        for lib in self.scenario_libraries:
            for sc in lib.get("scenarios", []):
                if isinstance(sc, dict):
                    name = sc.get("name", "")
                else:
                    name = sc
                if name:
                    names.append(name)
        return names

    def classify_scenarios(self):
        """Return two lists of scenario names grouped by category."""
        use_case = []
        sotif = []
        for lib in self.scenario_libraries:
            for sc in lib.get("scenarios", []):
                if isinstance(sc, dict):
                    name = sc.get("name", "")
                    if sc.get("tcs") or sc.get("fis") or sc.get("tc") or sc.get("fi") or sc.get("type") == "sotif":
                        sotif.append(name)
                    else:
                        use_case.append(name)
                else:
                    use_case.append(sc)
        return {"use_case": use_case, "sotif": sotif}

    def get_all_scenery_names(self):
        """Return the list of scenery/ODD element names."""
        names = []
        for lib in self.odd_libraries:
            for el in lib.get("elements", []):
                if isinstance(el, dict):
                    name = el.get("name") or el.get("element") or el.get("id")
                else:
                    name = str(el)
                if name:
                    names.append(name)
        return names

    def get_all_scenery_names(self):
        """Return the list of scenery/ODD element names."""
        names = []
        for lib in self.odd_libraries:
            for el in lib.get("elements", []):
                if isinstance(el, dict):
                    name = el.get("name") or el.get("element") or el.get("id")
                else:
                    name = str(el)
                if name:
                    names.append(name)
        return names

    def get_all_function_names(self):
        """Return unique function names from HAZOP entries."""
        names = set()
        for doc in getattr(self, "hazop_docs", []):
            for e in doc.entries:
                if getattr(e, "function", ""):
                    names.add(e.function)
        return sorted(names)

    def get_all_component_names(self):
        """Return unique component names from HAZOP and reliability analyses."""
        names = set()
        for doc in getattr(self, "hazop_docs", []):
            names.update(e.component for e in doc.entries if getattr(e, "component", ""))
        names.update(c.name for c in self.reliability_components)
        return sorted(n for n in names if n)

    def get_all_malfunction_names(self):
        """Return unique malfunction names from HAZOP entries."""
        names = set()
        for doc in getattr(self, "hazop_docs", []):
            names.update(e.malfunction for e in doc.entries if getattr(e, "malfunction", ""))
        return sorted(names)

    def update_odd_elements(self):
        """Aggregate elements from all ODD libraries into odd_elements list."""
        self.odd_elements = []
        for lib in self.odd_libraries:
            self.odd_elements.extend(lib.get("elements", []))

    def get_all_failure_modes(self):
        """Return list of all failure mode nodes from FTA, FMEAs and FMEDAs."""
        modes = list(self.get_all_basic_events())
        for doc in self.fmea_entries:
            modes.append(doc)
        for f in self.fmeas:
            modes.extend(f.get("entries", []))
        for d in self.fmedas:
            modes.extend(d.get("entries", []))
        unique = {}
        for m in modes:
            unique[getattr(m, "unique_id", id(m))] = m
        return list(unique.values())

    def get_failure_mode_node(self, node):
        ref = getattr(node, "failure_mode_ref", None)
        if ref:
            n = self.find_node_by_id_all(ref)
            if n:
                return n
        return node

    def format_failure_mode_label(self, node):
        comp = node.parents[0].user_name if node.parents else getattr(node, "fmea_component", "")
        label = node.description or (node.user_name or f"BE {node.unique_id}")
        return f"{comp}: {label}" if comp else label


    def get_all_nodes(self, node=None):
        if node is None:
            result = []
            for te in self.top_events:
                result.extend(self.get_all_nodes(te))
            return result

        visited = set()
        def rec(n):
            if n.unique_id in visited:
                return []
            visited.add(n.unique_id)
            # ---- Remove or comment out any code that returns [] if n is a page or if a parent is a page
            if n != self.root_node and any(parent.is_page for parent in n.parents):
                return []

            result = [n]
            for c in n.children:
                result.extend(rec(c))
            return result

        return rec(node)

    def update_views(self):
        self.treeview.delete(*self.treeview.get_children())
        for top_event in self.top_events:
            self.insert_node_in_tree("", top_event)
        # NEW: Compute the occurrence counts from the current tree:
        self.occurrence_counts = self.compute_occurrence_counts()

        if hasattr(self, "page_diagram") and self.page_diagram is not None:
            if self.page_diagram.canvas.winfo_exists():
                self.page_diagram.redraw_canvas()
            else:
                self.page_diagram = None
        elif hasattr(self, "canvas") and self.canvas.winfo_exists():
            if self.selected_node is not None:
                self.redraw_canvas()
            else:
                self.canvas.delete("all")

    def update_basic_event_probabilities(self):
        """Update failure probabilities for all basic events.

        The calculation uses the selected probability formula on each
        event or its associated failure mode. The FIT rate of the failure
        mode is converted to a failure rate in events per hour, then the
        probability is derived for the mission profile time ``tau``.
        """
        for be in self.get_all_basic_events():
            be.failure_prob = self.compute_failure_prob(be)

    def compute_failure_prob(self, node, failure_mode_ref=None, formula=None):
        """Return probability of failure for ``node`` based on FIT rate.

        When the constant formula is selected the ``failure_prob`` value
        stored on the node is returned directly so users can specify an
        arbitrary probability.
        """
        tau = 1.0
        if self.mission_profiles:
            tau = self.mission_profiles[0].tau
        if tau <= 0:
            tau = 1.0
        fm = self.find_node_by_id_all(failure_mode_ref) if failure_mode_ref else self.get_failure_mode_node(node)
        fit = getattr(fm, "fmeda_fit", getattr(node, "fmeda_fit", 0.0))
        t = tau
        formula = formula or getattr(node, "prob_formula", getattr(fm, "prob_formula", "linear"))
        f = str(formula).strip().lower()
        if f == "constant":
            try:
                return float(getattr(node, "failure_prob", 0.0))
            except (TypeError, ValueError):
                return 0.0
        if fit <= 0:
            return 0.0
        lam = fit / 1e9
        if f == "exponential":
            return 1 - math.exp(-lam * t)
        else:
            return lam * t

    def propagate_failure_mode_attributes(self, fm_node):
        """Update basic events referencing ``fm_node`` and recompute probability."""
        for be in self.get_all_basic_events():
            if getattr(be, "failure_mode_ref", None) == fm_node.unique_id:
                be.fmeda_fit = fm_node.fmeda_fit
                be.fmeda_diag_cov = fm_node.fmeda_diag_cov
                # Always propagate the formula so edits take effect
                be.prob_formula = fm_node.prob_formula
                be.failure_prob = self.compute_failure_prob(be)

    def insert_node_in_tree(self, parent_item, node):
        # If the node has no parent (i.e. it's a top-level event), display it.
        if not node.parents or node.node_type.upper() == "TOP EVENT" or node.is_page:
            txt = node.name
            item_id = self.treeview.insert(parent_item, "end", text=txt, open=True, tags=(str(node.unique_id),))
            # Recursively insert all children regardless of their type.
            for child in node.children:
                self.insert_node_in_tree(item_id, child)
        else:
            # If the node is not top-level, still check its children.
            for child in node.children:
                self.insert_node_in_tree(parent_item, child)

    def redraw_canvas(self):
        if not hasattr(self, "canvas") or not self.canvas.winfo_exists():
            return
        self.canvas.delete("all")
        self.draw_grid()
        drawn_ids = set()
        for top_event in self.top_events:
            self.draw_connections(top_event, drawn_ids)
        all_nodes = []
        for top_event in self.top_events:
            all_nodes.extend(self.get_all_nodes(top_event))
        for node in all_nodes:
            self.draw_node(node)
        self.canvas.config(scrollregion=self.canvas.bbox("all"))

    def draw_grid(self):
        if not self.project_properties.get("show_grid", True):
            return
        spacing = self.grid_size * self.zoom
        width = self.canvas.winfo_width()
        height = self.canvas.winfo_height()
        if width < 10:
            width = 800
        if height < 10:
            height = 600
        for x in range(0, int(width), int(spacing)):
            self.canvas.create_line(x, 0, x, height, fill="#ddd", tags="grid")
        for y in range(0, int(height), int(spacing)):
            self.canvas.create_line(0, y, width, y, fill="#ddd", tags="grid")

    def create_diagram_image_without_grid(self):
        if hasattr(self, "canvas") and self.canvas.winfo_exists():
            target_canvas = self.canvas
        elif hasattr(self, "page_diagram") and self.page_diagram is not None:
            target_canvas = self.page_diagram.canvas
        else:
            return None
        grid_items = target_canvas.find_withtag("grid")
        target_canvas.delete("grid")
        target_canvas.update()
        bbox = target_canvas.bbox("all")
        if not bbox:
            return None
        x, y, w, h = bbox[0], bbox[1], bbox[2]-bbox[0], bbox[3]-bbox[1]
        ps = target_canvas.postscript(colormode="color", x=x, y=y, width=w, height=h)
        from io import BytesIO
        ps_bytes = BytesIO(ps.encode("utf-8"))
        img = Image.open(ps_bytes)
        img.load(scale=3)
        if target_canvas == self.canvas:
            self.redraw_canvas()
        else:
            self.page_diagram.redraw_canvas()
        return img.convert("RGB")

    def draw_connections(self, node, drawn_ids=set()):
        if id(node) in drawn_ids:
            return
        drawn_ids.add(id(node))
        if node.is_page and node.is_primary_instance:
            return
        if node.children:
            region_width = 100 * self.zoom
            parent_bottom = (node.x * self.zoom, node.y * self.zoom + 40 * self.zoom)
            N = len(node.children)
            for i, child in enumerate(node.children):
                parent_conn = (node.x * self.zoom - region_width/2 + (i+0.5)*(region_width/N), parent_bottom[1])
                child_top = (child.x * self.zoom, child.y * self.zoom - 45 * self.zoom)
                fta_drawing_helper.draw_90_connection(self.canvas, parent_conn, child_top, outline_color="dimgray", line_width=1)
            for child in node.children:
                self.draw_connections(child, drawn_ids)

    def draw_node(self, node):
        """
        Draws the given node on the main canvas.
        For clones, it always uses the original’s non-positional attributes (like display_label,
        description, etc.) so that any changes to the original are reflected on all clones.
        """
        # If the node is a clone, use its original for configuration (non-positional attributes)
        source = node if node.is_primary_instance else node.original

        # For display purposes, show the clone marker on the clone's display_label.
        if node.is_primary_instance:
            display_label = source.display_label
        else:
            display_label = source.display_label + " (clone)"

        # Build a short top_text string from the source's attributes.
        subtype_text = source.input_subtype if source.input_subtype else "N/A"
        top_text = (
            f"Type: {source.node_type}\n"
            f"Subtype: {subtype_text}\n"
            f"{display_label}\n"
            f"Desc: {source.description}\n\n"
            f"Rationale: {source.rationale}"
        )
        # For the bottom text, you may choose to display the node's name (which for a clone is
        # usually the same as the original’s name)
        bottom_text = source.name

        # Compute the effective position using the clone’s own (positional) values
        eff_x = node.x * self.zoom
        eff_y = node.y * self.zoom

        # Highlight if selected or in diff list
        if node == self.selected_node:
            outline_color = "red"
            line_width = 2
        elif node.unique_id in self.diff_nodes:
            outline_color = "blue"
            line_width = 2
        else:
            outline_color = "dimgray"
            line_width = 1

        # Determine the fill color (this function already uses the original's display_label)
        fill_color = self.get_node_fill_color(node)
        font_obj = self.diagram_font

        # For shape selection, use the source’s node type and gate type.
        node_type_upper = source.node_type.upper()

        if not node.is_primary_instance:
            # For clones, draw them in a “clone” style.
            if source.is_page:
                fta_drawing_helper.draw_triangle_shape(
                    self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
            elif node_type_upper in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if source.gate_type.upper() == "OR":
                    fta_drawing_helper.draw_rotated_or_gate_clone_shape(
                        self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                        top_text=top_text, bottom_text=bottom_text,
                        fill=fill_color, outline_color=outline_color,
                        line_width=line_width, font_obj=font_obj
                    )
                else:
                    fta_drawing_helper.draw_rotated_and_gate_clone_shape(
                        self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                        top_text=top_text, bottom_text=bottom_text,
                        fill=fill_color, outline_color=outline_color,
                        line_width=line_width, font_obj=font_obj
                    )
            elif node_type_upper in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
            else:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
        else:
            # Primary node: use normal drawing routines.
            if node_type_upper in ["TOP EVENT", "GATE", "RIGOR LEVEL"]:
                if source.is_page and source != self.root_node:
                    fta_drawing_helper.draw_triangle_shape(
                        self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                        top_text=top_text, bottom_text=bottom_text,
                        fill=fill_color, outline_color=outline_color,
                        line_width=line_width, font_obj=font_obj
                    )
                else:
                    if source.gate_type.upper() == "OR":
                        fta_drawing_helper.draw_rotated_or_gate_shape(
                            self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                            top_text=top_text, bottom_text=bottom_text,
                            fill=fill_color, outline_color=outline_color,
                            line_width=line_width, font_obj=font_obj
                        )
                    else:
                        fta_drawing_helper.draw_rotated_and_gate_shape(
                            self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                            top_text=top_text, bottom_text=bottom_text,
                            fill=fill_color, outline_color=outline_color,
                            line_width=line_width, font_obj=font_obj
                        )
            elif node_type_upper in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
            else:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )

        # Draw any additional text (such as equations) from the source.
        if source.equation:
            self.canvas.create_text(
                eff_x - 80 * self.zoom, eff_y - 15 * self.zoom,
                text=source.equation, anchor="e", fill="gray",
                font=self.diagram_font
            )
        if source.detailed_equation:
            self.canvas.create_text(
                eff_x - 80 * self.zoom, eff_y + 15 * self.zoom,
                text=source.detailed_equation, anchor="e", fill="gray",
                font=self.diagram_font
            )

        # Finally, if the node appears multiple times, draw a shared marker.
        if self.occurrence_counts.get(node.unique_id, 0) > 1:
            marker_x = eff_x + 30 * self.zoom
            marker_y = eff_y - 30 * self.zoom
            fta_drawing_helper.draw_shared_marker(self.canvas, marker_x, marker_y, self.zoom)

        if self.review_data:
            unresolved = any(c.node_id == node.unique_id and not c.resolved for c in self.review_data.comments)
            if unresolved:
                self.canvas.create_oval(eff_x + 35 * self.zoom, eff_y + 35 * self.zoom,
                                        eff_x + 45 * self.zoom, eff_y + 45 * self.zoom,
                                        fill='yellow', outline='black')

        if self.review_data:
            unresolved = any(c.node_id == node.unique_id and not c.resolved for c in self.review_data.comments)
            if unresolved:
                self.canvas.create_oval(eff_x + 35 * self.zoom, eff_y + 35 * self.zoom,
                                        eff_x + 45 * self.zoom, eff_y + 45 * self.zoom,
                                        fill='yellow', outline='black')

        if self.review_data:
            unresolved = any(c.node_id == node.unique_id and not c.resolved for c in self.review_data.comments)
            if unresolved:
                self.canvas.create_oval(eff_x + 35 * self.zoom, eff_y + 35 * self.zoom,
                                        eff_x + 45 * self.zoom, eff_y + 45 * self.zoom,
                                        fill='yellow', outline='black')

        if self.review_data:
            unresolved = any(c.node_id == node.unique_id and not c.resolved for c in self.review_data.comments)
            if unresolved:
                self.canvas.create_oval(eff_x + 35 * self.zoom, eff_y + 35 * self.zoom,
                                        eff_x + 45 * self.zoom, eff_y + 45 * self.zoom,
                                        fill='yellow', outline='black')

    def find_node_by_id(self, node, unique_id, visited=None):
        if visited is None:
            visited = set()
        if node.unique_id in visited:
            return None
        visited.add(node.unique_id)
        if node.unique_id == unique_id:
            return node
        for c in node.children:
            res = self.find_node_by_id(c, unique_id, visited)
            if res:
                return res
        return None

    def is_descendant(self, node, possible_ancestor):
        if node == possible_ancestor:
            return True
        for p in node.parents:
            if self.is_descendant(p, possible_ancestor):
                return True
        return False

    def add_node_of_type(self, event_type):
        # If a node is selected, ensure it is a primary instance.
        if self.selected_node:
            if not self.selected_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", 
                    "Cannot add new elements to a clone node.\nPlease select the original node instead.")
                return
            parent_node = self.selected_node
        else:
            sel = self.treeview.selection()
            if sel:
                try:
                    node_id = int(self.treeview.item(sel[0], "tags")[0])
                except (IndexError, ValueError):
                    messagebox.showwarning("No selection", "Select a parent node from the tree.")
                    return
                parent_node = self.find_node_by_id_all(node_id)
            else:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return

        # Prevent adding to base events.
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return

        # Now create the new node.
        if event_type.upper() == "CONFIDENCE LEVEL":
            new_node = FaultTreeNode("", "Confidence Level", parent=parent_node)
            new_node.quant_value = 1
        elif event_type.upper() == "ROBUSTNESS SCORE":
            new_node = FaultTreeNode("", "Robustness Score", parent=parent_node)
            new_node.quant_value = 1
        elif event_type.upper() == "GATE":
            new_node = FaultTreeNode("", "GATE", parent=parent_node)
            new_node.gate_type = "AND"
        elif event_type.upper() == "BASIC EVENT":
            new_node = FaultTreeNode("", "Basic Event", parent=parent_node)
            new_node.failure_prob = 0.0
        elif event_type.upper() == "TRIGGERING CONDITION":
            new_node = FaultTreeNode("", "Triggering Condition", parent=parent_node)
        elif event_type.upper() == "FUNCTIONAL INSUFFICIENCY":
            new_node = FaultTreeNode("", "Functional Insufficiency", parent=parent_node)
        else:
            new_node = FaultTreeNode("", event_type, parent=parent_node)
        new_node.x = parent_node.x + 100
        new_node.y = parent_node.y + 100
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()

    def add_basic_event_from_fmea(self):
        events = list(self.fmea_entries)
        for doc in self.fmeas:
            events.extend(doc.get("entries", []))
        for doc in self.fmedas:
            events.extend(doc.get("entries", []))
        if not events:
            messagebox.showinfo("No Failure Modes", "No FMEA or FMEDA failure modes available.")
            return
        dialog = self.SelectBaseEventDialog(self.root, events)
        selected = dialog.selected
        if not selected:
            return
        if self.selected_node:
            parent_node = self.selected_node
            if not parent_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", "Cannot add to a clone node. Select the original.")
                return
        else:
            sel = self.treeview.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return
            try:
                node_id = int(self.treeview.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                messagebox.showwarning("No selection", "Select a parent node from the tree.")
                return
            parent_node = self.find_node_by_id_all(node_id)
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return
        data = selected.to_dict()
        data.pop("unique_id", None)
        data["children"] = []
        new_node = FaultTreeNode.from_dict(data, parent_node)
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()

    def add_basic_event_from_fmea(self):
        events = list(self.fmea_entries)
        for doc in self.fmeas:
            events.extend(doc.get("entries", []))
        for doc in self.fmedas:
            events.extend(doc.get("entries", []))
        if not events:
            messagebox.showinfo("No Failure Modes", "No FMEA or FMEDA failure modes available.")
            return
        dialog = self.SelectBaseEventDialog(self.root, events)
        selected = dialog.selected
        if not selected:
            return
        if self.selected_node:
            parent_node = self.selected_node
            if not parent_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", "Cannot add to a clone node. Select the original.")
                return
        else:
            sel = self.treeview.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return
            try:
                node_id = int(self.treeview.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                messagebox.showwarning("No selection", "Select a parent node from the tree.")
                return
            parent_node = self.find_node_by_id_all(node_id)
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return
        data = selected.to_dict()
        data.pop("unique_id", None)
        data["children"] = []
        new_node = FaultTreeNode.from_dict(data, parent_node)
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()

    def add_basic_event_from_fmea(self):
        events = list(self.fmea_entries)
        for doc in self.fmeas:
            events.extend(doc.get("entries", []))
        for doc in self.fmedas:
            events.extend(doc.get("entries", []))
        if not events:
            messagebox.showinfo("No Failure Modes", "No FMEA or FMEDA failure modes available.")
            return
        dialog = self.SelectBaseEventDialog(self.root, events)
        selected = dialog.selected
        if not selected:
            return
        if self.selected_node:
            parent_node = self.selected_node
            if not parent_node.is_primary_instance:
                messagebox.showwarning("Invalid Operation", "Cannot add to a clone node. Select the original.")
                return
        else:
            sel = self.treeview.selection()
            if not sel:
                messagebox.showwarning("No selection", "Select a parent node to paste into.")
                return
            try:
                node_id = int(self.treeview.item(sel[0], "tags")[0])
            except (IndexError, ValueError):
                messagebox.showwarning("No selection", "Select a parent node from the tree.")
                return
            parent_node = self.find_node_by_id_all(node_id)
        if parent_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE", "BASIC EVENT"]:
            messagebox.showwarning("Invalid", "Base events cannot have children.")
            return
        data = selected.to_dict()
        data.pop("unique_id", None)
        data["children"] = []
        new_node = FaultTreeNode.from_dict(data, parent_node)
        parent_node.children.append(new_node)
        new_node.parents.append(parent_node)
        self.update_views()


    def remove_node(self):
        sel = self.treeview.selection()
        target = None
        if sel:
            tags = self.treeview.item(sel[0], "tags")
            target = self.find_node_by_id(self.root_node, int(tags[0]))
        elif self.selected_node:
            target = self.selected_node
        if target and target != self.root_node:
            if target.parents:
                for p in target.parents:
                    if target in p.children:
                        p.children.remove(target)
                target.parents = []
            self.update_views()
        else:
            messagebox.showwarning("Invalid", "Cannot remove the root node.")

    def remove_connection(self, node):
        if node and node != self.root_node:
            if node.parents:
                for p in node.parents:
                    if node in p.children:
                        p.children.remove(node)
                node.parents = []
                if node not in self.top_events:
                    self.top_events.append(node)
                self.update_views()
                messagebox.showinfo("Remove Connection",
                                    f"Disconnected {node.name} from its parent(s) and made it a top-level event.")
            else:
                messagebox.showwarning("Remove Connection", "Node has no parent connection.")
        else:
            messagebox.showwarning("Remove Connection", "Cannot disconnect the root node.")

    def delete_node_and_subtree(self, node):
        if node:
            if node in self.top_events:
                self.top_events.remove(node)
            else:
                for p in node.parents:
                    if node in p.children:
                        p.children.remove(node)
                node.parents = []
            self.update_views()
            messagebox.showinfo("Delete Node", f"Deleted {node.name} and its subtree.")
        else:
            messagebox.showwarning("Delete Node", "Select a node to delete.")

    def calculate_overall(self):
        for top_event in self.top_events:
            AD_RiskAssessment_Helper.calculate_assurance_recursive(top_event, self.top_events)
        self.update_views()
        results = ""
        for top_event in self.top_events:
            if top_event.quant_value is not None:
                disc = AD_RiskAssessment_Helper.discretize_level(top_event.quant_value)
                results += (f"Top Event {top_event.display_label}\n"
                            f"(Continuous: {top_event.quant_value:.2f}, Discrete: {disc})\n\n")
        messagebox.showinfo("Calculation", results.strip())

    def calculate_pmfh(self):
        self.update_basic_event_probabilities()
        spf = 0.0
        lpf = 0.0
        for be in self.get_all_basic_events():
            fm = self.get_failure_mode_node(be)
            fit = getattr(be, "fmeda_fit", None)
            if fit is None or fit == 0.0:
                fit = getattr(fm, "fmeda_fit", 0.0)
            dc = getattr(be, "fmeda_diag_cov", getattr(fm, "fmeda_diag_cov", 0.0))
            if be.fmeda_fault_type == "permanent":
                spf += fit * (1 - dc)
            else:
                lpf += fit * (1 - dc)
        self.spfm = spf
        self.lpfm = lpf

        pmhf = 0.0
        for te in self.top_events:
            prob = AD_RiskAssessment_Helper.calculate_probability_recursive(te)
            te.probability = prob
            pmhf += prob

        self.update_views()
        lines = [f"Total PMHF: {pmhf:.2e}"]
        overall_ok = True
        for te in self.top_events:
            asil = te.safety_goal_asil or "QM"
            target = PMHF_TARGETS.get(asil, 1.0)
            ok = te.probability <= target
            overall_ok = overall_ok and ok
            symbol = CHECK_MARK if ok else CROSS_MARK
            lines.append(f"{te.user_name or te.display_label}: {te.probability:.2e} <= {target:.1e} {symbol}")
        self.pmhf_var.set("\n".join(lines))
        self.pmhf_label.config(foreground="green" if overall_ok else "red", font=("Segoe UI", 10, "bold"))

    def show_requirements_matrix(self):
        """Display a matrix table of requirements vs. basic events."""
        self.update_requirement_statuses()
        basic_events = [n for n in self.get_all_nodes(self.root_node)
                        if n.node_type.upper() == "BASIC EVENT"]
        reqs = list(global_requirements.values())
        reqs.sort(key=lambda r: r.get("req_type", ""))

        win = tk.Toplevel(self.root)
        win.title("Requirements Matrix")

        columns = [
            "Req ID",
            "ASIL",
            "Type",
            "Status",
            "Parent",
            "Text",
        ] + [be.user_name or f"BE {be.unique_id}" for be in basic_events]
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col not in ["Text"] else 300, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)


        for req in reqs:
            row = [
                req.get("id", ""),
                req.get("asil", ""),
                req.get("req_type", ""),
                req.get("status", "draft"),
                req.get("parent_id", ""),
                req.get("text", ""),
            ]
            for be in basic_events:
                linked = any(r.get("id") == req.get("id") for r in getattr(be, "safety_requirements", []))
                row.append("X" if linked else "")
            tree.insert("", "end", values=row)

        # Show allocation and safety goal traceability below the table
        frame = tk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True)
        vbar = tk.Scrollbar(frame, orient="vertical")
        text = tk.Text(frame, wrap="word", yscrollcommand=vbar.set, height=8)
        text.tag_configure("added", foreground="blue")
        text.tag_configure("removed", foreground="red")
        vbar.config(command=text.yview)
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)

        base_data = self.versions[-1]["data"] if self.versions else None

        def alloc_from_data(req_id):
            if not base_data:
                return ""
            names = []
            def traverse(d):
                if any(r.get("id") == req_id for r in d.get("safety_requirements", [])):
                    names.append(d.get("user_name") or f"Node {d.get('unique_id')}")
                for ch in d.get("children", []):
                    traverse(ch)
            for t in base_data.get("top_events", []):
                traverse(t)
            for fmea in base_data.get("fmeas", []):
                for e in fmea.get("entries", []):
                    if any(r.get("id") == req_id for r in e.get("safety_requirements", [])):
                        name = e.get("description") or e.get("user_name", f"BE {e.get('unique_id','')}")
                        names.append(f"{fmea['name']}:{name}")
            return ", ".join(names)

        def goals_from_data(req_id):
            if not base_data:
                return ""
            nodes = []
            def gather(n):
                nodes.append(n)
                for ch in n.get("children", []):
                    gather(ch)
            for t in base_data.get("top_events", []):
                gather(t)
            id_map = {n["unique_id"]: n for n in nodes}
            def collect_goal_names(nd, acc):
                if nd.get("node_type", "").upper() == "TOP EVENT":
                    acc.add(nd.get("safety_goal_description") or nd.get("user_name") or f"SG {nd.get('unique_id')}")
                for p in nd.get("parents", []):
                    pid = p.get("unique_id")
                    if pid and pid in id_map:
                        collect_goal_names(id_map[pid], acc)
            goals = set()
            for n in nodes:
                if any(r.get("id") == req_id for r in n.get("safety_requirements", [])):
                    collect_goal_names(n, goals)
            for fmea in base_data.get("fmeas", []):
                for e in fmea.get("entries", []):
                    if any(r.get("id") == req_id for r in e.get("safety_requirements", [])):
                        parents = e.get("parents", [])
                        if parents:
                            pid = parents[0].get("unique_id")
                            if pid and pid in id_map:
                                collect_goal_names(id_map[pid], goals)
            return ", ".join(sorted(goals))

        import difflib

        def insert_diff(widget, old, new):
            matcher = difflib.SequenceMatcher(None, old, new)
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == "equal":
                    widget.insert(tk.END, new[j1:j2])
                elif tag == "delete":
                    widget.insert(tk.END, old[i1:i2], "removed")
                elif tag == "insert":
                    widget.insert(tk.END, new[j1:j2], "added")
                elif tag == "replace":
                    widget.insert(tk.END, old[i1:i2], "removed")
                    widget.insert(tk.END, new[j1:j2], "added")

        def insert_list_diff(widget, old, new):
            old_items = [s.strip() for s in old.split(',') if s.strip()]
            new_items = [s.strip() for s in new.split(',') if s.strip()]
            old_set = set(old_items)
            new_set = set(new_items)
            first = True
            for item in new_items:
                if not first:
                    widget.insert(tk.END, ", ")
                first = False
                if item not in old_set:
                    widget.insert(tk.END, item, "added")
                else:
                    widget.insert(tk.END, item)
            for item in old_items:
                if item not in new_set:
                    if not first:
                        widget.insert(tk.END, ", ")
                    first = False
                    widget.insert(tk.END, item, "removed")

        for req in reqs:
            rid = req.get("id")
            alloc = ", ".join(self.get_requirement_allocation_names(rid))
            goals = ", ".join(self.get_requirement_goal_names(rid))
            text.insert(tk.END, f"[{rid}] {req.get('text','')}\n")
            text.insert(tk.END, "  Allocated to: ")
            if base_data:
                insert_list_diff(text, alloc_from_data(rid), alloc)
            else:
                text.insert(tk.END, alloc)
            text.insert(tk.END, "\n  Safety Goals: ")
            if base_data:
                insert_diff(text, goals_from_data(rid), goals)
            else:
                text.insert(tk.END, goals)
            text.insert(tk.END, "\n\n")

        tk.Button(win, text="Open Requirements Editor", command=self.show_requirements_editor).pack(pady=5)

    def show_requirements_editor(self):
        """Open an editor to manage global requirements and traceability."""
        self.update_requirement_statuses()
        win = tk.Toplevel(self.root)
        win.title("Requirements Editor")

        columns = ["ID", "ASIL", "Type", "Status", "Parent", "Text"]
        tree = ttk.Treeview(win, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=120 if col != "Text" else 300, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh_tree():
            tree.delete(*tree.get_children())
            for req in global_requirements.values():
                tree.insert("", "end", iid=req.get("id"), values=[
                    req.get("id", ""),
                    req.get("asil", ""),
                    req.get("req_type", ""),
                    req.get("status", "draft"),
                    req.get("parent_id", ""),
                    req.get("text", ""),
                ])

        class ReqDialog(simpledialog.Dialog):
            def __init__(self, parent, title, initial=None):
                self.initial = initial or {}
                super().__init__(parent, title=title)

            def body(self, master):
                ttk.Label(master, text="ID:").grid(row=0, column=0, sticky="e")
                self.id_var = tk.StringVar(value=self.initial.get("id", ""))
                tk.Entry(master, textvariable=self.id_var).grid(row=0, column=1, padx=5, pady=5)

                ttk.Label(master, text="Type:").grid(row=1, column=0, sticky="e")
                self.type_var = tk.StringVar(value=self.initial.get("req_type", "vehicle"))
                ttk.Combobox(master, textvariable=self.type_var, values=["vehicle", "operational"], state="readonly").grid(row=1, column=1, padx=5, pady=5)

                ttk.Label(master, text="ASIL:").grid(row=2, column=0, sticky="e")
                self.asil_var = tk.StringVar(value=self.initial.get("asil", "QM"))
                ttk.Combobox(master, textvariable=self.asil_var, values=ASIL_LEVEL_OPTIONS, state="readonly", width=8).grid(row=2, column=1, padx=5, pady=5)

                ttk.Label(master, text="Parent ID:").grid(row=3, column=0, sticky="e")
                self.parent_var = tk.StringVar(value=self.initial.get("parent_id", ""))
                tk.Entry(master, textvariable=self.parent_var).grid(row=3, column=1, padx=5, pady=5)

                ttk.Label(master, text="Status:").grid(row=4, column=0, sticky="e")
                self.status_var = tk.StringVar(value=self.initial.get("status", "draft"))
                ttk.Combobox(master, textvariable=self.status_var,
                             values=["draft", "in review", "peer reviewed", "pending approval", "approved"],
                             state="readonly").grid(row=4, column=1, padx=5, pady=5)

                ttk.Label(master, text="Text:").grid(row=5, column=0, sticky="e")
                self.text_var = tk.StringVar(value=self.initial.get("text", ""))
                tk.Entry(master, textvariable=self.text_var, width=40).grid(row=5, column=1, padx=5, pady=5)
                return master

            def apply(self):
                self.result = {
                    "id": self.id_var.get().strip() or str(uuid.uuid4()),
                    "req_type": self.type_var.get().strip(),
                    "asil": self.asil_var.get().strip(),
                    "parent_id": self.parent_var.get().strip(),
                    "status": self.status_var.get().strip(),
                    "text": self.text_var.get().strip(),
                }

            def validate(self):
                rid = self.id_var.get().strip()
                if rid and rid != self.initial.get("id") and rid in global_requirements:
                    messagebox.showerror("ID", "ID already exists")
                    return False
                return True

        class TraceDialog(simpledialog.Dialog):
            def __init__(self, parent, app, requirement):
                self.requirement = requirement
                self.app = app
                super().__init__(parent, title="Edit Traceability")

            def body(self, master):
                self.vars = {}
                canvas = tk.Canvas(master)
                frame = tk.Frame(canvas)
                vsb = tk.Scrollbar(master, orient="vertical", command=canvas.yview)
                canvas.configure(yscrollcommand=vsb.set)
                canvas.pack(side="left", fill="both", expand=True)
                vsb.pack(side="right", fill="y")
                canvas.create_window((0, 0), window=frame, anchor="nw")

                def on_config(event):
                    canvas.configure(scrollregion=canvas.bbox("all"))

                frame.bind("<Configure>", on_config)

                for n in [n for n in self.app.get_all_nodes_in_model() if n.node_type.upper() == "BASIC EVENT"]:
                    var = tk.BooleanVar(value=any(r.get("id") == self.requirement.get("id") for r in getattr(n, "safety_requirements", [])))
                    self.vars[n] = var
                    ttk.Checkbutton(frame, text=n.user_name or f"BE {n.unique_id}", variable=var).pack(anchor="w")
                return frame

            def apply(self):
                self.result = {n: v.get() for n, v in self.vars.items()}

        def add_req():
            dlg = ReqDialog(win, "Add Requirement")
            if dlg.result:
                global_requirements[dlg.result["id"]] = dlg.result
                refresh_tree()

        def edit_req():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            dlg = ReqDialog(win, "Edit Requirement", global_requirements.get(rid))
            if dlg.result:
                global_requirements[rid].update(dlg.result)
                refresh_tree()

        def del_req():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            if messagebox.askyesno("Delete", "Delete requirement?"):
                del global_requirements[rid]
                for n in self.get_all_nodes(self.root_node):
                    reqs = getattr(n, "safety_requirements", [])
                    n.safety_requirements = [r for r in reqs if r.get("id") != rid]
                for fmea in self.fmeas:
                    for e in fmea.get("entries", []):
                        reqs = e.get("safety_requirements", [])
                        e["safety_requirements"] = [r for r in reqs if r.get("id") != rid]
                refresh_tree()

        def edit_trace():
            sel = tree.selection()
            if not sel:
                return
            rid = sel[0]
            req = global_requirements.get(rid)
            dlg = TraceDialog(win, self, req)
            if dlg.result is not None:
                for node, val in dlg.result.items():
                    reqs = getattr(node, "safety_requirements", [])
                    present = any(r.get("id") == rid for r in reqs)
                    if val and not present:
                        reqs.append(req)
                    if not val and present:
                        node.safety_requirements = [r for r in reqs if r.get("id") != rid]
                refresh_tree()

        btn = tk.Frame(win)
        btn.pack(fill=tk.X)
        tk.Button(btn, text="Add", command=add_req).pack(side=tk.LEFT)
        tk.Button(btn, text="Edit", command=edit_req).pack(side=tk.LEFT)
        tk.Button(btn, text="Delete", command=del_req).pack(side=tk.LEFT)
        tk.Button(btn, text="Traceability", command=edit_trace).pack(side=tk.LEFT)

        refresh_tree()


    def show_fmea_list(self):
        win = tk.Toplevel(self.root)
        win.title("FMEA List")
        listbox = tk.Listbox(win, height=10, width=40)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for fmea in self.fmeas:
            listbox.insert(tk.END, fmea['name'])

        def open_selected(event=None):
            sel = listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            win.destroy()
            self.show_fmea_table(self.fmeas[idx])

        def add_fmea():
            name = simpledialog.askstring("New FMEA", "Enter FMEA name:")
            if name:
                file_name = f"fmea_{name}.csv"
                self.fmeas.append({'name': name, 'entries': [], 'file': file_name})
                listbox.insert(tk.END, name)

        def delete_fmea():
            sel = listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            del self.fmeas[idx]
            listbox.delete(idx)

        listbox.bind("<Double-1>", open_selected)
        btn_frame = ttk.Frame(win)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Button(btn_frame, text="Open", command=open_selected).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Add", command=add_fmea).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Delete", command=delete_fmea).pack(fill=tk.X)

    def show_fmeda_list(self):
        win = tk.Toplevel(self.root)
        win.title("FMEDA List")
        listbox = tk.Listbox(win, height=10, width=40)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        for doc in self.fmedas:
            listbox.insert(tk.END, doc['name'])

        def open_selected(event=None):
            sel = listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            win.destroy()
            self.show_fmea_table(self.fmedas[idx], fmeda=True)

        def add_fmeda():
            name = simpledialog.askstring("New FMEDA", "Enter FMEDA name:")
            if name:
                file_name = f"fmeda_{name}.csv"
                self.fmedas.append({'name': name, 'entries': [], 'file': file_name, 'bom': ''})
                listbox.insert(tk.END, name)

        def delete_fmeda():
            sel = listbox.curselection()
            if not sel:
                return
            idx = sel[0]
            del self.fmedas[idx]
            listbox.delete(idx)

        listbox.bind("<Double-1>", open_selected)
        btn_frame = ttk.Frame(win)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)
        ttk.Button(btn_frame, text="Open", command=open_selected).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Add", command=add_fmeda).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Delete", command=delete_fmeda).pack(fill=tk.X)
        
    def show_triggering_condition_list(self):
        win = tk.Toplevel(self.root)
        win.title("Triggering Conditions")
        lb = tk.Listbox(win, height=10, width=40)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        for tc in self.get_all_triggering_conditions():
            lb.insert(tk.END, tc.user_name or f"TC {tc.unique_id}")
        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV","*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["ID","Name","Description"])
                for n in self.get_all_triggering_conditions():
                    w.writerow([n.unique_id, n.user_name, n.description])
            messagebox.showinfo("Export","Triggering conditions exported.")
        ttk.Button(win, text="Export CSV", command=export_csv).pack(side=tk.RIGHT, padx=5, pady=5)

    def show_functional_insufficiency_list(self):
        win = tk.Toplevel(self.root)
        win.title("Functional Insufficiencies")
        lb = tk.Listbox(win, height=10, width=40)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        for fi in self.get_all_functional_insufficiencies():
            lb.insert(tk.END, fi.user_name or f"FI {fi.unique_id}")
        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV","*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["ID","Name","Description"])
                for n in self.get_all_functional_insufficiencies():
                    w.writerow([n.unique_id, n.user_name, n.description])
            messagebox.showinfo("Export","Functional insufficiencies exported.")
        ttk.Button(win, text="Export CSV", command=export_csv).pack(side=tk.RIGHT, padx=5, pady=5)

    class FMEARowDialog(simpledialog.Dialog):
        def __init__(self, parent, node, app, fmea_entries, mechanisms=None, hide_diagnostics=False):
            self.node = node
            self.app = app
            self.fmea_entries = fmea_entries
            self.mechanisms = mechanisms or []
            self.hide_diagnostics = hide_diagnostics
            super().__init__(parent, title="Edit FMEA Entry")
            self.app.selected_node = node

        def body(self, master):
            self.resizable(False, False)

            ttk.Label(master, text="Component:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
            if self.node.parents:
                comp = self.node.parents[0].user_name or f"Node {self.node.parents[0].unique_id}"
            else:
                comp = getattr(self.node, "fmea_component", "")
            comp_names = {c.name for c in self.app.reliability_components}
            basic_events = self.app.get_all_basic_events()
            for be in basic_events + self.fmea_entries:
                parent = be.parents[0] if be.parents else None
                if parent and parent.user_name:
                    comp_names.add(parent.user_name)
                else:
                    name = getattr(be, "fmea_component", "")
                    if name:
                        comp_names.add(name)
            self.comp_var = tk.StringVar(value=comp)
            self.comp_combo = ttk.Combobox(master, textvariable=self.comp_var,
                                           values=sorted(comp_names), width=30)
            self.comp_combo.grid(row=0, column=1, padx=5, pady=5)

            ttk.Label(master, text="Failure Mode:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
            # Include failure modes from both the FTA and any FMEA specific
            # entries so the combo box always lists all available modes.
            self.mode_map = {
                be.description or (be.user_name or f"BE {be.unique_id}"): be
                for be in basic_events + self.fmea_entries
            }
            for doc in self.app.hazop_docs:
                for e in doc.entries:
                    label = f"{e.function}: {e.malfunction}"
                    obj = types.SimpleNamespace(
                        description=e.malfunction,
                        user_name=label,
                        parents=[],
                        fmea_component=e.component,
                    )
                    self.mode_map[label] = obj
            mode_names = list(self.mode_map.keys())
            self.mode_var = tk.StringVar(value=self.node.description or self.node.user_name)
            self.mode_combo = ttk.Combobox(master, textvariable=self.mode_var,
                                          values=mode_names, width=30)
            self.mode_combo.grid(row=1, column=1, padx=5, pady=5)

            def mode_sel(_):
                label = self.mode_var.get()
                src = self.mode_map.get(label)
                if src:
                    if src.parents:
                        comp_name = src.parents[0].user_name or f"Node {src.parents[0].unique_id}"
                    else:
                        comp_name = getattr(src, "fmea_component", "")
                    if comp_name:
                        self.comp_var.set(comp_name)

            self.mode_combo.bind("<<ComboboxSelected>>", mode_sel)

            ttk.Label(master, text="Failure Effect:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
            self.effect_text = tk.Text(master, width=30, height=3)
            self.effect_text.insert("1.0", self.node.fmea_effect)
            self.effect_text.grid(row=2, column=1, padx=5, pady=5)

            ttk.Label(master, text="Potential Cause:").grid(row=3, column=0, sticky="e", padx=5, pady=5)
            self.cause_text = tk.Text(master, width=30, height=3)
            self.cause_text.insert("1.0", getattr(self.node, 'fmea_cause', ''))
            self.cause_text.grid(row=3, column=1, padx=5, pady=5)

            ttk.Label(master, text="Related Malfunction:").grid(row=4, column=0, sticky="e", padx=5, pady=5)
            self.mal_var = tk.StringVar(value=getattr(self.node, 'fmeda_malfunction', ''))
            ttk.Entry(master, textvariable=self.mal_var, width=30).grid(row=4, column=1, padx=5, pady=5)

            ttk.Label(master, text="Violates Safety Goal:").grid(row=5, column=0, sticky="e", padx=5, pady=5)
            self.sg_var = tk.StringVar(value=getattr(self.node, 'fmeda_safety_goal', ''))
            ttk.Entry(master, textvariable=self.sg_var, width=30).grid(row=5, column=1, padx=5, pady=5)

            ttk.Label(master, text="Severity (1-10):").grid(row=6, column=0, sticky="e", padx=5, pady=5)
            self.sev_spin = tk.Spinbox(master, from_=1, to=10, width=5)
            self.sev_spin.delete(0, tk.END)
            self.sev_spin.insert(0, str(self.node.fmea_severity))
            self.sev_spin.grid(row=6, column=1, sticky="w", padx=5, pady=5)

            ttk.Label(master, text="Occurrence (1-10):").grid(row=7, column=0, sticky="e", padx=5, pady=5)
            self.occ_spin = tk.Spinbox(master, from_=1, to=10, width=5)
            self.occ_spin.delete(0, tk.END)
            self.occ_spin.insert(0, str(self.node.fmea_occurrence))
            self.occ_spin.grid(row=7, column=1, sticky="w", padx=5, pady=5)

            ttk.Label(master, text="Detection (1-10):").grid(row=8, column=0, sticky="e", padx=5, pady=5)
            self.det_spin = tk.Spinbox(master, from_=1, to=10, width=5)
            self.det_spin.delete(0, tk.END)
            self.det_spin.insert(0, str(self.node.fmea_detection))
            self.det_spin.grid(row=8, column=1, sticky="w", padx=5, pady=5)

            row = 9
            if not self.hide_diagnostics:
                ttk.Label(master, text="Diag Coverage (0-1):").grid(row=row, column=0, sticky="e", padx=5, pady=5)
                self.dc_var = tk.DoubleVar(value=getattr(self.node, 'fmeda_diag_cov', 0.0))
                ttk.Entry(master, textvariable=self.dc_var, width=5).grid(row=row, column=1, sticky="w", padx=5, pady=5)
                row += 1

                ttk.Label(master, text="Mechanism:").grid(row=row, column=0, sticky="e", padx=5, pady=5)
                self.mech_var = tk.StringVar(value=getattr(self.node, 'fmeda_mechanism', ''))
                self.mech_combo = ttk.Combobox(master, textvariable=self.mech_var, values=[m.name for m in self.mechanisms], state='readonly', width=30)
                self.mech_combo.grid(row=row, column=1, padx=5, pady=5)

                def mech_sel(_):
                    name = self.mech_var.get()
                    for m in self.mechanisms:
                        if m.name == name:
                            self.dc_var.set(m.coverage)
                            break

                self.mech_combo.bind("<<ComboboxSelected>>", mech_sel)
                row += 1
            else:
                self.dc_var = tk.DoubleVar(value=0.0)
                self.mech_var = tk.StringVar(value="")

            ttk.Label(master, text="Fault Type:").grid(row=row, column=0, sticky="e", padx=5, pady=5)
            self.ftype_var = tk.StringVar(value=getattr(self.node, 'fmeda_fault_type', 'permanent'))
            ttk.Combobox(master, textvariable=self.ftype_var, values=['permanent', 'transient'], state='readonly', width=10).grid(row=row, column=1, sticky="w", padx=5, pady=5)

            row += 1
            ttk.Label(master, text="Fault Fraction:").grid(row=row, column=0, sticky="e", padx=5, pady=5)
            self.ffrac_var = tk.DoubleVar(value=getattr(self.node, 'fmeda_fault_fraction', 0.0))
            ttk.Entry(master, textvariable=self.ffrac_var, width=5).grid(row=row, column=1, sticky="w", padx=5, pady=5)

            row += 1
            ttk.Label(master, text="FIT Rate:").grid(row=row, column=0, sticky="e", padx=5, pady=5)
            self.fit_var = tk.DoubleVar(value=getattr(self.node, 'fmeda_fit', 0.0))
            ttk.Entry(master, textvariable=self.fit_var, width=10).grid(row=row, column=1, sticky="w", padx=5, pady=5)

            row += 1
            ttk.Label(master, text="Requirements:").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
            self.req_frame = ttk.Frame(master)
            self.req_frame.grid(row=row, column=1, padx=5, pady=5, sticky="w")
            self.req_listbox = tk.Listbox(self.req_frame, height=4, width=40)
            self.req_listbox.grid(row=0, column=0, columnspan=3, sticky="w")
            if not hasattr(self.node, "safety_requirements"):
                self.node.safety_requirements = []
            for req in self.node.safety_requirements:
                desc = f"[{req['req_type']}] [{req.get('asil','')}] {req['text']}"
                self.req_listbox.insert(tk.END, desc)
            ttk.Button(self.req_frame, text="Add New", command=self.add_safety_requirement).grid(row=1, column=0, padx=2, pady=2)
            ttk.Button(self.req_frame, text="Edit", command=self.edit_safety_requirement).grid(row=1, column=1, padx=2, pady=2)
            ttk.Button(self.req_frame, text="Delete", command=self.delete_safety_requirement).grid(row=1, column=2, padx=2, pady=2)
            ttk.Button(self.req_frame, text="Add Existing", command=self.add_existing_requirement).grid(row=1, column=3, padx=2, pady=2)
            ttk.Button(self.req_frame, text="Comment", command=self.comment_requirement).grid(row=1, column=4, padx=2, pady=2)
            ttk.Button(self.req_frame, text="Comment FMEA", command=self.comment_fmea).grid(row=1, column=5, padx=2, pady=2)
            return self.effect_text

        def apply(self):
            comp = self.comp_var.get()
            if self.node.parents:
                self.node.parents[0].user_name = comp
            # Always store the component name so it can be restored on load
            self.node.fmea_component = comp
            self.node.description = self.mode_var.get()
            self.node.fmea_effect = self.effect_text.get("1.0", "end-1c")
            self.node.fmea_cause = self.cause_text.get("1.0", "end-1c")
            try:
                self.node.fmea_severity = int(self.sev_spin.get())
            except ValueError:
                self.node.fmea_severity = 1
            try:
                self.node.fmea_occurrence = int(self.occ_spin.get())
            except ValueError:
                self.node.fmea_occurrence = 1
            try:
                self.node.fmea_detection = int(self.det_spin.get())
            except ValueError:
                self.node.fmea_detection = 1
            self.node.fmeda_malfunction = self.mal_var.get()
            self.node.fmeda_safety_goal = self.sg_var.get()
            try:
                self.node.fmeda_diag_cov = float(self.dc_var.get())
            except ValueError:
                self.node.fmeda_diag_cov = 0.0
            self.node.fmeda_mechanism = self.mech_var.get()
            if self.hide_diagnostics:
                self.node.fmeda_diag_cov = 0.0
                self.node.fmeda_mechanism = ""
            self.node.fmeda_fault_type = self.ftype_var.get()
            try:
                self.node.fmeda_fault_fraction = float(self.ffrac_var.get())
            except ValueError:
                self.node.fmeda_fault_fraction = 0.0
            try:
                self.node.fmeda_fit = float(self.fit_var.get())
            except ValueError:
                self.node.fmeda_fit = 0.0
            self.app.propagate_failure_mode_attributes(self.node)

        def add_existing_requirement(self):
            global global_requirements
            if not global_requirements:
                messagebox.showinfo("No Existing Requirements", "There are no existing requirements to add.")
                return
            dialog = EditNodeDialog.SelectExistingRequirementsDialog(self, title="Select Existing Requirements")
            if dialog.result:
                if not hasattr(self.node, "safety_requirements"):
                    self.node.safety_requirements = []
                for req_id in dialog.result:
                    req = global_requirements.get(req_id)
                    if req and not any(r["id"] == req_id for r in self.node.safety_requirements):
                        self.node.safety_requirements.append(req)
                        desc = f"[{req['req_type']}] [{req.get('asil','')}] {req['text']}"
                self.req_listbox.insert(tk.END, desc)
            else:
                messagebox.showinfo("No Selection", "No existing requirements were selected.")

        def comment_requirement(self):
            sel = self.req_listbox.curselection()
            if not sel:
                messagebox.showwarning("Comment", "Select a requirement")
                return
            req = self.node.safety_requirements[sel[0]]
            self.app.selected_node = self.node
            # include the node id as well so the toolbox has full context
            self.app.comment_target = ("requirement", self.node.unique_id, req.get("id"))
            self.app.open_review_toolbox()

        def comment_fmea(self):
            self.app.selected_node = self.node
            self.app.comment_target = ("fmea", self.node.unique_id)
            self.app.open_review_toolbox()


        def add_safety_requirement(self):
            global global_requirements
            dialog = EditNodeDialog.RequirementDialog(self, title="Add Safety Requirement")
            if dialog.result is None or dialog.result["text"] == "":
                return
            custom_id = dialog.result.get("custom_id", "").strip()
            if not custom_id:
                custom_id = str(uuid.uuid4())
            if custom_id in global_requirements:
                req = global_requirements[custom_id]
                req["req_type"] = dialog.result["req_type"]
                req["text"] = dialog.result["text"]
                req["asil"] = dialog.result.get("asil", "QM")
            else:
                req = {
                    "id": custom_id,
                    "req_type": dialog.result["req_type"],
                    "text": dialog.result["text"],
                    "custom_id": custom_id,
                    "asil": dialog.result.get("asil", "QM")
                }
                global_requirements[custom_id] = req
            if not hasattr(self.node, "safety_requirements"):
                self.node.safety_requirements = []
            if not any(r["id"] == custom_id for r in self.node.safety_requirements):
                self.node.safety_requirements.append(req)
                desc = f"[{req['req_type']}] [{req.get('asil','')}] {req['text']}"
                self.req_listbox.insert(tk.END, desc)

        def edit_safety_requirement(self):
            selected = self.req_listbox.curselection()
            if not selected:
                messagebox.showwarning("Edit Requirement", "Select a requirement to edit.")
                return
            index = selected[0]
            current_req = self.node.safety_requirements[index]
            initial_req = current_req.copy()
            dialog = EditNodeDialog.RequirementDialog(self, title="Edit Safety Requirement", initial_req=initial_req)
            if dialog.result is None or dialog.result["text"] == "":
                return
            new_custom_id = dialog.result["custom_id"].strip() or current_req.get("custom_id") or current_req.get("id") or str(uuid.uuid4())
            current_req["req_type"] = dialog.result["req_type"]
            current_req["text"] = dialog.result["text"]
            current_req["asil"] = dialog.result.get("asil", "QM")
            current_req["custom_id"] = new_custom_id
            current_req["id"] = new_custom_id
            global_requirements[new_custom_id] = current_req
            self.node.safety_requirements[index] = current_req
            self.req_listbox.delete(index)
            desc = f"[{current_req['req_type']}] [{current_req.get('asil','')}] {current_req['text']}"
            self.req_listbox.insert(index, desc)

        def delete_safety_requirement(self):
            selected = self.req_listbox.curselection()
            if not selected:
                messagebox.showwarning("Delete Requirement", "Select a requirement to delete.")
                return
            index = selected[0]
            del self.node.safety_requirements[index]
            self.req_listbox.delete(index)

    class SelectBaseEventDialog(simpledialog.Dialog):
        def __init__(self, parent, events, allow_new=False):
            self.events = events
            self.allow_new = allow_new
            self.selected = None
            super().__init__(parent, title="Select Base Event")

        def body(self, master):
            self.listbox = tk.Listbox(master, height=10, width=40)
            for be in self.events:
                label = be.description or (be.user_name or f"BE {be.unique_id}")
                self.listbox.insert(tk.END, label)
            if self.allow_new:
                self.listbox.insert(tk.END, "<Create New Failure Mode>")
            self.listbox.grid(row=0, column=0, padx=5, pady=5)
            return self.listbox

        def apply(self):
            sel = self.listbox.curselection()
            if sel:
                idx = sel[0]
                if self.allow_new and idx == len(self.events):
                    self.selected = "NEW"
                else:
                    self.selected = self.events[idx]

    def show_fmea_table(self, fmea=None, fmeda=False):
        """Display an editable AIAG-compliant FMEA or FMEDA table."""
        basic_events = self.get_all_basic_events()
        entries = self.fmea_entries if fmea is None else fmea['entries']
        win = tk.Toplevel(self.root)
        title = f"FMEA Table - {fmea['name']}" if fmea else "FMEA Table"
        win.title(title)

        # give the table a nicer look similar to professional FMEA tools
        style = ttk.Style(win)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(
            "FMEA.Treeview",
            font=("Segoe UI", 10),
            rowheight=22,
        )
        style.configure(
            "FMEA.Treeview.Heading",
            font=("Segoe UI", 10, "bold"),
            background="#d0d0d0",
        )

        columns = [
            "Component",
            "Parent",
            "Failure Mode",
            "Failure Effect",
            "Cause",
            "S",
            "O",
            "D",
            "RPN",
            "Requirements",
        ]
        if fmeda:
            columns.extend([
                "Malfunction",
                "Safety Goal",
                "FaultType",
                "Fraction",
                "FIT",
                "DiagCov",
                "Mechanism",
            ])
        btn_frame = ttk.Frame(win)
        btn_frame.pack(side=tk.TOP, pady=2)
        add_btn = ttk.Button(btn_frame, text="Add Failure Mode")
        add_btn.pack(side=tk.LEFT, padx=2)
        remove_btn = ttk.Button(btn_frame, text="Remove from FMEA")
        remove_btn.pack(side=tk.LEFT, padx=2)
        del_btn = ttk.Button(btn_frame, text="Delete Selected")
        del_btn.pack(side=tk.LEFT, padx=2)
        comment_btn = ttk.Button(btn_frame, text="Comment")
        comment_btn.pack(side=tk.LEFT, padx=2)
        if fmeda:
            calc_btn = ttk.Button(btn_frame, text="Calculate FMEDA", command=lambda: refresh_tree())
            calc_btn.pack(side=tk.LEFT, padx=2)
            ttk.Label(btn_frame, text="BOM:").pack(side=tk.LEFT, padx=2)
            bom_var = tk.StringVar(value=fmea.get('bom', ''))
            bom_combo = ttk.Combobox(
                btn_frame,
                textvariable=bom_var,
                values=[ra.name for ra in self.reliability_analyses],
                state="readonly",
                width=20,
            )
            bom_combo.pack(side=tk.LEFT, padx=2)

            def add_component():
                dlg = tk.Toplevel(win)
                dlg.title("New Component")
                ttk.Label(dlg, text="Name").grid(row=0, column=0, padx=5, pady=5, sticky="e")
                name_var = tk.StringVar()
                ttk.Entry(dlg, textvariable=name_var).grid(row=0, column=1, padx=5, pady=5)
                ttk.Label(dlg, text="Type").grid(row=1, column=0, padx=5, pady=5, sticky="e")
                type_var = tk.StringVar(value="capacitor")
                ttk.Combobox(
                    dlg,
                    textvariable=type_var,
                    values=list(COMPONENT_ATTR_TEMPLATES.keys()),
                    state="readonly",
                ).grid(row=1, column=1, padx=5, pady=5)
                ttk.Label(dlg, text="Quantity").grid(row=2, column=0, padx=5, pady=5, sticky="e")
                qty_var = tk.IntVar(value=1)
                ttk.Entry(dlg, textvariable=qty_var).grid(row=2, column=1, padx=5, pady=5)
                ttk.Label(dlg, text="Qualification").grid(row=3, column=0, padx=5, pady=5, sticky="e")
                qual_var = tk.StringVar(value="None")
                ttk.Combobox(dlg, textvariable=qual_var, values=QUALIFICATIONS, state="readonly").grid(row=3, column=1, padx=5, pady=5)
                passive_var = tk.BooleanVar(value=False)
                ttk.Checkbutton(dlg, text="Passive", variable=passive_var).grid(row=4, column=0, columnspan=2, pady=5)

                def ok():
                    comp = ReliabilityComponent(
                        name_var.get(),
                        type_var.get(),
                        qty_var.get(),
                        {},
                        qual_var.get(),
                        is_passive=passive_var.get(),
                    )
                    template = COMPONENT_ATTR_TEMPLATES.get(comp.comp_type, {})
                    for k, v in template.items():
                        comp.attributes[k] = v[0] if isinstance(v, list) else v
                    self.reliability_components.append(comp)
                    dlg.destroy()
                    refresh_tree()

                ttk.Button(dlg, text="Add", command=ok).grid(row=5, column=0, columnspan=2, pady=5)
                dlg.grab_set()
                dlg.wait_window()

            ttk.Button(btn_frame, text="Add Component", command=add_component).pack(side=tk.LEFT, padx=2)

            selected_libs = self.selected_mechanism_libraries

            def choose_libs():
                dlg = tk.Toplevel(win)
                dlg.title("Select Libraries")
                vars = {}
                for i, lib in enumerate(self.mechanism_libraries):
                    var = tk.BooleanVar(value=lib in selected_libs)
                    tk.Checkbutton(dlg, text=lib.name, variable=var).pack(anchor="w")
                    vars[i] = (var, lib)

                def ok():
                    selected_libs.clear()
                    for _, (v, lib) in vars.items():
                        if v.get():
                            selected_libs.append(lib)
                    dlg.destroy()

                ttk.Button(dlg, text="OK", command=ok).pack(pady=5)
                dlg.grab_set()
                dlg.wait_window()

            ttk.Button(btn_frame, text="Libraries", command=choose_libs).pack(side=tk.LEFT, padx=2)

            def load_bom(*_):
                name = bom_var.get()
                ra = next((r for r in self.reliability_analyses if r.name == name), None)
                if ra:
                    self.reliability_components = copy.deepcopy(ra.components)
                    self.reliability_total_fit = ra.total_fit
                    self.spfm = ra.spfm
                    self.lpfm = ra.lpfm
                    if fmea is not None:
                        fmea['bom'] = name
                    refresh_tree()

            bom_combo.bind("<<ComboboxSelected>>", load_bom)
            if bom_var.get():
                load_bom()

        tree_frame = ttk.Frame(win)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        tree = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="tree headings",
            style="FMEA.Treeview",
        )
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for col in columns:
            tree.heading(col, text=col)
            width = 120
            if col in ["Requirements", "Failure Effect", "Cause", "Safety Goal", "Malfunction"]:
                width = 200
            elif col == "Parent":
                width = 150
            elif col in ["FaultType", "Fraction", "FIT", "DiagCov", "Mechanism"]:
                width = 80
            tree.column(col, width=width, anchor="center")
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        metrics_lbl = ttk.Label(win, text="")
        metrics_lbl.pack(anchor="w", padx=5, pady=2)

        # alternating row colours and high RPN highlight
        tree.tag_configure("component", background="#e2e2e2", font=("Segoe UI", 10, "bold"))
        tree.tag_configure("evenrow", background="#ffffff")
        tree.tag_configure("oddrow", background="#f5f5f5")
        tree.tag_configure("highrpn", background="#ffe6e6")

        node_map = {}
        comp_items = {}

        def refresh_tree():
            tree.delete(*tree.get_children())
            node_map.clear()
            comp_items.clear()
            # remove any duplicate nodes based on unique_id
            unique = {}
            for be in entries:
                unique[be.unique_id] = be
            entries[:] = list(unique.values())
            events = entries

            comp_fit = {c.name: c.fit * c.quantity for c in self.reliability_components}
            frac_totals = {}
            for be in events:
                src = self.get_failure_mode_node(be)
                comp_name = (
                    src.parents[0].user_name if src.parents else getattr(src, "fmea_component", "")
                )
                fit = comp_fit.get(comp_name, 0.0)
                frac = src.fmeda_fault_fraction
                if frac > 1.0:
                    frac /= 100.0
                be.fmeda_fit = fit * frac
                if src.fmeda_fault_type == "permanent":
                    be.fmeda_spfm = be.fmeda_fit * (1 - src.fmeda_diag_cov)
                    be.fmeda_lpfm = 0.0
                else:
                    be.fmeda_lpfm = be.fmeda_fit * (1 - src.fmeda_diag_cov)
                    be.fmeda_spfm = 0.0
                frac_totals[comp_name] = frac_totals.get(comp_name, 0.0) + frac

            warnings = [f"{name} fractions={val:.2f}" for name, val in frac_totals.items() if abs(val - 1.0) > 0.01]
            if warnings:
                messagebox.showwarning("Distribution", "Fault fraction sum != 1:\n" + "\n".join(warnings))

            for idx, be in enumerate(events):
                src = self.get_failure_mode_node(be)
                parent = src.parents[0] if src.parents else None
                if parent:
                    comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                    parent_name = comp
                else:
                    comp = getattr(src, "fmea_component", "") or "N/A"
                    parent_name = ""
                if comp not in comp_items:
                    comp_items[comp] = tree.insert(
                        "",
                        "end",
                        text=comp,
                        values=[comp, "", "", "", "", "", "", "", "", ""],
                        tags=("component",),
                    )
                comp_iid = comp_items[comp]
                req_ids = "; ".join(
                    [f"{req['req_type']}:{req['text']}" for req in getattr(src, "safety_requirements", [])]
                )
                rpn = src.fmea_severity * src.fmea_occurrence * src.fmea_detection
                failure_mode = src.description or (src.user_name or f"BE {src.unique_id}")
                vals = [
                    "",
                    parent_name,
                    failure_mode,
                    src.fmea_effect,
                    src.fmea_cause,
                    src.fmea_severity,
                    src.fmea_occurrence,
                    src.fmea_detection,
                    rpn,
                    req_ids,
                ]
                if fmeda:
                    vals.extend([
                        src.fmeda_malfunction,
                        src.fmeda_safety_goal,
                        src.fmeda_fault_type,
                        f"{src.fmeda_fault_fraction:.2f}",
                        f"{src.fmeda_fit:.2f}",
                        f"{src.fmeda_diag_cov:.2f}",
                        getattr(src, "fmeda_mechanism", ""),
                    ])
                tags = ["evenrow" if idx % 2 == 0 else "oddrow"]
                if rpn >= 100:
                    tags.append("highrpn")
                iid = tree.insert(comp_iid, "end", text="", values=vals, tags=tags)
                node_map[iid] = be
            for iid in comp_items.values():
                tree.item(iid, open=True)

            if fmeda:
                total = 0.0
                spf = 0.0
                lpf = 0.0
                asil = "QM"
                for be in events:
                    src = self.get_failure_mode_node(be)
                    fit_mode = be.fmeda_fit
                    total += fit_mode
                    if src.fmeda_fault_type == "permanent":
                        spf += fit_mode * (1 - src.fmeda_diag_cov)
                    else:
                        lpf += fit_mode * (1 - src.fmeda_diag_cov)
                    sg = getattr(src, "fmeda_safety_goal", "")
                    a = self.get_safety_goal_asil(sg)
                    if ASIL_ORDER.get(a,0) > ASIL_ORDER.get(asil,0):
                        asil = a
                dc = (total - (spf + lpf)) / total if total else 0.0
                self.reliability_total_fit = total
                self.spfm = spf
                self.lpfm = lpf
                spfm_metric = 1 - spf / total if total else 0.0
                lpfm_metric = 1 - lpf / (total - spf) if total > spf else 0.0
                thresh = ASIL_TARGETS.get(asil, ASIL_TARGETS["QM"])
                ok_dc = dc >= thresh["dc"]
                ok_spf = spfm_metric >= thresh["spfm"]
                ok_lpf = lpfm_metric >= thresh["lpfm"]
                text = (
                    f"Total FIT: {total:.2f}  DC: {dc:.2f}{CHECK_MARK if ok_dc else CROSS_MARK}"
                    f"  SPFM: {spfm_metric:.2f}{CHECK_MARK if ok_spf else CROSS_MARK}"
                    f"  LPFM: {lpfm_metric:.2f}{CHECK_MARK if ok_lpf else CROSS_MARK}"
                    f"  (ASIL {asil})"
                )
                metrics_lbl.config(text=text)

        refresh_tree()

        def on_double(event):
            sel = tree.focus()
            node = node_map.get(sel)
            if node:
                mechs = []
                for lib in selected_libs:
                    mechs.extend(lib.mechanisms)
                comp_name = node.parents[0].user_name if node.parents else getattr(node, "fmea_component", "")
                is_passive = any(c.name == comp_name and c.is_passive for c in self.reliability_components)
                self.FMEARowDialog(win, node, self, entries, mechanisms=mechs, hide_diagnostics=is_passive)
                refresh_tree()

        tree.bind("<Double-1>", on_double)

        def add_failure_mode():
            dialog = self.SelectBaseEventDialog(win, basic_events, allow_new=True)
            node = dialog.selected
            if node == "NEW":
                node = FaultTreeNode("", "Basic Event")
                entries.append(node)
                mechs = []
                for lib in selected_libs:
                    mechs.extend(lib.mechanisms)
                comp_name = getattr(node, "fmea_component", "")
                is_passive = any(c.name == comp_name and c.is_passive for c in self.reliability_components)
                self.FMEARowDialog(win, node, self, entries, mechanisms=mechs, hide_diagnostics=is_passive)
            elif node:
                # gather all failure modes under the same component/parent
                if node.parents:
                    parent_id = node.parents[0].unique_id
                    related = [
                        be
                        for be in basic_events
                        if be.parents and be.parents[0].unique_id == parent_id
                    ]
                else:
                    comp = getattr(node, "fmea_component", "")
                    related = [
                        be
                        for be in basic_events
                        if not be.parents and getattr(be, "fmea_component", "") == comp
                    ]
                if node not in related:
                    related.append(node)
                existing_ids = {be.unique_id for be in entries}
                for be in related:
                    if be.unique_id not in existing_ids:
                        entries.append(be)
                        existing_ids.add(be.unique_id)
                    mechs = []
                    for lib in selected_libs:
                        mechs.extend(lib.mechanisms)
                    comp_name = be.parents[0].user_name if be.parents else getattr(be, "fmea_component", "")
                    is_passive = any(c.name == comp_name and c.is_passive for c in self.reliability_components)
                    self.FMEARowDialog(win, be, self, entries, mechanisms=mechs, hide_diagnostics=is_passive)
            refresh_tree()

        add_btn.config(command=add_failure_mode)

        def remove_from_fmea():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Remove Entry", "Select a row to remove.")
                return
            for iid in sel:
                node = node_map.get(iid)
                if node in entries:
                    entries.remove(node)
            refresh_tree()

        remove_btn.config(command=remove_from_fmea)

        def delete_failure_mode():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Delete Failure Mode", "Select a row to delete.")
                return
            if not messagebox.askyesno("Delete Failure Mode", "Remove selected failure modes from the FMEA?"):
                return
            for iid in sel:
                node = node_map.get(iid)
                if node in entries:
                    entries.remove(node)
            refresh_tree()

        del_btn.config(command=delete_failure_mode)

        def comment_fmea_entry():
            sel = tree.selection()
            if not sel:
                messagebox.showwarning("Comment", "Select a row to comment.")
                return
            node = node_map.get(sel[0])
            if not node:
                return
            self.selected_node = node
            self.comment_target = ("fmea", node.unique_id)
            self.open_review_toolbox()

        comment_btn.config(command=comment_fmea_entry)

        def on_close():
            if fmea is not None:
                if fmeda:
                    self.export_fmeda_to_csv(fmea, fmea['file'])
                else:
                    self.export_fmea_to_csv(fmea, fmea['file'])
                if fmeda:
                    fmea['bom'] = bom_var.get()
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", on_close)

    def export_fmea_to_csv(self, fmea, path):
        columns = ["Component", "Parent", "Failure Mode", "Failure Effect", "Cause", "S", "O", "D", "RPN", "Requirements"]
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for be in fmea['entries']:
                parent = be.parents[0] if be.parents else None
                if parent:
                    comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                    if parent.description:
                        comp = f"{comp} - {parent.description}"
                    parent_name = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                else:
                    comp = getattr(be, "fmea_component", "") or "N/A"
                    parent_name = ""
                req_ids = "; ".join([f"{req['req_type']}:{req['text']}" for req in getattr(be, 'safety_requirements', [])])
                rpn = be.fmea_severity * be.fmea_occurrence * be.fmea_detection
                failure_mode = be.description or (be.user_name or f"BE {be.unique_id}")
                row = [comp, parent_name, failure_mode, be.fmea_effect, be.fmea_cause, be.fmea_severity, be.fmea_occurrence, be.fmea_detection, rpn, req_ids]
                writer.writerow(row)

    def export_fmeda_to_csv(self, fmeda, path):
        columns = [
            "Component",
            "Parent",
            "Failure Mode",
            "Failure Effect",
            "Cause",
            "S",
            "O",
            "D",
            "RPN",
            "Requirements",
            "Malfunction",
            "Safety Goal",
            "FaultType",
            "Fraction",
            "FIT",
            "DiagCov",
            "Mechanism",
        ]
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for be in fmeda['entries']:
                parent = be.parents[0] if be.parents else None
                if parent:
                    comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                    if parent.description:
                        comp = f"{comp} - {parent.description}"
                    parent_name = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                else:
                    comp = getattr(be, "fmea_component", "") or "N/A"
                    parent_name = ""
                req_ids = "; ".join([f"{req['req_type']}:{req['text']}" for req in getattr(be, 'safety_requirements', [])])
                rpn = be.fmea_severity * be.fmea_occurrence * be.fmea_detection
                failure_mode = be.description or (be.user_name or f"BE {be.unique_id}")
                row = [
                    comp,
                    parent_name,
                    failure_mode,
                    be.fmea_effect,
                    be.fmea_cause,
                    be.fmea_severity,
                    be.fmea_occurrence,
                    be.fmea_detection,
                    rpn,
                    req_ids,
                    getattr(be, "fmeda_malfunction", ""),
                    getattr(be, "fmeda_safety_goal", ""),
                    getattr(be, "fmeda_fault_type", ""),
                    be.fmeda_fault_fraction,
                    be.fmeda_fit,
                    be.fmeda_diag_cov,
                    getattr(be, "fmeda_mechanism", ""),
                ]
                writer.writerow(row)

    def export_fmeda_to_csv(self, fmeda, path):
        columns = [
            "Component",
            "Parent",
            "Failure Mode",
            "Failure Effect",
            "Cause",
            "S",
            "O",
            "D",
            "RPN",
            "Requirements",
            "Malfunction",
            "Safety Goal",
            "FaultType",
            "Fraction",
            "FIT",
            "DiagCov",
            "Mechanism",
        ]
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for be in fmeda['entries']:
                parent = be.parents[0] if be.parents else None
                if parent:
                    comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                    if parent.description:
                        comp = f"{comp} - {parent.description}"
                    parent_name = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                else:
                    comp = getattr(be, "fmea_component", "") or "N/A"
                    parent_name = ""
                req_ids = "; ".join([f"{req['req_type']}:{req['text']}" for req in getattr(be, 'safety_requirements', [])])
                rpn = be.fmea_severity * be.fmea_occurrence * be.fmea_detection
                failure_mode = be.description or (be.user_name or f"BE {be.unique_id}")
                row = [
                    comp,
                    parent_name,
                    failure_mode,
                    be.fmea_effect,
                    be.fmea_cause,
                    be.fmea_severity,
                    be.fmea_occurrence,
                    be.fmea_detection,
                    rpn,
                    req_ids,
                    getattr(be, "fmeda_malfunction", ""),
                    getattr(be, "fmeda_safety_goal", ""),
                    getattr(be, "fmeda_fault_type", ""),
                    be.fmeda_fault_fraction,
                    be.fmeda_fit,
                    be.fmeda_diag_cov,
                    getattr(be, "fmeda_mechanism", ""),
                ]
                writer.writerow(row)


    def show_traceability_matrix(self):
        """Display a traceability matrix linking FTA basic events to FMEA components."""
        basic_events = [n for n in self.get_all_nodes(self.root_node)
                        if n.node_type.upper() == "BASIC EVENT"]
        win = tk.Toplevel(self.root)
        win.title("FTA-FMEA Traceability")
        columns = ["Basic Event", "Component"]
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=200, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        for be in basic_events:
            parent = be.parents[0] if be.parents else None
            if parent:
                if parent.user_name:
                    comp = parent.user_name
                else:
                    comp = "Node {}".format(parent.unique_id)
            else:
                comp = "N/A"
            tree.insert(
                "",
                "end",
                values=[be.user_name or f"BE {be.unique_id}", comp],
            )

    def collect_requirements_recursive(self, node):
        reqs = list(getattr(node, "safety_requirements", []))
        for child in node.children:
            reqs.extend(self.collect_requirements_recursive(child))
        return reqs

    def show_safety_goals_matrix(self):
        """Display safety goals and derived requirements in a tree view."""
        win = tk.Toplevel(self.root)
        win.title("Safety Goals Matrix")
        tree = ttk.Treeview(win, columns=["ID", "ASIL", "SafeState", "Text"], show="tree headings")
        tree.heading("ID", text="Requirement ID")
        tree.heading("ASIL", text="ASIL")
        tree.heading("SafeState", text="Safe State")
        tree.heading("Text", text="Text")
        tree.column("ID", width=120)
        tree.column("ASIL", width=60)
        tree.column("SafeState", width=100)
        tree.column("Text", width=300)
        tree.pack(fill=tk.BOTH, expand=True)

        for te in self.top_events:
            sg_text = te.safety_goal_description or (te.user_name or f"SG {te.unique_id}")
            sg_id = te.user_name or f"SG {te.unique_id}"
            parent_iid = tree.insert(
                "", "end", text=sg_text,
                values=[sg_id, te.safety_goal_asil, te.safe_state, sg_text],
            )
            reqs = self.collect_requirements_recursive(te)
            seen_ids = set()
            for req in reqs:
                req_id = req.get("id")
                if req_id in seen_ids:
                    continue
                seen_ids.add(req_id)
                tree.insert(
                    parent_iid,
                    "end",
                    text="",
                    values=[req_id, req.get("asil", ""), req.get("text", "")],
                )

    def show_safety_goals_editor(self):
        """Allow editing of top-level safety goals."""
        win = tk.Toplevel(self.root)
        win.title("Safety Goals Editor")

        columns = ["ID", "ASIL", "Safe State", "FTTI", "Acceptance", "Description"]
        tree = ttk.Treeview(win, columns=columns, show="headings", selectmode="browse")
        for c in columns:
            tree.heading(c, text=c)
            tree.column(c, width=120 if c != "Description" else 300, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh_tree():
            tree.delete(*tree.get_children())
            for sg in self.top_events:
                tree.insert(
                    "",
                    "end",
                    iid=sg.unique_id,
                    values=[
                        sg.user_name or f"SG {sg.unique_id}",
                        sg.safety_goal_asil,
                        sg.safe_state,
                        getattr(sg, "ftti", ""),
                        getattr(sg, "acceptance_criteria", ""),
                        sg.safety_goal_description,
                    ],
                )

        class SGDialog(simpledialog.Dialog):
            def __init__(self, parent, title, initial=None):
                self.initial = initial
                super().__init__(parent, title=title)

            def body(self, master):
                ttk.Label(master, text="ID:").grid(row=0, column=0, sticky="e")
                self.id_var = tk.StringVar(value=getattr(self.initial, "user_name", ""))
                tk.Entry(master, textvariable=self.id_var).grid(row=0, column=1, padx=5, pady=5)

                ttk.Label(master, text="ASIL:").grid(row=1, column=0, sticky="e")
                self.asil_var = tk.StringVar(value=getattr(self.initial, "safety_goal_asil", "QM"))
                ttk.Combobox(master, textvariable=self.asil_var, values=ASIL_LEVEL_OPTIONS, state="readonly", width=8).grid(row=1, column=1, padx=5, pady=5)

                ttk.Label(master, text="Safe State:").grid(row=2, column=0, sticky="e")
                self.state_var = tk.StringVar(value=getattr(self.initial, "safe_state", ""))
                tk.Entry(master, textvariable=self.state_var).grid(row=2, column=1, padx=5, pady=5)

                ttk.Label(master, text="FTTI:").grid(row=3, column=0, sticky="e")
                self.ftti_var = tk.StringVar(value=getattr(self.initial, "ftti", ""))
                tk.Entry(master, textvariable=self.ftti_var).grid(row=3, column=1, padx=5, pady=5)

                ttk.Label(master, text="Acceptance Criteria:").grid(row=4, column=0, sticky="ne")
                self.acc_text = tk.Text(master, width=30, height=3, wrap="word")
                self.acc_text.insert("1.0", getattr(self.initial, "acceptance_criteria", ""))
                self.acc_text.grid(row=4, column=1, padx=5, pady=5)

                ttk.Label(master, text="Description:").grid(row=5, column=0, sticky="ne")
                self.desc_text = tk.Text(master, width=30, height=3, wrap="word")
                self.desc_text.insert("1.0", getattr(self.initial, "safety_goal_description", ""))
                self.desc_text.grid(row=5, column=1, padx=5, pady=5)
                return master

            def apply(self):
                self.result = {
                    "id": self.id_var.get().strip(),
                    "asil": self.asil_var.get().strip(),
                    "state": self.state_var.get().strip(),
                    "ftti": self.ftti_var.get().strip(),
                    "accept": self.acc_text.get("1.0", "end-1c"),
                    "desc": self.desc_text.get("1.0", "end-1c"),
                }

        def add_sg():
            dlg = SGDialog(win, "Add Safety Goal")
            if dlg.result:
                node = FaultTreeNode(dlg.result["id"], "TOP EVENT")
                node.safety_goal_asil = dlg.result["asil"]
                node.safe_state = dlg.result["state"]
                node.ftti = dlg.result["ftti"]
                node.acceptance_criteria = dlg.result["accept"]
                node.safety_goal_description = dlg.result["desc"]
                self.top_events.append(node)
                refresh_tree()
                self.update_views()

        def edit_sg():
            sel = tree.selection()
            if not sel:
                return
            uid = int(sel[0])
            sg = self.find_node_by_id_all(uid)
            dlg = SGDialog(win, "Edit Safety Goal", sg)
            if dlg.result:
                sg.user_name = dlg.result["id"]
                sg.safety_goal_asil = dlg.result["asil"]
                sg.safe_state = dlg.result["state"]
                sg.ftti = dlg.result["ftti"]
                sg.acceptance_criteria = dlg.result["accept"]
                sg.safety_goal_description = dlg.result["desc"]
                refresh_tree()
                self.update_views()

        def del_sg():
            sel = tree.selection()
            if not sel:
                return
            uid = int(sel[0])
            sg = self.find_node_by_id_all(uid)
            if sg and messagebox.askyesno("Delete", "Delete safety goal?"):
                self.top_events = [t for t in self.top_events if t.unique_id != uid]
                refresh_tree()
                self.update_views()

        btn = ttk.Frame(win)
        btn.pack(fill=tk.X)
        ttk.Button(btn, text="Add", command=add_sg).pack(side=tk.LEFT)
        ttk.Button(btn, text="Edit", command=edit_sg).pack(side=tk.LEFT)
        ttk.Button(btn, text="Delete", command=del_sg).pack(side=tk.LEFT)

        refresh_tree()

    def export_safety_goal_requirements(self):
        """Export requirements traced to safety goals including their ASIL."""
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return

        columns = ["Safety Goal", "SG ASIL", "Safe State", "Requirement ID", "Req ASIL", "Text"]
        with open(path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(columns)
            for te in self.top_events:
                sg_text = te.safety_goal_description or (te.user_name or f"SG {te.unique_id}")
                sg_asil = te.safety_goal_asil
                reqs = self.collect_requirements_recursive(te)
                seen = set()
                for req in reqs:
                    rid = req.get("id")
                    if rid in seen:
                        continue
                    seen.add(rid)
                    writer.writerow([sg_text, sg_asil, te.safe_state, rid, req.get("asil", ""), req.get("text", "")])
        messagebox.showinfo("Export", "Safety goal requirements exported.")

    def show_cut_sets(self):
        """Display minimal cut sets for every top event."""
        if not self.top_events:
            return
        win = tk.Toplevel(self.root)
        win.title("FTA Cut Sets")
        columns = ("Top Event", "Cut Set #", "Basic Events")
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for c in columns:
            tree.heading(c, text=c)
        tree.pack(fill=tk.BOTH, expand=True)

        for te in self.top_events:
            nodes_by_id = {}

            def map_nodes(n):
                nodes_by_id[n.unique_id] = n
                for child in n.children:
                    map_nodes(child)

            map_nodes(te)
            cut_sets = self.calculate_cut_sets(te)
            te_label = te.user_name or f"Top Event {te.unique_id}"
            for idx, cs in enumerate(cut_sets, start=1):
                names = ", ".join(
                    f"{nodes_by_id[uid].user_name or nodes_by_id[uid].node_type} [{uid}]"
                    for uid in sorted(cs)
                )
                tree.insert("", "end", values=(te_label, idx, names))
                te_label = ""

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Top Event", "Cut Set #", "Basic Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Cut sets exported")

        ttk.Button(win, text="Export CSV", command=export_csv).pack(pady=5)

    def show_common_cause_view(self):
        win = tk.Toplevel(self.root)
        win.title("Common Cause Toolbox")
        var_fmea = tk.BooleanVar(value=True)
        var_fmeda = tk.BooleanVar(value=True)
        var_fta = tk.BooleanVar(value=True)
        chk_frame = ttk.Frame(win)
        chk_frame.pack(anchor="w")
        ttk.Checkbutton(chk_frame, text="FMEA", variable=var_fmea).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FMEDA", variable=var_fmeda).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FTA", variable=var_fta).pack(side=tk.LEFT)
        tree = ttk.Treeview(win, columns=["Cause", "Events"], show="headings")
        for c in ["Cause", "Events"]:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh():
            tree.delete(*tree.get_children())
            events_by_cause = {}
            if var_fmea.get():
                for fmea in self.fmeas:
                    for be in fmea["entries"]:
                        cause = be.description
                        label = f"{fmea['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fmeda.get():
                for fmeda in self.fmedas:
                    for be in fmeda["entries"]:
                        cause = be.description
                        label = f"{fmeda['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fta.get():
                for be in self.get_all_basic_events():
                    cause = be.description or ""
                    label = be.user_name or f"BE {be.unique_id}"
                    events_by_cause.setdefault(cause, set()).add(label)
            for cause, evts in events_by_cause.items():
                if len(evts) > 1:
                    tree.insert("", "end", values=[cause, ", ".join(sorted(evts))])

        refresh()

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Cause", "Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Common cause data exported")

        btn_frame = ttk.Frame(win)
        btn_frame.pack()
        ttk.Button(btn_frame, text="Refresh", command=refresh).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(btn_frame, text="Export CSV", command=export_csv).pack(side=tk.LEFT, padx=5, pady=5)

    def show_cut_sets(self):
        """Display minimal cut sets for every top event."""
        if not self.top_events:
            return
        win = tk.Toplevel(self.root)
        win.title("FTA Cut Sets")
        columns = ("Top Event", "Cut Set #", "Basic Events")
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for c in columns:
            tree.heading(c, text=c)
        tree.pack(fill=tk.BOTH, expand=True)

        for te in self.top_events:
            nodes_by_id = {}

            def map_nodes(n):
                nodes_by_id[n.unique_id] = n
                for child in n.children:
                    map_nodes(child)

            map_nodes(te)
            cut_sets = self.calculate_cut_sets(te)
            te_label = te.user_name or f"Top Event {te.unique_id}"
            for idx, cs in enumerate(cut_sets, start=1):
                names = ", ".join(
                    f"{nodes_by_id[uid].user_name or nodes_by_id[uid].node_type} [{uid}]"
                    for uid in sorted(cs)
                )
                tree.insert("", "end", values=(te_label, idx, names))
                te_label = ""

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Top Event", "Cut Set #", "Basic Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Cut sets exported")

        ttk.Button(win, text="Export CSV", command=export_csv).pack(pady=5)

    def show_common_cause_view(self):
        win = tk.Toplevel(self.root)
        win.title("Common Cause Toolbox")
        var_fmea = tk.BooleanVar(value=True)
        var_fmeda = tk.BooleanVar(value=True)
        var_fta = tk.BooleanVar(value=True)
        chk_frame = ttk.Frame(win)
        chk_frame.pack(anchor="w")
        ttk.Checkbutton(chk_frame, text="FMEA", variable=var_fmea).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FMEDA", variable=var_fmeda).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FTA", variable=var_fta).pack(side=tk.LEFT)
        tree = ttk.Treeview(win, columns=["Cause", "Events"], show="headings")
        for c in ["Cause", "Events"]:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh():
            tree.delete(*tree.get_children())
            events_by_cause = {}
            if var_fmea.get():
                for fmea in self.fmeas:
                    for be in fmea["entries"]:
                        cause = be.description
                        label = f"{fmea['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fmeda.get():
                for fmeda in self.fmedas:
                    for be in fmeda["entries"]:
                        cause = be.description
                        label = f"{fmeda['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fta.get():
                for be in self.get_all_basic_events():
                    cause = be.description or ""
                    label = be.user_name or f"BE {be.unique_id}"
                    events_by_cause.setdefault(cause, set()).add(label)
            for cause, evts in events_by_cause.items():
                if len(evts) > 1:
                    tree.insert("", "end", values=[cause, ", ".join(sorted(evts))])

        refresh()

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Cause", "Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Common cause data exported")

        btn_frame = ttk.Frame(win)
        btn_frame.pack()
        ttk.Button(btn_frame, text="Refresh", command=refresh).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(btn_frame, text="Export CSV", command=export_csv).pack(side=tk.LEFT, padx=5, pady=5)

    def show_cut_sets(self):
        """Display minimal cut sets for every top event."""
        if not self.top_events:
            return
        win = tk.Toplevel(self.root)
        win.title("FTA Cut Sets")
        columns = ("Top Event", "Cut Set #", "Basic Events")
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for c in columns:
            tree.heading(c, text=c)
        tree.pack(fill=tk.BOTH, expand=True)

        for te in self.top_events:
            nodes_by_id = {}

            def map_nodes(n):
                nodes_by_id[n.unique_id] = n
                for child in n.children:
                    map_nodes(child)

            map_nodes(te)
            cut_sets = self.calculate_cut_sets(te)
            te_label = te.user_name or f"Top Event {te.unique_id}"
            for idx, cs in enumerate(cut_sets, start=1):
                names = ", ".join(
                    f"{nodes_by_id[uid].user_name or nodes_by_id[uid].node_type} [{uid}]"
                    for uid in sorted(cs)
                )
                tree.insert("", "end", values=(te_label, idx, names))
                te_label = ""

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Top Event", "Cut Set #", "Basic Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Cut sets exported")

        ttk.Button(win, text="Export CSV", command=export_csv).pack(pady=5)

    def show_common_cause_view(self):
        win = tk.Toplevel(self.root)
        win.title("Common Cause Toolbox")
        var_fmea = tk.BooleanVar(value=True)
        var_fmeda = tk.BooleanVar(value=True)
        var_fta = tk.BooleanVar(value=True)
        chk_frame = ttk.Frame(win)
        chk_frame.pack(anchor="w")
        ttk.Checkbutton(chk_frame, text="FMEA", variable=var_fmea).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FMEDA", variable=var_fmeda).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FTA", variable=var_fta).pack(side=tk.LEFT)
        tree = ttk.Treeview(win, columns=["Cause", "Events"], show="headings")
        for c in ["Cause", "Events"]:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh():
            tree.delete(*tree.get_children())
            events_by_cause = {}
            if var_fmea.get():
                for fmea in self.fmeas:
                    for be in fmea["entries"]:
                        cause = be.description
                        label = f"{fmea['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fmeda.get():
                for fmeda in self.fmedas:
                    for be in fmeda["entries"]:
                        cause = be.description
                        label = f"{fmeda['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fta.get():
                for be in self.get_all_basic_events():
                    cause = be.description or ""
                    label = be.user_name or f"BE {be.unique_id}"
                    events_by_cause.setdefault(cause, set()).add(label)
            for cause, evts in events_by_cause.items():
                if len(evts) > 1:
                    tree.insert("", "end", values=[cause, ", ".join(sorted(evts))])

        refresh()

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Cause", "Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Common cause data exported")

        btn_frame = ttk.Frame(win)
        btn_frame.pack()
        ttk.Button(btn_frame, text="Refresh", command=refresh).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(btn_frame, text="Export CSV", command=export_csv).pack(side=tk.LEFT, padx=5, pady=5)

    def show_cut_sets(self):
        """Display minimal cut sets for every top event."""
        if not self.top_events:
            return
        win = tk.Toplevel(self.root)
        win.title("FTA Cut Sets")
        columns = ("Top Event", "Cut Set #", "Basic Events")
        tree = ttk.Treeview(win, columns=columns, show="headings")
        for c in columns:
            tree.heading(c, text=c)
        tree.pack(fill=tk.BOTH, expand=True)

        for te in self.top_events:
            nodes_by_id = {}

            def map_nodes(n):
                nodes_by_id[n.unique_id] = n
                for child in n.children:
                    map_nodes(child)

            map_nodes(te)
            cut_sets = self.calculate_cut_sets(te)
            te_label = te.user_name or f"Top Event {te.unique_id}"
            for idx, cs in enumerate(cut_sets, start=1):
                names = ", ".join(
                    f"{nodes_by_id[uid].user_name or nodes_by_id[uid].node_type} [{uid}]"
                    for uid in sorted(cs)
                )
                tree.insert("", "end", values=(te_label, idx, names))
                te_label = ""

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Top Event", "Cut Set #", "Basic Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Cut sets exported")

        ttk.Button(win, text="Export CSV", command=export_csv).pack(pady=5)

    def show_common_cause_view(self):
        win = tk.Toplevel(self.root)
        win.title("Common Cause Toolbox")
        var_fmea = tk.BooleanVar(value=True)
        var_fmeda = tk.BooleanVar(value=True)
        var_fta = tk.BooleanVar(value=True)
        chk_frame = ttk.Frame(win)
        chk_frame.pack(anchor="w")
        ttk.Checkbutton(chk_frame, text="FMEA", variable=var_fmea).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FMEDA", variable=var_fmeda).pack(side=tk.LEFT)
        ttk.Checkbutton(chk_frame, text="FTA", variable=var_fta).pack(side=tk.LEFT)
        tree = ttk.Treeview(win, columns=["Cause", "Events"], show="headings")
        for c in ["Cause", "Events"]:
            tree.heading(c, text=c)
            tree.column(c, width=150)
        tree.pack(fill=tk.BOTH, expand=True)

        def refresh():
            tree.delete(*tree.get_children())
            events_by_cause = {}
            if var_fmea.get():
                for fmea in self.fmeas:
                    for be in fmea["entries"]:
                        cause = be.description
                        label = f"{fmea['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fmeda.get():
                for fmeda in self.fmedas:
                    for be in fmeda["entries"]:
                        cause = be.description
                        label = f"{fmeda['name']}:{be.user_name or be.description or be.unique_id}"
                        events_by_cause.setdefault(cause, set()).add(label)
            if var_fta.get():
                for be in self.get_all_basic_events():
                    cause = be.description or ""
                    label = be.user_name or f"BE {be.unique_id}"
                    events_by_cause.setdefault(cause, set()).add(label)
            for cause, evts in events_by_cause.items():
                if len(evts) > 1:
                    tree.insert("", "end", values=[cause, ", ".join(sorted(evts))])

        refresh()

        def export_csv():
            path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
            if not path:
                return
            with open(path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["Cause", "Events"])
                for iid in tree.get_children():
                    writer.writerow(tree.item(iid, "values"))
            messagebox.showinfo("Export", "Common cause data exported")

        btn_frame = ttk.Frame(win)
        btn_frame.pack()
        ttk.Button(btn_frame, text="Refresh", command=refresh).pack(side=tk.LEFT, padx=5, pady=5)
        ttk.Button(btn_frame, text="Export CSV", command=export_csv).pack(side=tk.LEFT, padx=5, pady=5)

    def manage_mission_profiles(self):
        win = tk.Toplevel(self.root)
        win.title("Mission Profiles")
        listbox = tk.Listbox(win, height=8, width=40)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        btn_frame = ttk.Frame(win)
        btn_frame.pack(side=tk.RIGHT, fill=tk.Y)

        def refresh():
            listbox.delete(0, tk.END)
            for mp in self.mission_profiles:
                if mp is None:
                    continue
                info = (
                    f"{mp.name} (on: {mp.tau_on}h, off: {mp.tau_off}h, "
                    f"board: {mp.board_temp}\u00b0C, ambient: {mp.ambient_temp}\u00b0C)"
                )
                listbox.insert(tk.END, info)

        class MPDialog(simpledialog.Dialog):
            def __init__(self, master, mp=None):
                self.mp = mp
                super().__init__(master)

            def body(self, master):
                self.vars = {}
                fields = [
                    ("Name", "name"),
                    ("TAU On (h)", "tau_on"),
                    ("TAU Off (h)", "tau_off"),
                    ("Board Temp (\u00b0C)", "board_temp"),
                    ("Board Temp Min (\u00b0C)", "board_temp_min"),
                    ("Board Temp Max (\u00b0C)", "board_temp_max"),
                    ("Ambient Temp (\u00b0C)", "ambient_temp"),
                    ("Ambient Temp Min (\u00b0C)", "ambient_temp_min"),
                    ("Ambient Temp Max (\u00b0C)", "ambient_temp_max"),
                    ("Humidity (%)", "humidity"),
                    ("Duty Cycle", "duty_cycle"),
                    ("Notes", "notes"),
                ]
                self.entries = {}
                for row, (label, key) in enumerate(fields):
                    ttk.Label(master, text=label).grid(row=row, column=0, padx=5, pady=5, sticky="e")
                    var = tk.StringVar()
                    if self.mp:
                        var.set(str(getattr(self.mp, key)))
                    state = "readonly" if key == "duty_cycle" else "normal"
                    entry = ttk.Entry(master, textvariable=var, state=state)
                    entry.grid(row=row, column=1, padx=5, pady=5)
                    self.vars[key] = var
                    self.entries[key] = entry

                def update_dc(*_):
                    try:
                        on = float(self.vars["tau_on"].get() or 0)
                        off = float(self.vars["tau_off"].get() or 0)
                        total = on + off
                        dc = on / total if total else 0.0
                    except ValueError:
                        dc = 0.0
                    self.vars["duty_cycle"].set(str(dc))

                self.vars["tau_on"].trace_add("write", update_dc)
                self.vars["tau_off"].trace_add("write", update_dc)
                update_dc()

            def apply(self):
                vals = {k: v.get() for k, v in self.vars.items()}
                tau_on = float(vals.get("tau_on") or 0.0)
                tau_off = float(vals.get("tau_off") or 0.0)
                total = tau_on + tau_off
                dc = tau_on / total if total else 0.0
                if self.mp is None:
                    mp = MissionProfile(
                        vals["name"],
                        tau_on,
                        tau_off,
                        float(vals["board_temp"] or 25.0),
                        float(vals["board_temp_min"] or 25.0),
                        float(vals["board_temp_max"] or 25.0),
                        float(vals["ambient_temp"] or 25.0),
                        float(vals["ambient_temp_min"] or 25.0),
                        float(vals["ambient_temp_max"] or 25.0),
                        float(vals["humidity"] or 50.0),
                        dc,
                        vals["notes"],
                    )
                    self.result = mp
                else:
                    self.mp.name = vals["name"]
                    self.mp.tau_on = tau_on
                    self.mp.tau_off = tau_off
                    self.mp.board_temp = float(vals["board_temp"] or 25.0)
                    self.mp.board_temp_min = float(vals["board_temp_min"] or 25.0)
                    self.mp.board_temp_max = float(vals["board_temp_max"] or 25.0)
                    self.mp.ambient_temp = float(vals["ambient_temp"] or 25.0)
                    self.mp.ambient_temp_min = float(vals["ambient_temp_min"] or 25.0)
                    self.mp.ambient_temp_max = float(vals["ambient_temp_max"] or 25.0)
                    self.mp.humidity = float(vals["humidity"] or 50.0)
                    self.mp.duty_cycle = dc
                    self.mp.notes = vals["notes"]
                    self.result = self.mp

        def add_profile():
            dlg = MPDialog(win)
            if getattr(dlg, "result", None) is not None:
                self.mission_profiles.append(dlg.result)
                refresh()

        def edit_profile():
            sel = listbox.curselection()
            if not sel:
                return
            mp = self.mission_profiles[sel[0]]
            dlg = MPDialog(win, mp)
            if getattr(dlg, "result", None) is not None:
                refresh()

        def delete_profile():
            sel = listbox.curselection()
            if not sel:
                return
            del self.mission_profiles[sel[0]]
            refresh()

        ttk.Button(btn_frame, text="Add", command=add_profile).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Edit", command=edit_profile).pack(fill=tk.X)
        ttk.Button(btn_frame, text="Delete", command=delete_profile).pack(fill=tk.X)

        refresh()

    def load_default_mechanisms(self):
        if self.mechanism_libraries:
            return
        lib = MechanismLibrary("ISO 26262 Annex D", ANNEX_D_MECHANISMS.copy())
        self.mechanism_libraries.append(lib)

    def manage_mechanism_libraries(self):
        win = tk.Toplevel(self.root)
        win.title("Mechanism Libraries")
        lib_lb = tk.Listbox(win, height=8, width=25)
        lib_lb.grid(row=0, column=0, rowspan=4, sticky="nsew")
        mech_tree = ttk.Treeview(win, columns=("cov", "desc"), show="headings")
        mech_tree.heading("cov", text="Coverage")
        mech_tree.column("cov", width=80)
        mech_tree.heading("desc", text="Description")
        mech_tree.column("desc", width=200)
        mech_tree.grid(row=0, column=1, columnspan=3, sticky="nsew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(1, weight=1)

        def refresh_libs():
            lib_lb.delete(0, tk.END)
            for lib in self.mechanism_libraries:
                lib_lb.insert(tk.END, lib.name)
            refresh_mechs()

        def refresh_mechs(*_):
            mech_tree.delete(*mech_tree.get_children())
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.mechanism_libraries[sel[0]]
            for mech in lib.mechanisms:
                mech_tree.insert("", tk.END, values=(f"{mech.coverage:.2f}", mech.description), text=mech.name)

        def add_lib():
            name = simpledialog.askstring("New Library", "Library name:")
            if not name:
                return
            self.mechanism_libraries.append(MechanismLibrary(name))
            refresh_libs()

        def edit_lib():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.mechanism_libraries[sel[0]]
            name = simpledialog.askstring("Edit Library", "Library name:", initialvalue=lib.name)
            if name:
                lib.name = name
                refresh_libs()

        def del_lib():
            sel = lib_lb.curselection()
            if not sel:
                return
            del self.mechanism_libraries[sel[0]]
            refresh_libs()

        def add_mech():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.mechanism_libraries[sel[0]]
            dlg = simpledialog.Dialog(win, title="Add Mechanism")
            class MForm(simpledialog.Dialog):
                def body(self, master):
                    ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                    self.name_var = tk.StringVar()
                    ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1)
                    ttk.Label(master, text="Coverage").grid(row=1, column=0, sticky="e")
                    self.cov_var = tk.StringVar(value="1.0")
                    ttk.Entry(master, textvariable=self.cov_var).grid(row=1, column=1)
                    ttk.Label(master, text="Description").grid(row=2, column=0, sticky="e")
                    self.desc_var = tk.StringVar()
                    ttk.Entry(master, textvariable=self.desc_var).grid(row=2, column=1)

                def apply(self):
                    self.result = (
                        self.name_var.get(),
                        float(self.cov_var.get() or 1.0),
                        self.desc_var.get(),
                    )

            form = MForm(win)
            if hasattr(form, "result"):
                name, cov, desc = form.result
                lib.mechanisms.append(DiagnosticMechanism(name, cov, desc))
                refresh_mechs()

        def edit_mech():
            sel_lib = lib_lb.curselection()
            sel_mech = mech_tree.selection()
            if not sel_lib or not sel_mech:
                return
            lib = self.mechanism_libraries[sel_lib[0]]
            idx = mech_tree.index(sel_mech[0])
            mech = lib.mechanisms[idx]

            class MForm(simpledialog.Dialog):
                def body(self, master):
                    ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                    self.name_var = tk.StringVar(value=mech.name)
                    ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1)
                    ttk.Label(master, text="Coverage").grid(row=1, column=0, sticky="e")
                    self.cov_var = tk.StringVar(value=str(mech.coverage))
                    ttk.Entry(master, textvariable=self.cov_var).grid(row=1, column=1)
                    ttk.Label(master, text="Description").grid(row=2, column=0, sticky="e")
                    self.desc_var = tk.StringVar(value=mech.description)
                    ttk.Entry(master, textvariable=self.desc_var).grid(row=2, column=1)

                def apply(self):
                    mech.name = self.name_var.get()
                    mech.coverage = float(self.cov_var.get() or 1.0)
                    mech.description = self.desc_var.get()

            MForm(win)
            refresh_mechs()

        def del_mech():
            sel_lib = lib_lb.curselection()
            sel_mech = mech_tree.selection()
            if not sel_lib or not sel_mech:
                return
            lib = self.mechanism_libraries[sel_lib[0]]
            idx = mech_tree.index(sel_mech[0])
            del lib.mechanisms[idx]
            refresh_mechs()

        btnf = ttk.Frame(win)
        btnf.grid(row=1, column=1, columnspan=3, sticky="ew")
        ttk.Button(btnf, text="Add Lib", command=add_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Edit Lib", command=edit_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Lib", command=del_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Add Mech", command=add_mech).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="Edit Mech", command=edit_mech).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Mech", command=del_mech).pack(side=tk.LEFT)

        lib_lb.bind("<<ListboxSelect>>", refresh_mechs)
        refresh_libs()

    def manage_scenario_libraries(self):
        win = tk.Toplevel(self.root)
        win.title("Scenario Libraries")
        lib_lb = tk.Listbox(win, height=8, width=25)
        lib_lb.grid(row=0, column=0, rowspan=4, sticky="nsew")
        scen_tree = ttk.Treeview(
            win,
            columns=("beh", "sce", "tc", "fi", "desc"),
            show="tree headings",
        )
        scen_tree.heading("#0", text="Name")
        scen_tree.column("#0", width=150)
        scen_tree.heading("beh", text="Other Users")
        scen_tree.column("beh", width=140)
        scen_tree.heading("sce", text="Scenery")
        scen_tree.column("sce", width=140)
        scen_tree.heading("tc", text="TC")
        scen_tree.column("tc", width=80)
        scen_tree.heading("fi", text="FI")
        scen_tree.column("fi", width=80)
        scen_tree.heading("desc", text="Description")
        scen_tree.column("desc", width=200)
        scen_tree.grid(row=0, column=1, columnspan=3, sticky="nsew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(1, weight=1)

        def refresh_libs():
            lib_lb.delete(0, tk.END)
            for lib in self.scenario_libraries:
                lib_lb.insert(tk.END, lib.get("name", ""))
            refresh_scenarios()

        def refresh_scenarios(*_):
            scen_tree.delete(*scen_tree.get_children())
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.scenario_libraries[sel[0]]
            for sc in lib.get("scenarios", []):
                if isinstance(sc, dict):
                    name = sc.get("name", "")
                    beh = sc.get("behavior", "")
                    sce = sc.get("scenery", "")
                    tc = sc.get("tc", "")
                    fi = sc.get("fi", "")
                    desc = sc.get("description", "")
                else:
                    name = str(sc)
                    beh = sce = tc = fi = desc = ""
                scen_tree.insert("", tk.END, text=name, values=(beh, sce, tc, fi, desc))

        class LibraryDialog(simpledialog.Dialog):
            def __init__(self, parent, app, data=None):
                self.app = app
                self.data = data or {"name": "", "odds": []}
                super().__init__(parent, title="Edit Library")

            def body(self, master):
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")
                ttk.Label(master, text="ODD Libraries").grid(row=1, column=0, sticky="ne")
                self.lb = tk.Listbox(master, selectmode=tk.MULTIPLE, height=5)
                for i, lib in enumerate(self.app.odd_libraries):
                    self.lb.insert(tk.END, lib.get("name", ""))
                    if lib.get("name", "") in self.data.get("odds", []):
                        self.lb.selection_set(i)
                self.lb.grid(row=1, column=1, sticky="nsew")
                master.grid_rowconfigure(1, weight=1)
                master.grid_columnconfigure(1, weight=1)

            def apply(self):
                self.data["name"] = self.name_var.get()
                sels = self.lb.curselection()
                self.data["odds"] = [self.app.odd_libraries[i].get("name", "") for i in sels]

        class ScenarioDialog(simpledialog.Dialog):
            def __init__(self, parent, app, lib, data=None):
                self.app = app
                self.lib = lib
                self.data = data or {
                    "name": "",
                    "behavior": "",
                    "scenery": "",
                    "tc": "",
                    "fi": "",
                    "description": "",
                }
                self.tag_counter = 0
                super().__init__(parent, title="Edit Scenario")

            def body(self, master):
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")

                ttk.Label(master, text="Other Road Users").grid(row=1, column=0, sticky="e")
                self.beh_var = tk.StringVar(value=self.data.get("behavior", ""))
                ttk.Entry(master, textvariable=self.beh_var).grid(row=1, column=1, sticky="ew")

                ttk.Label(master, text="Scenery").grid(row=2, column=0, sticky="e")
                self.sce_var = tk.StringVar(value=self.data.get("scenery", ""))
                ttk.Entry(master, textvariable=self.sce_var).grid(row=2, column=1, sticky="ew")

                elems = []
                for name in self.lib.get("odds", []):
                    for l in self.app.odd_libraries:
                        if l.get("name") == name:
                            for el in l.get("elements", []):
                                if isinstance(el, dict):
                                    val = el.get("name") or el.get("element") or el.get("id")
                                else:
                                    val = str(el)
                                if val:
                                    elems.append(val)

                ttk.Label(master, text="ODD Element").grid(row=3, column=0, sticky="e")
                self.elem_var = tk.StringVar()
                self.elem_combo = ttk.Combobox(master, textvariable=self.elem_var, values=elems, state="readonly")
                self.elem_combo.grid(row=3, column=1, sticky="ew")
                ttk.Button(master, text="To Scenery", command=self.insert_elem).grid(row=3, column=2, padx=2)
                ttk.Button(master, text="To Desc", command=self.insert_desc_elem).grid(row=3, column=3, padx=2)

                tc_names = [n.user_name or f"TC {n.unique_id}" for n in self.app.get_all_triggering_conditions()]
                fi_names = [n.user_name or f"FI {n.unique_id}" for n in self.app.get_all_functional_insufficiencies()]
                ttk.Label(master, text="Triggering Condition").grid(row=4, column=0, sticky="e")
                self.tc_var = tk.StringVar(value=self.data.get("tc", ""))
                ttk.Combobox(master, textvariable=self.tc_var, values=tc_names, state="readonly").grid(row=4, column=1, sticky="ew")
                ttk.Label(master, text="Functional Insufficiency").grid(row=5, column=0, sticky="e")
                self.fi_var = tk.StringVar(value=self.data.get("fi", ""))
                ttk.Combobox(master, textvariable=self.fi_var, values=fi_names, state="readonly").grid(row=5, column=1, sticky="ew")

                ttk.Label(master, text="Description").grid(row=6, column=0, sticky="ne")
                self.desc = tk.Text(master, height=4, width=40, wrap="word")
                self.desc.grid(row=6, column=1, columnspan=3, sticky="nsew")
                self.load_desc_links()
                master.grid_columnconfigure(1, weight=1)

            def insert_elem(self):
                el = self.elem_var.get()
                if el:
                    cur = self.sce_var.get().strip()
                    if cur:
                        self.sce_var.set(f"{cur}, {el}")
                    else:
                        self.sce_var.set(el)

            def insert_desc_elem(self):
                el = self.elem_var.get()
                if not el:
                    return
                pos = self.desc.index(tk.INSERT)
                text = f"[[{el}]]"
                self.desc.insert(pos, text)
                tag = f"link{self.tag_counter}"
                self.tag_counter += 1
                end = self.desc.index(f"{pos}+{len(text)}c")
                self.desc.tag_add(tag, pos, end)
                self.desc.tag_config(tag, foreground="blue", underline=1)
                self.desc.tag_bind(tag, "<Button-1>", lambda e, n=el: self.show_elem(n))

            def load_desc_links(self):
                desc = self.data.get("description", "")
                self.desc.insert("1.0", desc)
                for m in re.finditer(r"\[\[(.+?)\]\]", desc):
                    name = m.group(1)
                    start = f"1.0+{m.start()}c"
                    end = f"1.0+{m.end()}c"
                    tag = f"link{self.tag_counter}"
                    self.tag_counter += 1
                    self.desc.tag_add(tag, start, end)
                    self.desc.tag_config(tag, foreground="blue", underline=1)
                    self.desc.tag_bind(tag, "<Button-1>", lambda e, n=name: self.show_elem(n))

            def show_elem(self, name):
                for lib_name in self.lib.get("odds", []):
                    for l in self.app.odd_libraries:
                        if l.get("name") == lib_name:
                            for el in l.get("elements", []):
                                val = el.get("name") or el.get("element") or el.get("id")
                                if val == name:
                                    msg = "\n".join(f"{k}: {v}" for k, v in el.items())
                                    messagebox.showinfo("ODD Element", msg)
                                    return
                messagebox.showinfo("ODD Element", f"{name}")

            def apply(self):
                self.data["name"] = self.name_var.get()
                self.data["behavior"] = self.beh_var.get()
                self.data["scenery"] = self.sce_var.get()
                self.data["tc"] = self.tc_var.get()
                self.data["fi"] = self.fi_var.get()
                self.data["description"] = self.desc.get("1.0", "end-1c")

        def add_lib():
            dlg = LibraryDialog(win, self)
            if dlg.data.get("name"):
                self.scenario_libraries.append({"name": dlg.data["name"], "scenarios": [], "odds": dlg.data["odds"]})
                refresh_libs()

        def edit_lib():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.scenario_libraries[sel[0]]
            dlg = LibraryDialog(win, self, lib)
            lib.update(dlg.data)
            refresh_libs()

        def delete_lib():
            sel = lib_lb.curselection()
            if sel:
                idx = sel[0]
                del self.scenario_libraries[idx]
                refresh_libs()

        def add_scen():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.scenario_libraries[sel[0]]
            dlg = ScenarioDialog(win, self, lib)
            if dlg.data.get("name"):
                lib.setdefault("scenarios", []).append(dlg.data)
                refresh_scenarios()

        def edit_scen():
            sel_lib = lib_lb.curselection()
            sel_sc = scen_tree.selection()
            if not sel_lib or not sel_sc:
                return
            lib = self.scenario_libraries[sel_lib[0]]
            idx = scen_tree.index(sel_sc[0])
            data = lib.get("scenarios", [])[idx]
            dlg = ScenarioDialog(win, self, lib, data)
            lib["scenarios"][idx] = dlg.data
            refresh_scenarios()

        def del_scen():
            sel_lib = lib_lb.curselection()
            sel_sc = scen_tree.selection()
            if not sel_lib or not sel_sc:
                return
            lib = self.scenario_libraries[sel_lib[0]]
            idx = scen_tree.index(sel_sc[0])
            del lib.get("scenarios", [])[idx]
            refresh_scenarios()

        btnf = ttk.Frame(win)
        btnf.grid(row=1, column=1, columnspan=3, sticky="ew")
        ttk.Button(btnf, text="Add Lib", command=add_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Edit Lib", command=edit_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Lib", command=delete_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Add Scen", command=add_scen).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="Edit Scen", command=edit_scen).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Scen", command=del_scen).pack(side=tk.LEFT)

        lib_lb.bind("<<ListboxSelect>>", refresh_scenarios)
        refresh_libs()

    def manage_odd_libraries(self):
        win = tk.Toplevel(self.root)
        win.title("ODD Libraries")
        lib_lb = tk.Listbox(win, height=8, width=25)
        lib_lb.grid(row=0, column=0, rowspan=4, sticky="nsew")
        elem_tree = ttk.Treeview(win, columns=("attrs",), show="tree headings")
        elem_tree.heading("#0", text="Name")
        elem_tree.column("#0", width=150)
        elem_tree.heading("attrs", text="Attributes")
        elem_tree.column("attrs", width=200)
        elem_tree.grid(row=0, column=1, columnspan=3, sticky="nsew")
        win.grid_rowconfigure(0, weight=1)
        win.grid_columnconfigure(1, weight=1)

        def refresh_libs():
            lib_lb.delete(0, tk.END)
            for lib in self.odd_libraries:
                lib_lb.insert(tk.END, lib.get("name", ""))
            refresh_elems()

        def refresh_elems(*_):
            elem_tree.delete(*elem_tree.get_children())
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.odd_libraries[sel[0]]
            for el in lib.get("elements", []):
                name = el.get("name") or el.get("element") or el.get("id")
                attrs = ", ".join(f"{k}={v}" for k, v in el.items() if k != "name")
                elem_tree.insert("", tk.END, values=(attrs,), text=name)

        class ElementDialog(simpledialog.Dialog):
            def __init__(self, parent, data=None):
                self.data = data or {"name": ""}
                super().__init__(parent, title="Edit Element")

            def add_attr_row(self, key="", val=""):
                r = len(self.attr_rows)
                k_var = tk.StringVar(value=key)
                v_var = tk.StringVar(value=str(val))
                ttk.Entry(self.attr_frame, textvariable=k_var).grid(row=r, column=0, padx=2, pady=2)
                ttk.Entry(self.attr_frame, textvariable=v_var).grid(row=r, column=1, padx=2, pady=2)
                self.attr_rows.append((k_var, v_var))

            def body(self, master):
                ttk.Label(master, text="Name").grid(row=0, column=0, sticky="e")
                self.name_var = tk.StringVar(value=self.data.get("name", ""))
                ttk.Entry(master, textvariable=self.name_var).grid(row=0, column=1, sticky="ew")
                self.attr_frame = ttk.Frame(master)
                self.attr_frame.grid(row=1, column=0, columnspan=2, sticky="nsew")
                self.attr_rows = []
                for k, v in self.data.items():
                    if k != "name":
                        self.add_attr_row(k, v)
                ttk.Button(master, text="Add Attribute", command=self.add_attr_row).grid(row=2, column=0, columnspan=2, pady=5)

            def apply(self):
                new_data = {"name": self.name_var.get()}
                for k_var, v_var in self.attr_rows:
                    key = k_var.get().strip()
                    if key:
                        new_data[key] = v_var.get()
                self.data = new_data

        def add_lib():
            name = simpledialog.askstring("New Library", "Library name:")
            if not name:
                return
            elems = []
            if messagebox.askyesno("Import", "Import elements from file?"):
                path = filedialog.askopenfilename(filetypes=[("CSV/Excel", "*.csv *.xlsx")])
                if path:
                    if path.lower().endswith(".csv"):
                        with open(path, newline="") as f:
                            elems = list(csv.DictReader(f))
                    elif path.lower().endswith(".xlsx"):
                        try:
                            if load_workbook is None:
                                raise ImportError
                            wb = load_workbook(path, read_only=True)
                            ws = wb.active
                            headers = [c.value for c in next(ws.iter_rows(max_row=1))]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                elem = {headers[i]: row[i] for i in range(len(headers))}
                                elems.append(elem)
                        except Exception:
                            messagebox.showerror("Import", "Failed to read Excel file. openpyxl required.")
            self.odd_libraries.append({"name": name, "elements": elems})
            refresh_libs()
            self.update_odd_elements()

        def edit_lib():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.odd_libraries[sel[0]]
            name = simpledialog.askstring("Edit Library", "Library name:", initialvalue=lib.get("name", ""))
            if name:
                lib["name"] = name
                refresh_libs()

        def delete_lib():
            sel = lib_lb.curselection()
            if sel:
                idx = sel[0]
                del self.odd_libraries[idx]
                refresh_libs()
                self.update_odd_elements()

        def add_elem():
            sel = lib_lb.curselection()
            if not sel:
                return
            lib = self.odd_libraries[sel[0]]
            dlg = ElementDialog(win)
            lib.setdefault("elements", []).append(dlg.data)
            refresh_elems()
            self.update_odd_elements()

        def edit_elem():
            sel_lib = lib_lb.curselection()
            sel_elem = elem_tree.selection()
            if not sel_lib or not sel_elem:
                return
            lib = self.odd_libraries[sel_lib[0]]
            idx = elem_tree.index(sel_elem[0])
            data = lib.get("elements", [])[idx]
            dlg = ElementDialog(win, data)
            lib["elements"][idx] = dlg.data
            refresh_elems()
            self.update_odd_elements()

        def del_elem():
            sel_lib = lib_lb.curselection()
            sel_elem = elem_tree.selection()
            if not sel_lib or not sel_elem:
                return
            lib = self.odd_libraries[sel_lib[0]]
            idx = elem_tree.index(sel_elem[0])
            del lib.get("elements", [])[idx]
            refresh_elems()
            self.update_odd_elements()

        btnf = ttk.Frame(win)
        btnf.grid(row=1, column=1, columnspan=3, sticky="ew")
        ttk.Button(btnf, text="Add Lib", command=add_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Edit Lib", command=edit_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Lib", command=delete_lib).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Add Elem", command=add_elem).pack(side=tk.LEFT, padx=5)
        ttk.Button(btnf, text="Edit Elem", command=edit_elem).pack(side=tk.LEFT)
        ttk.Button(btnf, text="Del Elem", command=del_elem).pack(side=tk.LEFT)

        lib_lb.bind("<<ListboxSelect>>", refresh_elems)
        refresh_libs()

    def open_reliability_window(self):
        if hasattr(self, "_rel_window") and self._rel_window.winfo_exists():
            self._rel_window.lift()
            return
        self._rel_window = ReliabilityWindow(self)

    def open_fmeda_window(self):
        self.show_fmeda_list()

    def open_hazop_window(self):
        if hasattr(self, "_hazop_window") and self._hazop_window.winfo_exists():
            self._hazop_window.lift()
            return
        self._hazop_window = HazopWindow(self)

    def open_hara_window(self):
        if hasattr(self, "_hara_window") and self._hara_window.winfo_exists():
            self._hara_window.lift()
            return
        self._hara_window = HaraWindow(self)

    def open_fi2tc_window(self):
        if hasattr(self, "_fi2tc_window") and self._fi2tc_window.winfo_exists():
            self._fi2tc_window.lift()
            return
        self._fi2tc_window = FI2TCWindow(self)

    def open_tc2fi_window(self):
        if hasattr(self, "_tc2fi_window") and self._tc2fi_window.winfo_exists():
            self._tc2fi_window.lift()
            return
        self._tc2fi_window = TC2FIWindow(self)
        
    def copy_node(self):
        if self.selected_node and self.selected_node != self.root_node:
            self.clipboard_node = self.selected_node
            self.cut_mode = False
        else:
            messagebox.showwarning("Copy", "Select a non-root node to copy.")

    def cut_node(self):
        """Store the currently selected node for a cut & paste operation."""
        if self.selected_node and self.selected_node != self.root_node:
            self.clipboard_node = self.selected_node
            self.cut_mode = True
        else:
            messagebox.showwarning("Cut", "Select a non-root node to cut.")

    def paste_node(self):
        # 1) Determine target from selection or current selected node.
        target = None
        sel = self.treeview.selection()
        if sel:
            tags = self.treeview.item(sel[0], "tags")
            if tags:
                target = self.find_node_by_id(self.root_node, int(tags[0]))
        if not target:
            target = self.selected_node
        if not target:
            messagebox.showwarning("Paste", "Select a target node to paste into.")
            return

        # 2) Do not allow pasting into base events.
        if target.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            messagebox.showwarning("Paste", "Cannot paste into a base event.")
            return

        # 3) Always use the primary instance of target.
        if not target.is_primary_instance:
            target = target.original

        # 4) Ensure clipboard is not empty.
        if not self.clipboard_node:
            messagebox.showwarning("Paste", "Clipboard is empty.")
            return

        # 5) Prevent self-pasting.
        if target.unique_id == self.clipboard_node.unique_id:
            messagebox.showwarning("Paste", "Cannot paste a node onto itself.")
            return
        for child in target.children:
            if child.unique_id == self.clipboard_node.unique_id:
                messagebox.showwarning("Paste", "This node is already a child of the target.")
                return

        # 6) If in cut mode, update parent's pointer, remove from top_events, and update coordinates.
        if self.cut_mode:
            if self.clipboard_node in self.top_events:
                self.top_events.remove(self.clipboard_node)
            for p in list(self.clipboard_node.parents):
                if self.clipboard_node in p.children:
                    p.children.remove(self.clipboard_node)
            self.clipboard_node.parents = []
            if self.clipboard_node.node_type.upper() == "TOP EVENT":
                # Demote top events so they no longer show in the tree.
                self.clipboard_node.node_type = "RIGOR LEVEL"
                self.clipboard_node.severity = None
                self.clipboard_node.is_page = False
                self.clipboard_node.input_subtype = "Capability"
            self.clipboard_node.is_primary_instance = True
            target.children.append(self.clipboard_node)
            self.clipboard_node.parents.append(target)
            # NEW: Update its position so it is offset relative to the new parent.
            self.clipboard_node.x = target.x + 100
            self.clipboard_node.y = target.y + 100
            # (Optional: remove any clone marker from its label.)
            self.clipboard_node.display_label = self.clipboard_node.display_label.replace(" (clone)", "")
            self.clipboard_node = None
            self.cut_mode = False
            messagebox.showinfo("Paste", "Node moved successfully (cut & pasted).")
        else:
            # 7) Copy branch: create a clone and attach it.
            cloned_node = self.clone_node_preserving_id(self.clipboard_node)
            target.children.append(cloned_node)
            cloned_node.parents.append(target)
            # NEW: Also update the cloned node’s position relative to the target.
            cloned_node.x = target.x + 100
            cloned_node.y = target.y + 100
            messagebox.showinfo("Paste", "Node pasted successfully (copied).")

        # 8) Recalculate and update views.
        AD_RiskAssessment_Helper.calculate_assurance_recursive(
            self.root_node,
            self.top_events,
        )
        self.update_views()
 
    def clone_node_preserving_id(self, node):
        # Create a new node with the same properties, but assign a new unique ID.
        new_node = FaultTreeNode(node.user_name, node.node_type)
        new_node.unique_id = AD_RiskAssessment_Helper.get_next_unique_id()
        new_node.quant_value = node.quant_value
        new_node.gate_type = node.gate_type
        new_node.description = node.description
        new_node.rationale = node.rationale
        # NEW: Offset the new node relative to the original.
        new_node.x = node.x + 100  
        new_node.y = node.y + 100
        new_node.severity = node.severity
        new_node.input_subtype = node.input_subtype
        new_node.display_label = node.display_label  # (do not append " (clone)" in the copy branch if you prefer)
        new_node.equation = node.equation
        new_node.detailed_equation = node.detailed_equation
        new_node.is_page = node.is_page
        new_node.is_primary_instance = False
        # Set the clone’s "original" pointer: if the original was primary, use it; otherwise, use its original.
        new_node.original = node if node.is_primary_instance else node.original
        new_node.children = []
        return new_node

    def sync_nodes_by_id(self, updated_node):
        # Always work with the primary instance.
        if not updated_node.is_primary_instance and updated_node.original:
            updated_node = updated_node.original
        updated_primary_id = updated_node.unique_id

        for node in self.get_all_nodes(self.root_node):
            # Skip the updated node itself.
            if node is updated_node:
                continue

            if node.is_primary_instance:
                if node.unique_id == updated_primary_id:
                    node.node_type = updated_node.node_type
                    node.user_name = updated_node.user_name
                    node.description = updated_node.description
                    node.rationale = updated_node.rationale
                    node.quant_value = updated_node.quant_value
                    node.gate_type = updated_node.gate_type
                    node.severity = updated_node.severity
                    node.input_subtype = updated_node.input_subtype
                    node.display_label = updated_node.display_label
                    node.equation = updated_node.equation
                    node.detailed_equation = updated_node.detailed_equation
                    node.is_page = updated_node.is_page
                    node.failure_prob = updated_node.failure_prob
                    node.prob_formula = updated_node.prob_formula
                    node.failure_mode_ref = updated_node.failure_mode_ref
                    node.fmea_effect = updated_node.fmea_effect
                    node.fmea_cause = updated_node.fmea_cause
                    node.fmea_severity = updated_node.fmea_severity
                    node.fmea_occurrence = updated_node.fmea_occurrence
                    node.fmea_detection = updated_node.fmea_detection
                    node.fmea_component = updated_node.fmea_component
                    node.fmeda_malfunction = updated_node.fmeda_malfunction
                    node.fmeda_safety_goal = updated_node.fmeda_safety_goal
                    node.fmeda_diag_cov = updated_node.fmeda_diag_cov
                    node.fmeda_fit = updated_node.fmeda_fit
                    node.fmeda_spfm = updated_node.fmeda_spfm
                    node.fmeda_lpfm = updated_node.fmeda_lpfm
                    node.fmeda_fault_type = updated_node.fmeda_fault_type
                    node.fmeda_fault_fraction = updated_node.fmeda_fault_fraction
            else:
                # Use the original pointer to compare.
                if node.original and node.original.unique_id == updated_primary_id:
                    node.user_name = updated_node.user_name
                    node.description = updated_node.description
                    node.rationale = updated_node.rationale
                    node.quant_value = updated_node.quant_value
                    node.gate_type = updated_node.gate_type
                    node.severity = updated_node.severity
                    node.input_subtype = updated_node.input_subtype
                    # Append a marker to the display label to indicate this is a clone.
                    node.display_label = updated_node.display_label + " (clone)"
                    node.equation = updated_node.equation
                    node.detailed_equation = updated_node.detailed_equation
                    # **The key change: update the page flag on clones as well.**
                    node.is_page = updated_node.is_page
                    node.failure_prob = updated_node.failure_prob
                    node.prob_formula = updated_node.prob_formula
                    node.failure_mode_ref = updated_node.failure_mode_ref
                    node.fmea_effect = updated_node.fmea_effect
                    node.fmea_cause = updated_node.fmea_cause
                    node.fmea_severity = updated_node.fmea_severity
                    node.fmea_occurrence = updated_node.fmea_occurrence
                    node.fmea_detection = updated_node.fmea_detection
                    node.fmea_component = updated_node.fmea_component
                    node.fmeda_malfunction = updated_node.fmeda_malfunction
                    node.fmeda_safety_goal = updated_node.fmeda_safety_goal
                    node.fmeda_diag_cov = updated_node.fmeda_diag_cov
                    node.fmeda_fit = updated_node.fmeda_fit
                    node.fmeda_spfm = updated_node.fmeda_spfm
                    node.fmeda_lpfm = updated_node.fmeda_lpfm
                    node.fmeda_fault_type = updated_node.fmeda_fault_type
                    node.fmeda_fault_fraction = updated_node.fmeda_fault_fraction

    def edit_user_name(self):
        if self.selected_node:
            new_name = simpledialog.askstring("Edit User Name", "Enter new user name:", initialvalue=self.selected_node.user_name)
            if new_name is not None:
                self.selected_node.user_name = new_name.strip()
                self.update_views()
        else:
            messagebox.showwarning("Edit User Name", "Select a node first.")

    def edit_description(self):
        if self.selected_node:
            new_desc = simpledialog.askstring("Edit Description", "Enter new description:", initialvalue=self.selected_node.description)
            if new_desc is not None:
                self.selected_node.description = new_desc
                self.update_views()
        else:
            messagebox.showwarning("Edit Description", "Select a node first.")

    def edit_rationale(self):
        if self.selected_node:
            new_rat = simpledialog.askstring("Edit Rationale", "Enter new rationale:", initialvalue=self.selected_node.rationale)
            if new_rat is not None:
                self.selected_node.rationale = new_rat
                self.update_views()
        else:
            messagebox.showwarning("Edit Rationale", "Select a node first.")

    def edit_value(self):
        if self.selected_node and self.selected_node.node_type.upper() in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
            try:
                new_val = simpledialog.askfloat("Edit Value", "Enter new value (1-5):", initialvalue=self.selected_node.quant_value)
                if new_val is not None and 1 <= new_val <= 5:
                    self.selected_node.quant_value = new_val
                    self.update_views()
                else:
                    messagebox.showerror("Error", "Value must be between 1 and 5.")
            except Exception:
                messagebox.showerror("Error", "Invalid input.")
        else:
            messagebox.showwarning("Edit Value", "Select a Confidence or Robustness node.")

    def edit_gate_type(self):
        if self.selected_node and self.selected_node.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
            new_gt = simpledialog.askstring("Edit Gate Type", "Enter new gate type (AND/OR):", initialvalue=self.selected_node.gate_type)
            if new_gt is not None and new_gt.upper() in ["AND", "OR"]:
                self.selected_node.gate_type = new_gt.upper()
                self.update_views()
            else:
                messagebox.showerror("Error", "Gate type must be AND or OR.")
        else:
            messagebox.showwarning("Edit Gate Type", "Select a gate-type node.")

    def edit_severity(self):
        messagebox.showinfo(
            "Severity",
            "Severity is determined from the HARA and cannot be edited here.",
        )

    def edit_controllability(self):
        messagebox.showinfo(
            "Controllability",
            "Controllability is determined from the HARA and cannot be edited here.",
        )

    def edit_page_flag(self):
        if not self.selected_node:
            messagebox.showwarning("Edit Page Flag", "Select a node first.")
            return
        # If this is a clone, update its original.
        target = self.selected_node if self.selected_node.is_primary_instance else self.selected_node.original

        if target.node_type.upper() in ["TOP EVENT", "BASIC EVENT"]:
            messagebox.showwarning("Edit Page Flag", "This node type cannot be a page.")
            return

        # Ask for the new page flag value.
        response = messagebox.askyesno("Edit Page Flag", f"Should node '{target.name}' be a page gate?")
        target.is_page = response

        # Sync the changes to all clones.
        self.sync_nodes_by_id(target)
        self.update_views()

    def set_last_saved_state(self):
        """Record the current model state for change detection."""
        self.last_saved_state = json.dumps(self.export_model_data(), sort_keys=True)

    def has_unsaved_changes(self):
        """Return True if the model differs from the last saved state."""
        current_state = json.dumps(self.export_model_data(), sort_keys=True)
        return current_state != getattr(self, "last_saved_state", None)

    def confirm_close(self):
        """Prompt to save if there are unsaved changes before closing."""
        if self.has_unsaved_changes():
            result = messagebox.askyesnocancel("Unsaved Changes", "Save changes before exiting?")
            if result is None:
                return
            if result:
                self.save_model()
        self.root.destroy()

    def export_model_data(self, include_versions=True):
        # Ensure aggregated ODD elements are up to date
        self.update_odd_elements()
        reviews = []
        for r in self.reviews:
            reviews.append({
                "name": r.name,
                "description": r.description,
                "mode": r.mode,
                "moderators": [asdict(m) for m in r.moderators],
                "approved": r.approved,
                "due_date": r.due_date,
                "closed": r.closed,
                "participants": [asdict(p) for p in r.participants],
                "comments": [asdict(c) for c in r.comments],
                "fta_ids": r.fta_ids,
                "fmea_names": r.fmea_names,
                "fmeda_names": getattr(r, 'fmeda_names', []),
                "hazop_names": getattr(r, 'hazop_names', []),
                "hara_names": getattr(r, 'hara_names', []),
            })
        current_name = self.review_data.name if self.review_data else None
        data = {
            "top_events": [event.to_dict() for event in self.top_events],
            "fmeas": [
                {
                    "name": f["name"],
                    "file": f["file"],
                    "entries": [e.to_dict() for e in f["entries"]],
                }
                for f in self.fmeas
            ],
            "fmedas": [
                {
                    "name": d["name"],
                    "file": d["file"],
                    "entries": [e.to_dict() for e in d["entries"]],
                    "bom": d.get("bom", ""),
                }
                for d in self.fmedas
            ],
            "mechanism_libraries": [
                {
                    "name": lib.name,
                    "mechanisms": [asdict(m) for m in lib.mechanisms],
                }
                for lib in self.mechanism_libraries
            ],
            "selected_mechanism_libraries": [lib.name for lib in self.selected_mechanism_libraries],
            "mission_profiles": [
                {
                    **asdict(mp),
                    "duty_cycle": mp.tau_on / (mp.tau_on + mp.tau_off)
                    if (mp.tau_on + mp.tau_off)
                    else 0.0,
                }
                for mp in self.mission_profiles
            ],
            "reliability_analyses": [
                {
                    "name": ra.name,
                    "standard": ra.standard,
                    "profile": ra.profile,
                    "components": [asdict(c) for c in ra.components],
                    "total_fit": ra.total_fit,
                    "spfm": ra.spfm,
                    "lpfm": ra.lpfm,
                    "dc": ra.dc,
                }
                for ra in self.reliability_analyses
            ],
            "hazops": [
                {
                    "name": doc.name,
                    "entries": [asdict(e) for e in doc.entries],
                }
                for doc in self.hazop_docs
            ],
            "haras": [
                {
                    "name": doc.name,
                    "hazops": getattr(doc, "hazops", []),
                    "entries": [asdict(e) for e in doc.entries],
                    "approved": getattr(doc, "approved", False),
                }
                for doc in self.hara_docs
            ],
            "hazop_entries": [asdict(e) for e in self.hazop_entries],
            "fi2tc_entries": self.fi2tc_entries,
            "tc2fi_entries": self.tc2fi_entries,
            "scenario_libraries": self.scenario_libraries,
            "odd_libraries": self.odd_libraries,
            "project_properties": self.project_properties,
            "global_requirements": global_requirements,
            "reviews": reviews,
            "current_review": current_name,
        }
        if include_versions:
            data["versions"] = self.versions
        return data

    def save_model(self):
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
        if path:
            for fmea in self.fmeas:
                self.export_fmea_to_csv(fmea, fmea['file'])
            for fmeda in self.fmedas:
                self.export_fmeda_to_csv(fmeda, fmeda['file'])
            data = self.export_model_data()
            with open(path, "w") as f:
                json.dump(data, f, indent=4)
            messagebox.showinfo("Saved", "Model saved with all configuration and safety goal information.")
            self.set_last_saved_state()

    def load_model(self):
        global AD_RiskAssessment_Helper
        # Reinitialize the helper so that the counter is reset.
        AD_RiskAssessment_Helper = ADRiskAssessmentHelper()
        
        path = filedialog.askopenfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
        if not path:
            return
        with open(path, "r") as f:
            raw = f.read()
        try:
            data = json.loads(raw)
        except json.JSONDecodeError as exc:
            import re

            def clean(text: str) -> str:
                text = re.sub(r"//.*", "", text)
                text = re.sub(r"#.*", "", text)
                text = re.sub(r"/\*.*?\*/", "", text, flags=re.S)
                text = re.sub(r",\s*(\]|\})", r"\1", text)
                return text

            try:
                data = json.loads(clean(raw))
            except json.JSONDecodeError:
                messagebox.showerror(
                    "Load Model",
                    f"Failed to parse JSON file:\n{exc}",
                )
                return

        if "top_events" in data:
            self.top_events = [FaultTreeNode.from_dict(e) for e in data["top_events"]]
        elif "root_node" in data:
            root = FaultTreeNode.from_dict(data["root_node"])
            self.top_events = [root]
        else:
            messagebox.showerror("Error", "Invalid model file format.")
            return

        self.fmeas = []
        for fmea_data in data.get("fmeas", []):
            entries = [FaultTreeNode.from_dict(e) for e in fmea_data.get("entries", [])]
            self.fmeas.append({"name": fmea_data.get("name", "FMEA"), "file": fmea_data.get("file", f"fmea_{len(self.fmeas)}.csv"), "entries": entries})
        if not self.fmeas and "fmea_entries" in data:
            entries = [FaultTreeNode.from_dict(e) for e in data.get("fmea_entries", [])]
            self.fmeas.append({"name": "Default FMEA", "file": "fmea_default.csv", "entries": entries})

        self.fmedas = []
        for doc in data.get("fmedas", []):
            entries = [FaultTreeNode.from_dict(e) for e in doc.get("entries", [])]
            self.fmedas.append({
                "name": doc.get("name", "FMEDA"),
                "file": doc.get("file", f"fmeda_{len(self.fmedas)}.csv"),
                "entries": entries,
                "bom": doc.get("bom", ""),
            })

        # Mechanism libraries and selections
        self.mechanism_libraries = []
        for lib in data.get("mechanism_libraries", []):
            mechs = [DiagnosticMechanism(**m) for m in lib.get("mechanisms", [])]
            self.mechanism_libraries.append(MechanismLibrary(lib.get("name", ""), mechs))
        self.selected_mechanism_libraries = []
        for name in data.get("selected_mechanism_libraries", []):
            found = next((l for l in self.mechanism_libraries if l.name == name), None)
            if found:
                self.selected_mechanism_libraries.append(found)
        if not self.mechanism_libraries:
            self.load_default_mechanisms()

        # Mission profiles
        self.mission_profiles = []
        for mp_data in data.get("mission_profiles", []):
            try:
                mp = MissionProfile(**mp_data)
                total = mp.tau_on + mp.tau_off
                mp.duty_cycle = mp.tau_on / total if total else 0.0
                self.mission_profiles.append(mp)
            except TypeError:
                pass

        # Reliability analyses
        self.reliability_analyses = []
        for ra in data.get("reliability_analyses", []):
            def load_comp(cdata):
                comp = ReliabilityComponent(
                    cdata.get("name", ""),
                    cdata.get("comp_type", ""),
                    cdata.get("quantity", 1),
                    cdata.get("attributes", {}),
                    cdata.get("qualification", cdata.get("safety_req", "")),
                    cdata.get("fit", 0.0),
                    cdata.get("is_passive", False),
                )
                comp.sub_boms = [
                    [load_comp(sc) for sc in bom]
                    for bom in cdata.get("sub_boms", [])
                ]
                return comp

            comps = [load_comp(c) for c in ra.get("components", [])]
            self.reliability_analyses.append(
                ReliabilityAnalysis(
                    ra.get("name", ""),
                    ra.get("standard", ""),
                    ra.get("profile", ""),
                    comps,
                    ra.get("total_fit", 0.0),
                    ra.get("spfm", 0.0),
                    ra.get("lpfm", 0.0),
                    ra.get("dc", 0.0),
                )
            )

        self.hazop_docs = []
        for d in data.get("hazops", []):
            entries = []
            for h in d.get("entries", []):
                h["safety"] = boolify(h.get("safety", False), False)
                h["covered"] = boolify(h.get("covered", False), False)
                entries.append(HazopEntry(**h))
            self.hazop_docs.append(
                HazopDoc(d.get("name", f"HAZOP {len(self.hazop_docs)+1}"), entries)
            )
        if not self.hazop_docs and "hazop_entries" in data:
            entries = []
            for h in data.get("hazop_entries", []):
                h["safety"] = boolify(h.get("safety", False), False)
                h["covered"] = boolify(h.get("covered", False), False)
                entries.append(HazopEntry(**h))
            self.hazop_docs.append(HazopDoc("Default", entries))
        self.active_hazop = self.hazop_docs[0] if self.hazop_docs else None
        self.hazop_entries = self.active_hazop.entries if self.active_hazop else []

        self.hara_docs = []
        for d in data.get("haras", []):
            entries = [HaraEntry(**e) for e in d.get("entries", [])]
            hazops = d.get("hazops")
            if not hazops:
                hazop = d.get("hazop")
                hazops = [hazop] if hazop else []
            self.hara_docs.append(
                HaraDoc(
                    d.get("name", f"HARA {len(self.hara_docs)+1}"),
                    hazops,
                    entries,
                    d.get("approved", False),
                )
            )
        if not self.hara_docs and "hara_entries" in data:
            hazop_name = self.hazop_docs[0].name if self.hazop_docs else ""
            self.hara_docs.append(
                HaraDoc("Default", [hazop_name] if hazop_name else [], [HaraEntry(**e) for e in data.get("hara_entries", [])])
            )
        self.active_hara = self.hara_docs[0] if self.hara_docs else None
        self.hara_entries = self.active_hara.entries if self.active_hara else []

        self.fi2tc_entries = data.get("fi2tc_entries", [])
        self.tc2fi_entries = data.get("tc2fi_entries", [])
        self.scenario_libraries = data.get("scenario_libraries", [])
        self.odd_libraries = data.get("odd_libraries", [])
        if not self.odd_libraries and "odd_elements" in data:
            self.odd_libraries = [{"name": "Default", "elements": data.get("odd_elements", [])}]
        self.update_odd_elements()

        self.fmedas = []
        for doc in data.get("fmedas", []):
            entries = [FaultTreeNode.from_dict(e) for e in doc.get("entries", [])]
            self.fmedas.append({
                "name": doc.get("name", "FMEDA"),
                "file": doc.get("file", f"fmeda_{len(self.fmedas)}.csv"),
                "entries": entries,
                "bom": doc.get("bom", ""),
            })

        # Fix clone references for each top event.
        for event in self.top_events:
            AD_RiskAssessment_Helper.fix_clone_references(self.top_events)

        # Update the unique ID counter.
        AD_RiskAssessment_Helper.update_unique_id_counter_for_top_events(self.top_events)
        
        # *** Add this loop to update your global_requirements database ***
        for event in self.top_events:
            self.update_global_requirements_from_nodes(event)

        # Propagate ASIL values from HARA entries to loaded safety goals
        if hasattr(self, "hara_entries"):
            self.sync_hara_to_safety_goals()
        
        # Load project properties.
        self.project_properties = data.get("project_properties", self.project_properties)
        self.reviews = []
        reviews_data = data.get("reviews")
        if reviews_data:
            for rd in reviews_data:
                participants = [ReviewParticipant(**p) for p in rd.get("participants", [])]
                comments = [ReviewComment(**c) for c in rd.get("comments", [])]
                moderators = [ReviewParticipant(**m) for m in rd.get("moderators", [])]
                if not moderators and rd.get("moderator"):
                    moderators = [ReviewParticipant(rd.get("moderator"), "", "moderator")]
                self.reviews.append(
                    ReviewData(
                        name=rd.get("name", ""),
                        description=rd.get("description", ""),
                        mode=rd.get("mode", "peer"),
                        moderators=moderators,
                        participants=participants,
                        comments=comments,
                        approved=rd.get("approved", False),
                        due_date=rd.get("due_date", ""),
                        closed=rd.get("closed", False),
                        fta_ids=rd.get("fta_ids", []),
                        fmea_names=rd.get("fmea_names", []),
                        fmeda_names=rd.get("fmeda_names", []),
                        hazop_names=rd.get("hazop_names", []),
                        hara_names=rd.get("hara_names", []),
                    )
                )
            current = data.get("current_review")
            self.review_data = None
            for r in self.reviews:
                if r.name == current:
                    self.review_data = r
                    break
        else:
            rd = data.get("review_data")
            if rd:
                participants = [ReviewParticipant(**p) for p in rd.get("participants", [])]
                comments = [ReviewComment(**c) for c in rd.get("comments", [])]
                moderators = [ReviewParticipant(**m) for m in rd.get("moderators", [])]
                if not moderators and rd.get("moderator"):
                    moderators = [ReviewParticipant(rd.get("moderator"), "", "moderator")]
                review = ReviewData(
                    name=rd.get("name", "Review 1"),
                    description=rd.get("description", ""),
                    mode=rd.get("mode", "peer"),
                    moderators=moderators,
                    participants=participants,
                    comments=comments,
                    approved=rd.get("approved", False),
                    due_date=rd.get("due_date", ""),
                    closed=rd.get("closed", False),
                    fta_ids=rd.get("fta_ids", []),
                    fmea_names=rd.get("fmea_names", []),
                    fmeda_names=rd.get("fmeda_names", []),
                    hazop_names=rd.get("hazop_names", []),
                    hara_names=rd.get("hara_names", []),
                )
                self.reviews = [review]
                self.review_data = review
            else:
                self.review_data = None
        self.versions = data.get("versions", [])

        self.selected_node = None
        if hasattr(self, "page_diagram") and self.page_diagram is not None:
            self.close_page_diagram()
        self.update_views()
        self.set_last_saved_state()
        
    def update_global_requirements_from_nodes(self,node):
        if hasattr(node, "safety_requirements"):
            for req in node.safety_requirements:
                # Use req["id"] as key; if already exists, you could update if needed.
                if req["id"] not in global_requirements:
                    global_requirements[req["id"]] = req
        for child in node.children:
            self.update_global_requirements_from_nodes(child)

    def generate_report(self):
        path = filedialog.asksaveasfilename(defaultextension=".html", filetypes=[("HTML", "*.html")])
        if path:
            html = self.build_html_report()
            with open(path, "w", encoding="utf-8") as f:
                f.write(html)
            messagebox.showinfo("Report", "HTML report generated.")

    def build_html_report(self):
        def node_to_html(n):
            txt = f"{n.name} ({n.node_type}"
            if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                txt += f", {n.gate_type}"
            txt += ")"
            if n.display_label:
                txt += f" => {n.display_label}"
            if n.description:
                txt += f"<br>Desc: {n.description}"
            if n.rationale:
                txt += f"<br>Rationale: {n.rationale}"
            content = f"<details open><summary>{txt}</summary>\n"
            for c in n.children:
                content += node_to_html(c)
            content += "</details>\n"
            return content
        return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Autonomous Driving Risk Assessment</title>
<style>body {{ font-family: Arial; }} details {{ margin-left: 20px; }}</style>
</head>
<body>
<h1>Autonomous Driving Risk Assessment</h1>
{node_to_html(self.root_node)}
</body>
</html>"""
    def resolve_original(self,node):
        # Walk the clone chain until you find a primary instance.
        while not node.is_primary_instance and node.original is not None and node.original != node:
            node = node.original
        return node

    def open_page_diagram(self, node, push_history=True):
        # Resolve the node to its original.
        resolved_node = self.resolve_original(node)
        if push_history and hasattr(self, "page_diagram") and self.page_diagram is not None:
            self.page_history.append(self.page_diagram.root_node)
        for widget in self.canvas_frame.winfo_children():
            widget.destroy()

        # Create header frame with the original node’s name.
        header_frame = ttk.Frame(self.canvas_frame)
        header_frame.grid(row=0, column=0, sticky="ew")
        header_frame.columnconfigure(0, weight=1)

        header = tk.Label(header_frame, text=f"Page Diagram: {resolved_node.name}",
                          font=("Arial", 14, "bold"))
        header.grid(row=0, column=0, sticky="w", padx=(5, 0))
        back_button = ttk.Button(header_frame, text="Go Back", command=self.go_back)
        back_button.grid(row=0, column=1, sticky="e", padx=5)

        page_canvas = tk.Canvas(self.canvas_frame, bg="white")
        page_canvas.grid(row=1, column=0, sticky="nsew")
        vbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=page_canvas.yview)
        vbar.grid(row=1, column=1, sticky="ns")
        self.canvas_frame.rowconfigure(0, weight=0)
        self.canvas_frame.rowconfigure(1, weight=1)
        self.canvas_frame.columnconfigure(0, weight=1)

        # Use the resolved (original) node for the page diagram.
        self.page_diagram = PageDiagram(self, resolved_node, page_canvas)
        self.page_diagram.redraw_canvas()

    def go_back(self):
        if self.page_history:
            # Pop one page off the history and open it without pushing the current page again.
            previous_page = self.page_history.pop()
            self.open_page_diagram(previous_page, push_history=False)
        #else:
            # If history is empty, remain on the current (root) page.
            #messagebox.showinfo("Back", "You are already at the root page.")

    def draw_page_subtree(self, page_root):
        self.page_canvas.delete("all")
        self.draw_page_grid()
        visited_ids = set()
        self.draw_page_connections_subtree(page_root, visited_ids)
        self.draw_page_nodes_subtree(page_root)
        bbox = self.page_canvas.bbox("all")
        if bbox:
            self.page_canvas.config(scrollregion=bbox)

    def draw_page_grid(self):
        spacing = 20
        width = self.page_canvas.winfo_width() or 800
        height = self.page_canvas.winfo_height() or 600
        for x in range(0, width, spacing):
            self.page_canvas.create_line(x, 0, x, height, fill="#ddd", tags="grid")
        for y in range(0, height, spacing):
            self.page_canvas.create_line(0, y, width, y, fill="#ddd", tags="grid")

    def draw_page_connections_subtree(self, node, visited_ids):
        if id(node) in visited_ids:
            return
        visited_ids.add(id(node))
        region_width = 100
        parent_bottom = (node.x, node.y + 40)
        N = len(node.children)
        for i, child in enumerate(node.children):
            parent_conn = (node.x - region_width/2 + (i+0.5)*(region_width/N), parent_bottom[1])
            child_top = (child.x, child.y - 45)
            draw_90_connection(self.page_canvas, parent_conn, child_top, outline_color="dimgray", line_width=1)
        for child in node.children:
            self.draw_page_connections_subtree(child, visited_ids)

    def draw_page_nodes_subtree(self, node):
        self.draw_node_on_page_canvas(node)
        for child in node.children:
            self.draw_page_nodes_subtree(child)

    def draw_node_on_page_canvas(self, canvas, node):
        # Use the clone's own display label and append a marker
        if not node.is_primary_instance:
            display_label = node.display_label + " (clone)"
        else:
            display_label = node.display_label
        
        fill_color = self.get_node_fill_color(node)
        eff_x, eff_y = node.x, node.y
        top_text = f"Type: {node.node_type}"
        if node.input_subtype:
            top_text += f" ({node.input_subtype})"
        if node.description:
            top_text += f"\nDesc: {node.description}"
        if node.rationale:
            top_text += f"\nRationale: {node.rationale}"
        bottom_text = node.name

        outline_color = "dimgray"
        line_width = 1
        if node.unique_id in getattr(self.app, "diff_nodes", []):
            outline_color = "blue"
            line_width = 2
        elif not node.is_primary_instance:
            outline_color = "red"
        
        # For page elements, assume they use a triangle shape.
        if node.is_page:
            fta_drawing_helper.draw_triangle_shape(
                canvas,
                eff_x,
                eff_y,
                scale=40,
                top_text=top_text,
                bottom_text=bottom_text,
                fill=fill_color,
                outline_color=outline_color,
                line_width=line_width,
                font_obj=self.diagram_font,
            )
        else:
            node_type_upper = node.node_type.upper()
            if node_type_upper in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if node.gate_type and node.gate_type.upper() == "OR":
                    fta_drawing_helper.draw_rotated_or_gate_shape(
                        self.page_canvas,
                        eff_x,
                        eff_y,
                        scale=40,
                        top_text=top_text,
                        bottom_text=bottom_text,
                        fill=fill_color,
                        outline_color=outline_color,
                        line_width=line_width,
                    )
                else:
                    fta_drawing_helper.draw_rotated_and_gate_shape(
                        self.page_canvas,
                        eff_x,
                        eff_y,
                        scale=40,
                        top_text=top_text,
                        bottom_text=bottom_text,
                        fill=fill_color,
                        outline_color=outline_color,
                        line_width=line_width,
                    )
            elif node_type_upper in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
                fta_drawing_helper.draw_circle_event_shape(
                    self.page_canvas,
                    eff_x,
                    eff_y,
                    45,
                    top_text=top_text,
                    bottom_text=bottom_text,
                    fill=fill_color,
                    outline_color=outline_color,
                    line_width=line_width,
                )
            else:
                fta_drawing_helper.draw_circle_event_shape(
                    self.page_canvas,
                    eff_x,
                    eff_y,
                    45,
                    top_text=top_text,
                    bottom_text=bottom_text,
                    fill=fill_color,
                    outline_color=outline_color,
                    line_width=line_width,
                )

        if self.review_data:
            unresolved = any(
                c.node_id == node.unique_id and not c.resolved
                for c in self.review_data.comments
            )
            if unresolved:
                canvas.create_oval(
                    eff_x + 35,
                    eff_y + 35,
                    eff_x + 45,
                    eff_y + 45,
                    fill="yellow",
                    outline="black",
                )

    def on_ctrl_mousewheel_page(self, event):
        if event.delta > 0:
            self.page_diagram.zoom_in()
        else:
            self.page_diagram.zoom_out()

    def close_page_diagram(self):
        if self.page_history:
            prev = self.page_history.pop()
            for widget in self.canvas_frame.winfo_children():
                widget.destroy()
            if prev is None:
                self.canvas = tk.Canvas(self.canvas_frame, bg="white")
                self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                self.hbar = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
                self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
                self.vbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
                self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
                self.canvas.config(xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set,
                                   scrollregion=(0, 0, 2000, 2000))
                self.canvas.bind("<ButtonPress-3>", self.on_right_mouse_press)
                self.canvas.bind("<B3-Motion>", self.on_right_mouse_drag)
                self.canvas.bind("<Button-1>", self.on_canvas_click)
                self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
                self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
                self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
                self.canvas.bind("<ButtonRelease-3>", self.show_context_menu)
                self.update_views()
                self.page_diagram = None
            else:
                self.open_page_diagram(prev)
        else:
            for widget in self.canvas_frame.winfo_children():
                widget.destroy()
            self.canvas = tk.Canvas(self.canvas_frame, bg="white")
            self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            self.hbar = ttk.Scrollbar(self.canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
            self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
            self.vbar = ttk.Scrollbar(self.canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
            self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.canvas.config(xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set,
                               scrollregion=(0, 0, 2000, 2000))
            self.canvas.bind("<ButtonPress-3>", self.on_right_mouse_press)
            self.canvas.bind("<B3-Motion>", self.on_right_mouse_drag)
            self.canvas.bind("<Button-1>", self.on_canvas_click)
            self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
            self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
            self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
            self.canvas.bind("<ButtonRelease-3>", self.show_context_menu)
            self.update_views()
            self.page_diagram = None

    # --- Review Toolbox Methods ---
    def start_peer_review(self):
        dialog = ParticipantDialog(self.root, joint=False)

        if dialog.result:
            moderators, parts = dialog.result
            name = simpledialog.askstring("Review Name", "Enter unique review name:")
            if not name:
                return
            description = simpledialog.askstring("Description", "Enter a short description:")
            if description is None:
                description = ""
            if not moderators:
                messagebox.showerror("Review", "Please specify a moderator")
                return
            if not parts:
                messagebox.showerror("Review", "At least one reviewer required")
                return
            due_date = simpledialog.askstring("Due Date", "Enter due date (YYYY-MM-DD):")
            if any(r.name == name for r in self.reviews):
                messagebox.showerror("Review", "Name already exists")
                return
            scope = ReviewScopeDialog(self.root, self)
            fta_ids, fmea_names, fmeda_names, hazop_names, hara_names = (
                scope.result if scope.result else ([], [], [], [], [])
            )
            review = ReviewData(name=name, description=description, mode='peer', moderators=moderators,
                               participants=parts, comments=[],
                               fta_ids=fta_ids, fmea_names=fmea_names, fmeda_names=fmeda_names,
                               hazop_names=hazop_names, hara_names=hara_names,
                               due_date=due_date)
            self.reviews.append(review)
            self.review_data = review
            self.current_user = moderators[0].name if moderators else parts[0].name
            ReviewDocumentDialog(self.root, self, review)
            self.send_review_email(review)
            self.open_review_toolbox()

    def start_joint_review(self):
        dialog = ParticipantDialog(self.root, joint=True)
        if dialog.result:
            moderators, participants = dialog.result
            name = simpledialog.askstring("Review Name", "Enter unique review name:")
            if not name:
                return
            description = simpledialog.askstring("Description", "Enter a short description:")
            if description is None:
                description = ""
            if not moderators:
                messagebox.showerror("Review", "Please specify a moderator")
                return
            if not any(p.role == 'reviewer' for p in participants):
                messagebox.showerror("Review", "At least one reviewer required")
                return
            if not any(p.role == 'approver' for p in participants):
                messagebox.showerror("Review", "At least one approver required")
                return
            due_date = simpledialog.askstring("Due Date", "Enter due date (YYYY-MM-DD):")
            if any(r.name == name for r in self.reviews):
                messagebox.showerror("Review", "Name already exists")
                return
            scope = ReviewScopeDialog(self.root, self)
            fta_ids, fmea_names, fmeda_names, hazop_names, hara_names = (
                scope.result if scope.result else ([], [], [], [], [])
            )

            # Ensure each selected element has an approved peer review
            for tid in fta_ids:
                if not any(r.mode == 'peer' and r.approved and tid in r.fta_ids for r in self.reviews):
                    messagebox.showerror("Review", "Peer review must be approved before starting joint review")
                    return
            for name_fta in fmea_names:
                if not any(r.mode == 'peer' and r.approved and name_fta in r.fmea_names for r in self.reviews):
                    messagebox.showerror("Review", "Peer review must be approved before starting joint review")
                    return
            for name_fd in fmeda_names:
                if not any(r.mode == 'peer' and r.approved and name_fd in r.fmeda_names for r in self.reviews):
                    messagebox.showerror("Review", "Peer review must be approved before starting joint review")
                    return
            review = ReviewData(name=name, description=description, mode='joint', moderators=moderators,
                               participants=participants, comments=[],
                               fta_ids=fta_ids, fmea_names=fmea_names, fmeda_names=fmeda_names,
                               hazop_names=hazop_names, hara_names=hara_names,
                               due_date=due_date)
            self.reviews.append(review)
            self.review_data = review
            self.current_user = moderators[0].name if moderators else participants[0].name
            ReviewDocumentDialog(self.root, self, review)
            self.send_review_email(review)
            self.open_review_toolbox()

    def open_review_toolbox(self):
        if not self.reviews:
            messagebox.showwarning("Review", "No reviews defined")
            return
        if not self.review_data and self.reviews:
            self.review_data = self.reviews[0]
        if self.review_window is None or not self.review_window.winfo_exists():
            self.review_window = ReviewToolbox(self.root, self)
        self.set_current_user()

    def send_review_email(self, review):
        """Send the review summary to all reviewers via configured SMTP."""
        recipients = [p.email for p in review.participants if p.role == 'reviewer' and p.email]
        if not recipients:
            return

        # Determine the current user's email if available
        current_email = next((p.email for p in review.participants
                              if p.name == self.current_user), "")

        if not getattr(self, "email_config", None):
            cfg = EmailConfigDialog(self.root, default_email=current_email).result
            self.email_config = cfg

        cfg = getattr(self, "email_config", None)
        if not cfg:
            return

        subject = f"Review: {review.name}"
        lines = [f"Review Name: {review.name}", f"Description: {review.description}", ""]
        if review.fta_ids:
            lines.append("FTAs:")
            for tid in review.fta_ids:
                te = next((t for t in self.top_events if t.unique_id == tid), None)
                if te:
                    lines.append(f" - {te.name}")
            lines.append("")
        if review.fmea_names:
            lines.append("FMEAs:")
            for name in review.fmea_names:
                lines.append(f" - {name}")
            lines.append("")
        if getattr(review, 'hazop_names', []):
            lines.append("HAZOPs:")
            for name in review.hazop_names:
                lines.append(f" - {name}")
            lines.append("")
        if getattr(review, 'hara_names', []):
            lines.append("HARAs:")
            for name in review.hara_names:
                lines.append(f" - {name}")
            lines.append("")
        content = "\n".join(lines)
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = cfg['email']
        msg['To'] = ', '.join(recipients)
        msg.set_content(content)

        html_lines = ["<html><body>", "<pre>", html.escape(content), "</pre>"]
        image_cids = []
        images = []
        for tid in review.fta_ids:
            node = self.find_node_by_id_all(tid)
            if not node:
                continue
            img = self.capture_diff_diagram(node)
            if img is None:
                continue
            buf = BytesIO()
            img.save(buf, format="PNG")
            buf.seek(0)
            cid = make_msgid()
            label = node.user_name or node.name or f"id{tid}"
            html_lines.append(f"<p><b>FTA: {html.escape(label)}</b><br>" +
                             f"<img src=\"cid:{cid[1:-1]}\" alt=\"{html.escape(label)}\"></p>")
            image_cids.append(cid)
            images.append(buf.getvalue())
        diff_html = self.build_requirement_diff_html(review)
        if diff_html:
            html_lines.append("<b>Requirements:</b><br>" + diff_html)
        html_lines.append("</body></html>")
        html_body = "\n".join(html_lines)
        msg.add_alternative(html_body, subtype="html")
        html_part = msg.get_payload()[1]
        for cid, data in zip(image_cids, images):
            html_part.add_related(data, "image", "png", cid=cid)

        # Attach FMEA tables as CSV files (can be opened with Excel)
        for name in review.fmea_names:
            fmea = next((f for f in self.fmeas if f["name"] == name), None)
            if not fmea:
                continue
            out = StringIO()
            writer = csv.writer(out)
            columns = [
                "Component",
                "Parent",
                "Failure Mode",
                "Failure Effect",
                "Cause",
                "S",
                "O",
                "D",
                "RPN",
                "Requirements",
            ]
            writer.writerow(columns)
            for be in fmea["entries"]:
                parent = be.parents[0] if be.parents else None
                if parent:
                    comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                    if parent.description:
                        comp = f"{comp} - {parent.description}"
                    parent_name = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                else:
                    comp = getattr(be, "fmea_component", "") or "N/A"
                    parent_name = ""
                req_ids = "; ".join(
                    [f"{req['req_type']}:{req['text']}" for req in getattr(be, 'safety_requirements', [])]
                )
                rpn = be.fmea_severity * be.fmea_occurrence * be.fmea_detection
                failure_mode = be.description or (be.user_name or f"BE {be.unique_id}")
                row = [
                    comp,
                    parent_name,
                    failure_mode,
                    be.fmea_effect,
                    getattr(be, "fmea_cause", ""),
                    be.fmea_severity,
                    be.fmea_occurrence,
                    be.fmea_detection,
                    rpn,
                    req_ids,
                ]
                writer.writerow(row)
            csv_bytes = out.getvalue().encode('utf-8')
            out.close()
            msg.add_attachment(
                csv_bytes,
                maintype="text",
                subtype="csv",
                filename=f"fmea_{name}.csv",
            )
        for name in getattr(review, 'fmeda_names', []):
            fmeda = next((f for f in self.fmedas if f["name"] == name), None)
            if not fmeda:
                continue
            out = StringIO()
            writer = csv.writer(out)
            columns = [
                "Component","Parent","Failure Mode","Malfunction","Safety Goal","FaultType","Fraction","FIT","DiagCov","Mechanism"
            ]
            writer.writerow(columns)
            for be in fmeda["entries"]:
                parent = be.parents[0] if be.parents else None
                if parent:
                    comp = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                    parent_name = parent.user_name if parent.user_name else f"Node {parent.unique_id}"
                else:
                    comp = getattr(be, "fmea_component", "") or "N/A"
                    parent_name = ""
                row = [
                    comp,
                    parent_name,
                    be.description or (be.user_name or f"BE {be.unique_id}"),
                    be.fmeda_malfunction,
                    be.fmeda_safety_goal,
                    be.fmeda_fault_type,
                    f"{be.fmeda_fault_fraction:.2f}",
                    f"{be.fmeda_fit:.2f}",
                    f"{be.fmeda_diag_cov:.2f}",
                    getattr(be, "fmeda_mechanism", ""),
                ]
                writer.writerow(row)
            csv_bytes = out.getvalue().encode('utf-8')
            out.close()
            msg.add_attachment(
                csv_bytes,
                maintype="text",
                subtype="csv",
                filename=f"fmeda_{name}.csv",
            )
        for name in getattr(review, 'hazop_names', []):
            doc = next((d for d in self.hazop_docs if d.name == name), None)
            if not doc:
                continue
            out = StringIO()
            writer = csv.writer(out)
            columns = [
                "Function",
                "Malfunction",
                "Type",
                "Scenario",
                "Conditions",
                "Hazard",
                "Safety",
                "Rationale",
                "Covered",
                "Covered By",
            ]
            writer.writerow(columns)
            for e in doc.entries:
                writer.writerow([
                    getattr(e, "function", e.get("function", "")),
                    getattr(e, "malfunction", e.get("malfunction", "")),
                    getattr(e, "mtype", e.get("mtype", "")),
                    getattr(e, "scenario", e.get("scenario", "")),
                    getattr(e, "conditions", e.get("conditions", "")),
                    getattr(e, "hazard", e.get("hazard", "")),
                    "Yes" if getattr(e, "safety", e.get("safety", False)) else "No",
                    getattr(e, "rationale", e.get("rationale", "")),
                    "Yes" if getattr(e, "covered", e.get("covered", False)) else "No",
                    getattr(e, "covered_by", e.get("covered_by", "")),
                ])
            csv_bytes = out.getvalue().encode("utf-8")
            out.close()
            msg.add_attachment(
                csv_bytes,
                maintype="text",
                subtype="csv",
                filename=f"hazop_{name}.csv",
            )
        for name in getattr(review, 'hara_names', []):
            doc = next((d for d in self.hara_docs if d.name == name), None)
            if not doc:
                continue
            out = StringIO()
            writer = csv.writer(out)
            columns = [
                "Malfunction",
                "Severity",
                "Severity Rationale",
                "Controllability",
                "Cont. Rationale",
                "Exposure",
                "Exp. Rationale",
                "ASIL",
                "Safety Goal",
            ]
            writer.writerow(columns)
            for e in doc.entries:
                writer.writerow([
                    getattr(e, "malfunction", e.get("malfunction", "")),
                    getattr(e, "severity", e.get("severity", "")),
                    getattr(e, "sev_rationale", e.get("sev_rationale", "")),
                    getattr(e, "controllability", e.get("controllability", "")),
                    getattr(e, "cont_rationale", e.get("cont_rationale", "")),
                    getattr(e, "exposure", e.get("exposure", "")),
                    getattr(e, "exp_rationale", e.get("exp_rationale", "")),
                    getattr(e, "asil", e.get("asil", "")),
                    getattr(e, "safety_goal", e.get("safety_goal", "")),
                ])
            csv_bytes = out.getvalue().encode("utf-8")
            out.close()
            msg.add_attachment(
                csv_bytes,
                maintype="text",
                subtype="csv",
                filename=f"hara_{name}.csv",
            )
        try:
            port = cfg.get('port', 465)
            if port == 465:
                with smtplib.SMTP_SSL(cfg['server'], port) as server:
                    server.login(cfg['email'], cfg['password'])
                    server.send_message(msg)
            else:
                with smtplib.SMTP(cfg['server'], port) as server:
                    server.starttls()
                    server.login(cfg['email'], cfg['password'])
                    server.send_message(msg)
        except smtplib.SMTPAuthenticationError:
            messagebox.showerror(
                "Email",
                "Login failed. If your account uses two-factor authentication, "
                "create an app password and use that instead of your normal password."
            )
            self.email_config = None
        except (socket.gaierror, ConnectionRefusedError, smtplib.SMTPConnectError) as e:
            messagebox.showerror(
                "Email",
                "Failed to connect to the SMTP server. Check the address, port and internet connection."
            )
            self.email_config = None
        except Exception as e:
            messagebox.showerror("Email", f"Failed to send review email: {e}")


    def review_is_closed(self):
        if not self.review_data:
            return False
        if getattr(self.review_data, "closed", False):
            return True
        if self.review_data.due_date:
            try:
                due = datetime.datetime.strptime(self.review_data.due_date, "%Y-%m-%d").date()
                if datetime.date.today() > due:
                    return True
            except ValueError:
                pass
        return False

    def review_is_closed_for(self, review):
        if not review:
            return False
        if getattr(review, "closed", False):
            return True
        if review.due_date:
            try:
                due = datetime.datetime.strptime(review.due_date, "%Y-%m-%d").date()
                if datetime.date.today() > due:
                    return True
            except ValueError:
                pass
        return False

    def get_requirements_for_review(self, review):
        """Return a set of requirement IDs included in the given review."""
        req_ids = set()
        for tid in getattr(review, "fta_ids", []):
            node = self.find_node_by_id_all(tid)
            if not node:
                continue
            for n in self.get_all_nodes(node):
                for r in getattr(n, "safety_requirements", []):
                    req_ids.add(r.get("id"))
        for name in getattr(review, "fmea_names", []):
            fmea = next((f for f in self.fmeas if f["name"] == name), None)
            if not fmea:
                continue
            for e in fmea.get("entries", []):
                for r in e.get("safety_requirements", []):
                    req_ids.add(r.get("id"))
        return req_ids

    def update_requirement_statuses(self):
        status_order = {
            "draft": 0,
            "in review": 1,
            "peer reviewed": 2,
            "pending approval": 3,
            "approved": 4,
        }
        for req in global_requirements.values():
            req.setdefault("status", "draft")
        for review in self.reviews:
            ids = self.get_requirements_for_review(review)
            closed = self.review_is_closed_for(review)
            for rid in ids:
                req = global_requirements.get(rid)
                if not req:
                    continue
                if review.mode == "joint":
                    if review.approved:
                        status = "approved"
                    elif closed:
                        status = "pending approval"
                    else:
                        status = "in review"
                else:  # peer review
                    if review.approved or closed:
                        status = "peer reviewed"
                    else:
                        status = "in review"
                if status_order[status] > status_order.get(req.get("status", "draft"), 0):
                    req["status"] = status


    def add_version(self):
        name = f"v{len(self.versions)+1}"
        # Exclude the versions list when capturing a snapshot to avoid
        # recursively embedding previous versions within each saved state.
        data = self.export_model_data(include_versions=False)
        self.versions.append({"name": name, "data": data})

    def compare_versions(self):
        if not self.versions:
            messagebox.showinfo("Versions", "No previous versions")
            return
        VersionCompareDialog(self.root, self)

    def merge_review_comments(self):
        path = filedialog.askopenfilename(defaultextension=".json", filetypes=[("JSON", "*.json")])
        if not path:
            return
        with open(path, "r") as f:
            data = json.load(f)

        for rd in data.get("reviews", []):
            participants = [ReviewParticipant(**p) for p in rd.get("participants", [])]
            comments = [ReviewComment(**c) for c in rd.get("comments", [])]
            moderators = [ReviewParticipant(**m) for m in rd.get("moderators", [])]
            if not moderators and rd.get("moderator"):
                moderators = [ReviewParticipant(rd.get("moderator"), "", "moderator")]
            review = next((r for r in self.reviews if r.name == rd.get("name", "")), None)
            if review is None:
                review = ReviewData(
                    name=rd.get("name", ""),
                    description=rd.get("description", ""),
                    mode=rd.get("mode", "peer"),
                    moderators=moderators,
                    participants=participants,
                    comments=comments,
                    approved=rd.get("approved", False),
                    fta_ids=rd.get("fta_ids", []),
                    fmea_names=rd.get("fmea_names", []),
                    fmeda_names=rd.get("fmeda_names", []),
                    due_date=rd.get("due_date", ""),
                    closed=rd.get("closed", False),
                )
                self.reviews.append(review)
                continue
            for p in participants:
                if all(p.name != ep.name for ep in review.participants):
                    review.participants.append(p)
            for m in moderators:
                if all(m.name != em.name for em in review.moderators):
                    review.moderators.append(m)
            review.due_date = rd.get("due_date", review.due_date)
            review.closed = rd.get("closed", review.closed)
            next_id = len(review.comments) + 1
            for c in comments:
                review.comments.append(ReviewComment(next_id, c.node_id, c.text, c.reviewer,
                                                     target_type=c.target_type, req_id=c.req_id,
                                                     field=c.field, resolved=c.resolved,
                                                     resolution=c.resolution))
                next_id += 1
        messagebox.showinfo("Merge", "Comments merged")

    def calculate_diff_nodes(self, old_data):
        old_map = self.node_map_from_data(old_data["top_events"])
        new_map = self.node_map_from_data([e.to_dict() for e in self.top_events])
        changed = []
        for nid, nd in new_map.items():
            if nid not in old_map:
                changed.append(nid)
            elif json.dumps(old_map[nid], sort_keys=True) != json.dumps(nd, sort_keys=True):
                changed.append(nid)
        return changed

    def calculate_diff_between(self, data1, data2):
        map1 = self.node_map_from_data(data1["top_events"])
        map2 = self.node_map_from_data(data2["top_events"])
        changed = []
        for nid, nd in map2.items():
            if nid not in map1 or json.dumps(map1.get(nid, {}), sort_keys=True) != json.dumps(nd, sort_keys=True):
                changed.append(nid)
        return changed

    def node_map_from_data(self, top_events):
        result = {}
        def visit(d):
            result[d["unique_id"]] = d
            for ch in d.get("children", []):
                visit(ch)
        for t in top_events:
            visit(t)
        return result

    def set_current_user(self):
        if not self.review_data:
            messagebox.showwarning("User", "Start a review first")
            return
        parts = self.review_data.participants + self.review_data.moderators
        dlg = UserSelectDialog(self.root, parts, initial_name=self.current_user)
        if not dlg.result:
            return
        name, _ = dlg.result
        allowed = [p.name for p in parts]
        if name not in allowed:
            messagebox.showerror("User", "Name not found in participants")
            return
        self.current_user = name

    def get_current_user_role(self):
        if not self.review_data:
            return None
        if self.current_user in [m.name for m in self.review_data.moderators]:
            return "moderator"
        for p in self.review_data.participants:
            if p.name == self.current_user:
                return p.role
        return None

    def focus_on_node(self, node):
        self.selected_node = node
        try:
            if hasattr(self, "canvas") and self.canvas.winfo_exists():
                self.redraw_canvas()
                bbox = self.canvas.bbox("all")
                if bbox:
                    self.canvas.xview_moveto(max(0, (node.x * self.zoom - self.canvas.winfo_width()/2) / bbox[2]))
                    self.canvas.yview_moveto(max(0, (node.y * self.zoom - self.canvas.winfo_height()/2) / bbox[3]))
        except tk.TclError:
            pass

    def get_review_targets(self):
        targets = []
        target_map = {}

        # Determine which FTAs and FMEAs are part of the current review.
        if self.review_data:
            allowed_ftas = set(self.review_data.fta_ids)
            allowed_fmeas = set(self.review_data.fmea_names)
            allowed_fmedas = set(getattr(self.review_data, 'fmeda_names', []))
        else:
            allowed_ftas = set()
            allowed_fmeas = set()
            allowed_fmedas = set()

        # Collect nodes from the selected FTAs (or all if none selected).
        nodes = []
        if allowed_ftas:
            for te in self.top_events:
                if te.unique_id in allowed_ftas:
                    nodes.extend(self.get_all_nodes(te))
        else:
            nodes = self.get_all_nodes_in_model()

        # Determine which nodes have FMEA entries in the selected FMEAs.
        fmea_node_ids = set()
        if allowed_fmeas or allowed_fmedas:
            for fmea in self.fmeas:
                if fmea["name"] in allowed_fmeas:
                    fmea_node_ids.update(be.unique_id for be in fmea["entries"])
            for d in self.fmedas:
                if d["name"] in allowed_fmedas:
                    fmea_node_ids.update(be.unique_id for be in d["entries"])
        else:
            # When no FMEA was selected, do not offer FMEA-related targets
            fmea_node_ids = set()

        for node in nodes:
            label = node.user_name or node.description or f"Node {node.unique_id}"
            targets.append(label)
            target_map[label] = ("node", node.unique_id)
            if hasattr(node, "safety_requirements"):
                for req in node.safety_requirements:
                    rlabel = f"{label} [Req {req.get('id')}]"
                    targets.append(rlabel)
                    target_map[rlabel] = ("requirement", node.unique_id, req.get("id"))

            if node.node_type.upper() == "BASIC EVENT" and node.unique_id in fmea_node_ids:
                flabel = f"{label} [FMEA]"
                targets.append(flabel)
                target_map[flabel] = ("fmea", node.unique_id)
                for field in ["Failure Mode", "Effect", "Cause", "Severity", "Occurrence", "Detection", "RPN"]:
                    slabel = f"{label} [FMEA {field}]"
                    key = field.lower().replace(' ', '_')
                    target_map[slabel] = ("fmea_field", node.unique_id, key)
                    targets.append(slabel)

        return targets, target_map

##########################################
# Node Model 
##########################################
class FaultTreeNode:
    def __init__(self, user_name, node_type, parent=None):
        self.unique_id = AD_RiskAssessment_Helper.get_next_unique_id()
        # Assign a sequential default name if none is provided
        self.user_name = user_name if user_name else f"Node {self.unique_id}"
        self.node_type = node_type
        self.children = []
        self.parents = []
        if parent is not None:
            self.parents.append(parent)
        self.quant_value = None
        self.gate_type = "AND" if node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else None
        self.description = ""
        self.rationale = ""
        self.x = 50
        self.y = 50
        # Severity and controllability now use a 1-3 scale
        # Default to the lowest level until linked to a HARA entry
        self.severity = 1 if node_type.upper() == "TOP EVENT" else None
        self.controllability = 1 if node_type.upper() == "TOP EVENT" else None
        self.input_subtype = None
        self.display_label = ""
        self.equation = ""
        self.detailed_equation = ""
        self.is_page = False
        self.is_primary_instance = True
        self.original = self
        self.safety_goal_description = ""
        self.safety_goal_asil = ""
        self.safe_state = ""
        self.ftti = ""
        self.acceptance_criteria = ""
        self.vehicle_safety_requirements = []          # List of vehicle safety requirements
        self.operational_safety_requirements = []        # List of operational safety requirements
        # Each requirement is a dict with keys: "id", "req_type" and "text"
        self.safety_requirements = []
        # --- FMEA attributes for basic events (AIAG style) ---
        self.fmea_effect = ""       # Description of effect/failure mode
        self.fmea_cause = ""        # Potential cause of failure
        self.fmea_severity = 1       # 1-10 scale
        self.fmea_occurrence = 1     # 1-10 scale
        self.fmea_detection = 1      # 1-10 scale
        self.fmea_component = ""     # Optional component name for FMEA-only nodes
        # --- FMEDA attributes ---
        self.fmeda_malfunction = ""
        self.fmeda_safety_goal = ""
        self.fmeda_diag_cov = 0.0
        self.fmeda_fit = 0.0
        self.fmeda_spfm = 0.0
        self.fmeda_lpfm = 0.0
        self.fmeda_fault_type = "permanent"
        self.fmeda_fault_fraction = 0.0
        # Reference to a unique failure mode this node represents
        self.failure_mode_ref = None
        # Probability values for classical FTA calculations
        self.failure_prob = 0.0
        self.probability = 0.0
        # Formula used to derive probability from FIT rate
        self.prob_formula = "linear"  # linear, exponential, or constant

    @property
    def name(self):
        orig = getattr(self, "original", self)
        uid = orig.unique_id if not self.is_primary_instance else self.unique_id
        base_name = self.user_name
        # Avoid repeating the ID if the user_name already matches the default
        if not base_name or base_name == f"Node {uid}":
            return f"Node {uid}"
        return f"Node {uid}: {base_name}"

    def to_dict(self):
        d = {
            "unique_id": self.unique_id,
            "user_name": self.user_name,
            "type": self.node_type,
            "quant_value": self.quant_value,
            "gate_type": self.gate_type,
            "description": self.description,
            "rationale": self.rationale,
            "x": self.x,
            "y": self.y,
            "severity": self.severity,
            "controllability": self.controllability,
            "input_subtype": self.input_subtype,
            "is_page": self.is_page,
            "is_primary_instance": self.is_primary_instance,
            "safety_goal_description": self.safety_goal_description,
            "safety_goal_asil": self.safety_goal_asil,
            "safe_state": self.safe_state,
            "ftti": self.ftti,
            "acceptance_criteria": self.acceptance_criteria,
            "fmea_effect": self.fmea_effect,
            "fmea_cause": self.fmea_cause,
            "fmea_severity": self.fmea_severity,
            "fmea_occurrence": self.fmea_occurrence,
            "fmea_detection": self.fmea_detection,
            "fmea_component": self.fmea_component,
            "fmeda_malfunction": self.fmeda_malfunction,
            "fmeda_safety_goal": self.fmeda_safety_goal,
            "fmeda_diag_cov": self.fmeda_diag_cov,
            "fmeda_fit": self.fmeda_fit,
            "fmeda_spfm": self.fmeda_spfm,
            "fmeda_lpfm": self.fmeda_lpfm,
            "fmeda_fault_type": self.fmeda_fault_type,
            "fmeda_fault_fraction": self.fmeda_fault_fraction,
            "failure_mode_ref": self.failure_mode_ref,
            # Save the safety requirements list (which now includes custom_id)
            "safety_requirements": self.safety_requirements,
            "failure_prob": self.failure_prob,
            "probability": self.probability,
            "prob_formula": self.prob_formula,
            "children": [child.to_dict() for child in self.children]
        }
        if not self.is_primary_instance and self.original and (self.original.unique_id != self.unique_id):
            d["original_id"] = self.original.unique_id
        return d

    @staticmethod
    def from_dict(data, parent=None):
        node = FaultTreeNode.__new__(FaultTreeNode)
        node.user_name = data.get("user_name", "")
        node.node_type = data.get("type", "")
        node.children = [FaultTreeNode.from_dict(child_data, parent=node) for child_data in data.get("children", [])]
        node.parents = []
        if parent is not None:
            node.parents.append(parent)
        node.quant_value = data.get("quant_value")
        node.gate_type = data.get("gate_type", "AND")
        node.description = data.get("description", "")
        node.rationale = data.get("rationale", "")
        node.x = data.get("x", 50)
        node.y = data.get("y", 50)
        node.severity = data.get("severity", 1) if node.node_type.upper() == "TOP EVENT" else None
        node.controllability = data.get("controllability", 1) if node.node_type.upper() == "TOP EVENT" else None
        node.input_subtype = data.get("input_subtype", None)
        node.is_page = boolify(data.get("is_page", False), False)
        node.is_primary_instance = boolify(data.get("is_primary_instance", True), True)
        node.safety_goal_description = data.get("safety_goal_description", "")
        node.safety_goal_asil = data.get("safety_goal_asil", "")
        node.safe_state = data.get("safe_state", "")
        node.ftti = data.get("ftti", "")
        node.acceptance_criteria = data.get("acceptance_criteria", "")
        node.fmea_effect = data.get("fmea_effect", "")
        node.fmea_cause = data.get("fmea_cause", "")
        node.fmea_severity = data.get("fmea_severity", 1)
        node.fmea_occurrence = data.get("fmea_occurrence", 1)
        node.fmea_detection = data.get("fmea_detection", 1)
        node.fmea_component = data.get("fmea_component", "")
        node.fmeda_malfunction = data.get("fmeda_malfunction", "")
        node.fmeda_safety_goal = data.get("fmeda_safety_goal", "")
        node.fmeda_diag_cov = data.get("fmeda_diag_cov", 0.0)
        node.fmeda_fit = data.get("fmeda_fit", 0.0)
        node.fmeda_spfm = data.get("fmeda_spfm", 0.0)
        node.fmeda_lpfm = data.get("fmeda_lpfm", 0.0)
        node.fmeda_fault_type = data.get("fmeda_fault_type", "permanent")
        node.fmeda_fault_fraction = data.get("fmeda_fault_fraction", 0.0)
        node.failure_mode_ref = data.get("failure_mode_ref")
        # NEW: Load safety_requirements (or default to empty list)
        node.safety_requirements = data.get("safety_requirements", [])
        node.failure_prob = data.get("failure_prob", 0.0)
        node.probability = data.get("probability", 0.0)
        node.prob_formula = data.get("prob_formula", "linear")
        node.display_label = ""
        node.equation = ""
        node.detailed_equation = ""
        if "unique_id" in data:
            node.unique_id = data["unique_id"]
        else:
            node.unique_id = AD_RiskAssessment_Helper.get_next_unique_id()
        if not node.is_primary_instance and "original_id" in data:
            node._original_id = data["original_id"]
        else:
            node._original_id = None
        return node
        
##########################################
# Page Diagram 
##########################################
class PageDiagram:
    def __init__(self, app, page_gate_node, canvas):
        self.app = app
        self.root_node = page_gate_node
        self.canvas = canvas
        self.zoom = 1.0
        self.diagram_font = tkFont.Font(family="Arial", size=int(8 * self.zoom))
        self.grid_size = 20
        self.selected_node = None
        self.dragging_node = None
        self.drag_offset_x = 0
        self.drag_offset_y = 0
        self.rc_dragged = False
        # Reference project properties for grid and color options
        self.project_properties = app.project_properties

        # Bind events – including right-click release
        self.canvas.bind("<ButtonPress-3>", self.on_right_mouse_press)
        self.canvas.bind("<B3-Motion>", self.on_right_mouse_drag)
        self.canvas.bind("<ButtonRelease-3>", self.on_right_mouse_release)
        self.canvas.bind("<Button-1>", self.on_canvas_click)
        self.canvas.bind("<B1-Motion>", self.on_canvas_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_canvas_release)
        self.canvas.bind("<Double-Button-1>", self.on_canvas_double_click)
        self.canvas.bind("<Control-MouseWheel>", self.on_ctrl_mousewheel)

    def on_right_mouse_press(self, event):
        self.rc_dragged = False
        self.canvas.scan_mark(event.x, event.y)

    def on_right_mouse_drag(self, event):
        self.rc_dragged = True
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def on_right_mouse_release(self, event):
        # If there was no significant drag, show the context menu.
        if not self.rc_dragged:
            self.show_context_menu(event)

    def find_node_at_position(self, x, y):
        # Adjust the radius (here using 45 as an example)
        radius_sq = (45 * self.zoom) ** 2
        for n in self.get_all_nodes(self.root_node):
            if (x - n.x) ** 2 + (y - n.y) ** 2 < radius_sq:
                return n
        return None
        
    def on_ctrl_mousewheel(self, event):
        if event.delta > 0:
            self.zoom_in()
        else:
            self.zoom_out()
            
    def get_all_nodes(self, node=None):
        if node is None:
            node = self.root_node
        visited = set()

        def rec(n):
            if n.unique_id in visited:
                return []
            visited.add(n.unique_id)

            # Skip nodes if a parent is page, but that page is NOT our root_node
            if n != self.root_node and any(p.is_page and p != self.root_node for p in n.parents):
                return []

            result = [n]
            for c in n.children:
                result.extend(rec(c))
            return result

        return rec(node)

    def rc_on_press(self, event):
        self.rc_start = (event.x, event.y)
        self.rc_dragged = False
        self.canvas.scan_mark(event.x, event.y)

    def rc_on_motion(self, event):
        self.rc_dragged = True
        self.canvas.scan_dragto(event.x, event.y, gain=1)

    def rc_on_release(self, event):
        if not self.rc_dragged:
            self.show_context_menu(event)

    def show_context_menu(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        node = None
        for n in self.get_all_nodes(self.root_node):
            radius = 60 if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else 45
            if (x - n.x)**2 + (y - n.y)**2 < radius**2:
                node = n
                break
        if not node:
            return
        self.selected_node = node
        self.app.selected_node = node
        menu = tk.Menu(self.app.root, tearoff=0)
        menu.add_command(label="Edit", command=lambda: self.context_edit(node))
        menu.add_command(label="Remove Connection", command=lambda: self.context_remove(node))
        menu.add_command(label="Delete Node", command=lambda: self.context_delete(node))
        menu.add_command(label="Copy", command=lambda: self.context_copy(node))
        menu.add_command(label="Cut", command=lambda: self.context_cut(node))
        menu.add_command(label="Paste", command=lambda: self.context_paste(node))
        if node.node_type.upper() not in ["TOP EVENT", "BASIC EVENT"]:
            menu.add_command(label="Edit Page Flag", command=lambda: self.context_edit_page_flag(node))
        menu.add_separator()
        menu.add_command(label="Add Confidence", command=lambda: self.context_add("Confidence Level"))
        menu.add_command(label="Add Robustness", command=lambda: self.context_add("Robustness Score"))
        menu.add_command(label="Add Gate", command=lambda: self.context_add("GATE"))
        menu.add_command(label="Add Basic Event", command=lambda: self.context_add("Basic Event"))
        menu.add_command(label="Add Triggering Condition", command=lambda: self.context_add("Triggering Condition"))
        menu.add_command(label="Add Functional Insufficiency", command=lambda: self.context_add("Functional Insufficiency"))
        menu.tk_popup(event.x_root, event.y_root)

    def context_edit(self, node):
        EditNodeDialog(self.canvas, node, self.app)
        self.redraw_canvas()
        self.app.update_views()

    def context_remove(self, node):
        self.selected_node = node
        self.app.remove_connection(node)
        self.redraw_canvas()
        self.app.update_views()

    def context_delete(self, node):
        self.selected_node = node
        self.app.delete_node_and_subtree(node)
        self.redraw_canvas()
        self.app.update_views()

    def context_copy(self, node):
        self.selected_node = node
        self.app.copy_node()

    def context_cut(self, node):
        self.selected_node = node
        self.app.cut_node()

    def context_paste(self, node):
        self.selected_node = node
        self.app.paste_node()

    def context_edit_page_flag(self, node):
        self.selected_node = node
        self.app.edit_page_flag()
        self.redraw_canvas()

    def context_add(self, event_type):
        self.app.selected_node = self.selected_node
        self.app.add_node_of_type(event_type)
        self.redraw_canvas()
        self.app.update_views()

    def on_canvas_click(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        clicked_node = None
        for n in self.get_all_nodes(self.root_node):
            radius = 60 if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else 45
            if (x - n.x)**2 + (y - n.y)**2 < radius**2:
                clicked_node = n
                break
        self.selected_node = clicked_node
        self.app.selected_node = clicked_node
        if clicked_node and clicked_node is not self.root_node:
            self.dragging_node = clicked_node
            self.drag_offset_x = x - clicked_node.x
            self.drag_offset_y = y - clicked_node.y
        else:
            self.dragging_node = None
        self.redraw_canvas()
        
    def on_canvas_double_click(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        clicked_node = self.find_node_at_position(x, y)
        if clicked_node:
            if not clicked_node.is_primary_instance:
                self.app.open_page_diagram(getattr(clicked_node, "original", clicked_node))
            else:
                if clicked_node.is_page:
                    self.app.open_page_diagram(clicked_node)
                else:
                    EditNodeDialog(self.app.root, clicked_node, self.app)
            self.app.update_views()

    def on_canvas_drag(self, event):
        if self.dragging_node:
            x = self.canvas.canvasx(event.x) / self.zoom
            y = self.canvas.canvasy(event.y) / self.zoom
            new_x = x - self.drag_offset_x
            new_y = y - self.drag_offset_y
            dx = new_x - self.dragging_node.x
            dy = new_y - self.dragging_node.y
            self.dragging_node.x = new_x
            self.dragging_node.y = new_y
            if self.dragging_node.is_primary_instance:
                self.app.move_subtree(self.dragging_node, dx, dy)
            self.app.sync_nodes_by_id(self.dragging_node)
            self.redraw_canvas()

    def on_canvas_release(self, event):
        if self.dragging_node:
            self.dragging_node.x = round(self.dragging_node.x/self.grid_size)*self.grid_size
            self.dragging_node.y = round(self.dragging_node.y/self.grid_size)*self.grid_size
        self.dragging_node = None
        self.drag_offset_x = 0
        self.drag_offset_y = 0

    def on_canvas_double_click(self, event):
        x = self.canvas.canvasx(event.x) / self.zoom
        y = self.canvas.canvasy(event.y) / self.zoom
        clicked_node = None
        for n in self.get_all_nodes(self.root_node):
            radius = 60 if n.node_type.upper() in ["GATE", "RIGOR LEVEL", "TOP EVENT"] else 45
            if (x - n.x)**2 + (y - n.y)**2 < radius**2:
                clicked_node = n
                break
        if clicked_node:
            if not clicked_node.is_primary_instance:
                self.app.open_page_diagram(getattr(clicked_node, "original", clicked_node))
            else:
                if clicked_node.is_page:
                    self.app.open_page_diagram(clicked_node)
                else:
                    EditNodeDialog(self.app.root, clicked_node, self.app)
            self.app.update_views()

    def zoom_in(self):
        self.zoom *= 1.2
        self.diagram_font.config(size=int(8 * self.zoom))
        self.redraw_canvas()

    def zoom_out(self):
        self.zoom /= 1.2
        self.diagram_font.config(size=int(8 * self.zoom))
        self.redraw_canvas()

    def auto_arrange(self):
        horizontal_gap = 150
        vertical_gap = 100
        next_y = [100]
        def layout(node, depth):
            node.x = depth * horizontal_gap + 100
            if not node.children:
                node.y = next_y[0]
                next_y[0] += vertical_gap
            else:
                for child in node.children:
                    layout(child, depth+1)
                node.y = (node.children[0].y + node.children[-1].y) / 2
        layout(self.root_node, 0)
        self.redraw_canvas()

    def redraw_canvas(self):
        # Clear the canvas and draw the grid first.
        if not hasattr(self, "canvas") or not self.canvas.winfo_exists():
            return
        self.canvas.delete("all")
        self.draw_grid()
        
        # Use the page's root node as the sole top-level event.
        drawn_ids = set()
        for top_event in [self.root_node]:
            self.draw_connections(top_event, drawn_ids)
        
        all_nodes = []
        for top_event in [self.root_node]:
            all_nodes.extend(self.get_all_nodes(top_event))
        for node in all_nodes:
            self.draw_node(node)
        
        # Update the scroll region.
        self.canvas.config(scrollregion=self.canvas.bbox("all"))

    def draw_grid(self):
        if not self.project_properties.get("show_grid", True):
            return
        spacing = self.grid_size * self.zoom
        width = self.canvas.winfo_width()
        height = self.canvas.winfo_height()
        if width < 10: 
            width = 800
        if height < 10: 
            height = 600
        for x in range(0, int(width), int(spacing)):
            self.canvas.create_line(x, 0, x, height, fill="#ddd", tags="grid")
        for y in range(0, int(height), int(spacing)):
            self.canvas.create_line(0, y, width, y, fill="#ddd", tags="grid")

    def draw_connections(self, node, drawn_ids=set()):
        if id(node) in drawn_ids:
            return
        drawn_ids.add(id(node))
        if node.is_page and node.is_primary_instance and node != self.root_node:
            return
        if node.children:
            region_width = 100 * self.zoom
            parent_bottom = (node.x * self.zoom, node.y * self.zoom + 40 * self.zoom)
            N = len(node.children)
            for i, child in enumerate(node.children):
                parent_conn = (node.x * self.zoom - region_width/2 + (i+0.5)*(region_width/N), parent_bottom[1])
                child_top = (child.x * self.zoom, child.y * self.zoom - 45 * self.zoom)
                fta_drawing_helper.draw_90_connection(self.canvas, parent_conn, child_top, outline_color="dimgray", line_width=1)
            for child in node.children:
                self.draw_connections(child, drawn_ids)

    def draw_node(self, node):
        """
        Draws the given node on the main canvas.
        For clones, it always uses the original’s non-positional attributes (like display_label,
        description, etc.) so that any changes to the original are reflected on all clones.
        """
        # If the node is a clone, use its original for configuration (non-positional attributes)
        source = node if node.is_primary_instance else node.original

        # For display purposes, show the clone marker on the clone's display_label.
        if node.is_primary_instance:
            display_label = source.display_label
        else:
            display_label = source.display_label + " (clone)"

        # Build a short top_text string from the source's attributes.
        subtype_text = source.input_subtype if source.input_subtype else "N/A"
        top_text = (
            f"Type: {source.node_type}\n"
            f"Subtype: {subtype_text}\n"
            f"{display_label}\n"
            f"Desc: {source.description}\n\n"
            f"Rationale: {source.rationale}"
        )
        # For the bottom text, you may choose to display the node's name (which for a clone is
        # usually the same as the original’s name)
        bottom_text = source.name

        # Compute the effective position using the clone’s own (positional) values
        eff_x = node.x * self.zoom
        eff_y = node.y * self.zoom

        # Highlight if selected
        outline_color = "red" if node == self.selected_node else "dimgray"
        line_width = 2 if node == self.selected_node else 1

        # Determine the fill color (this function already uses the original's display_label)
        fill_color = self.app.get_node_fill_color(node)
        font_obj = self.diagram_font

        # For shape selection, use the source’s node type and gate type.
        node_type_upper = source.node_type.upper()

        if not node.is_primary_instance:
            # For clones, draw them in a “clone” style.
            if source.is_page:
                fta_drawing_helper.draw_triangle_clone_shape(self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                                                             top_text=top_text,
                                                             bottom_text=bottom_text,
                                                             fill=fill_color,
                                                             outline_color=outline_color,
                                                             line_width=1,
                                                             font_obj=self.diagram_font)
            elif node_type_upper in ["GATE", "RIGOR LEVEL", "TOP EVENT"]:
                if source.gate_type.upper() == "OR":
                    fta_drawing_helper.draw_rotated_or_gate_clone_shape(
                        self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                        top_text=top_text, bottom_text=bottom_text,
                        fill=fill_color, outline_color=outline_color,
                        line_width=line_width, font_obj=font_obj
                    )
                else:
                    fta_drawing_helper.draw_rotated_and_gate_clone_shape(
                        self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                        top_text=top_text, bottom_text=bottom_text,
                        fill=fill_color, outline_color=outline_color,
                        line_width=line_width, font_obj=font_obj
                    )
            elif node_type_upper in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
            else:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
        else:
            # Primary node: use normal drawing routines.
            if node_type_upper in ["TOP EVENT", "GATE", "RIGOR LEVEL"]:
                if source.is_page and source != self.root_node:
                    fta_drawing_helper.draw_triangle_shape(
                        self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                        top_text=top_text, bottom_text=bottom_text,
                        fill=fill_color, outline_color=outline_color,
                        line_width=line_width, font_obj=font_obj
                    )
                else:
                    if source.gate_type.upper() == "OR":
                        fta_drawing_helper.draw_rotated_or_gate_shape(
                            self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                            top_text=top_text, bottom_text=bottom_text,
                            fill=fill_color, outline_color=outline_color,
                            line_width=line_width, font_obj=font_obj
                        )
                    else:
                        fta_drawing_helper.draw_rotated_and_gate_shape(
                            self.canvas, eff_x, eff_y, scale=40 * self.zoom,
                            top_text=top_text, bottom_text=bottom_text,
                            fill=fill_color, outline_color=outline_color,
                            line_width=line_width, font_obj=font_obj
                        )
            elif node_type_upper in ["CONFIDENCE LEVEL", "ROBUSTNESS SCORE"]:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )
            else:
                fta_drawing_helper.draw_circle_event_shape(
                    self.canvas, eff_x, eff_y, 45 * self.zoom,
                    top_text=top_text, bottom_text=bottom_text,
                    fill=fill_color, outline_color=outline_color,
                    line_width=line_width, font_obj=font_obj
                )

        # Draw any additional text (such as equations) from the source.
        if source.equation:
            self.canvas.create_text(
                eff_x - 80 * self.zoom, eff_y - 15 * self.zoom,
                text=source.equation, anchor="e", fill="gray",
                font=self.diagram_font
            )
        if source.detailed_equation:
            self.canvas.create_text(
                eff_x - 80 * self.zoom, eff_y + 15 * self.zoom,
                text=source.detailed_equation, anchor="e", fill="gray",
                font=self.diagram_font
            )

        # Finally, if the node appears multiple times, draw a shared marker.
        if self.app.occurrence_counts.get(node.unique_id, 0) > 1:
            marker_x = eff_x + 30 * self.zoom
            marker_y = eff_y - 30 * self.zoom
            fta_drawing_helper.draw_shared_marker(self.canvas, marker_x, marker_y, self.zoom)

def main():
    root = tk.Tk()
    # Create a fresh helper each session:
    global AD_RiskAssessment_Helper
    AD_RiskAssessment_Helper = ADRiskAssessmentHelper()
    
    app = FaultTreeApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
